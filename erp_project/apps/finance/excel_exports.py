"""
Excel Export Utilities for Finance Reports
Uses openpyxl for Excel generation.
"""
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from decimal import Decimal
from datetime import date, datetime


def create_excel_response(filename):
    """Create HttpResponse for Excel file download."""
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def style_header_row(ws, row_num, col_count):
    """Apply header styling to a row."""
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align


def style_title_row(ws, row_num, title, col_count):
    """Add and style a title row."""
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=col_count)
    cell = ws.cell(row=row_num, column=1, value=title)
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal='center')


def auto_width_columns(ws):
    """Auto-adjust column widths based on content."""
    from openpyxl.cell.cell import MergedCell
    
    for column_cells in ws.columns:
        max_length = 0
        column = None
        for cell in column_cells:
            # Skip merged cells
            if isinstance(cell, MergedCell):
                continue
            # Get column letter from first non-merged cell
            if column is None:
                column = cell.column_letter
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        if column:
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width


def format_currency(value):
    """Format decimal as currency string."""
    if value is None:
        return ''
    if isinstance(value, (int, float, Decimal)):
        return float(value)
    return value


# ============ TRIAL BALANCE EXPORT (Standard - As at Date) ============

def export_trial_balance(accounts, as_of_date, company_name=''):
    """
    Export Standard Trial Balance to Excel.
    Shows only net balances as Debit OR Credit per account.
    IFRS & UAE Audit Compliant.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Trial Balance'
    
    # Title
    style_title_row(ws, 1, f'Trial Balance (As at {as_of_date})', 5)
    if company_name:
        ws.cell(row=2, column=1, value=company_name)
    
    # Headers
    headers = ['Account Code', 'Account Name', 'Type', 'Debit (AED)', 'Credit (AED)']
    header_row = 4 if company_name else 3
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    row = header_row + 1
    total_debit = Decimal('0.00')
    total_credit = Decimal('0.00')
    
    for acc in accounts:
        ws.cell(row=row, column=1, value=acc.get('code', ''))
        ws.cell(row=row, column=2, value=acc.get('name', ''))
        ws.cell(row=row, column=3, value=acc.get('account_type', ''))
        
        debit = acc.get('debit', Decimal('0.00'))
        credit = acc.get('credit', Decimal('0.00'))
        
        ws.cell(row=row, column=4, value=format_currency(debit) if debit else '')
        ws.cell(row=row, column=5, value=format_currency(credit) if credit else '')
        
        total_debit += debit or Decimal('0.00')
        total_credit += credit or Decimal('0.00')
        
        # Highlight abnormal balances
        if acc.get('abnormal', False):
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color='FFF3CD', end_color='FFF3CD', fill_type='solid'
                )
        
        row += 1
    
    # Totals
    ws.cell(row=row, column=1, value='TOTAL')
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=4, value=format_currency(total_debit))
    ws.cell(row=row, column=4).font = Font(bold=True)
    ws.cell(row=row, column=5, value=format_currency(total_credit))
    ws.cell(row=row, column=5).font = Font(bold=True)
    
    # Add border to totals row
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = Border(top=Side(style='double'))
    
    # Balance check
    row += 2
    is_balanced = total_debit == total_credit
    if is_balanced:
        ws.cell(row=row, column=1, value='✓ Trial Balance is BALANCED')
        ws.cell(row=row, column=1).font = Font(bold=True, color='008000')
    else:
        ws.cell(row=row, column=1, value=f'✗ UNBALANCED - Difference: {format_currency(total_debit - total_credit)}')
        ws.cell(row=row, column=1).font = Font(bold=True, color='FF0000')
    
    auto_width_columns(ws)
    
    response = create_excel_response(f'trial_balance_as_at_{as_of_date}.xlsx')
    wb.save(response)
    return response


# ============ TRIAL BALANCE WITH MOVEMENTS EXPORT ============

def export_trial_balance_with_movements(accounts, start_date, end_date, totals=None, company_name=''):
    """
    Export Trial Balance with Movements to Excel.
    Shows Opening Balance, Period Movement, and Closing Balance.
    IFRS & UAE Audit Compliant.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'TB with Movements'
    
    # Title
    style_title_row(ws, 1, f'Trial Balance with Movements', 9)
    ws.cell(row=2, column=1, value=f'Period: {start_date} to {end_date}')
    if company_name:
        ws.cell(row=3, column=1, value=company_name)
    
    # Headers Row 1 - Groups
    ws.cell(row=5, column=1, value='Account Code')
    ws.cell(row=5, column=2, value='Account Name')
    ws.cell(row=5, column=3, value='Type')
    ws.merge_cells('D5:E5')
    ws.cell(row=5, column=4, value='Opening Balance')
    ws.cell(row=5, column=4).alignment = Alignment(horizontal='center')
    ws.merge_cells('F5:G5')
    ws.cell(row=5, column=6, value='Period Movement')
    ws.cell(row=5, column=6).alignment = Alignment(horizontal='center')
    ws.merge_cells('H5:I5')
    ws.cell(row=5, column=8, value='Closing Balance')
    ws.cell(row=5, column=8).alignment = Alignment(horizontal='center')
    
    # Headers Row 2 - Columns
    headers = ['Account Code', 'Account Name', 'Type', 
               'Opening Dr', 'Opening Cr', 
               'Period Dr', 'Period Cr', 
               'Closing Dr', 'Closing Cr']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        cell.border = Border(
            bottom=Side(style='thin'),
            top=Side(style='thin')
        )
    
    # Data
    row = 7
    for acc in accounts:
        ws.cell(row=row, column=1, value=acc.get('code', ''))
        ws.cell(row=row, column=2, value=acc.get('name', ''))
        ws.cell(row=row, column=3, value=acc.get('account_type', ''))
        
        # Opening
        ws.cell(row=row, column=4, value=format_currency(acc.get('opening_debit', Decimal('0.00'))))
        ws.cell(row=row, column=5, value=format_currency(acc.get('opening_credit', Decimal('0.00'))))
        
        # Period
        ws.cell(row=row, column=6, value=format_currency(acc.get('period_debit', Decimal('0.00'))))
        ws.cell(row=row, column=7, value=format_currency(acc.get('period_credit', Decimal('0.00'))))
        
        # Closing
        ws.cell(row=row, column=8, value=format_currency(acc.get('closing_debit', Decimal('0.00'))))
        ws.cell(row=row, column=9, value=format_currency(acc.get('closing_credit', Decimal('0.00'))))
        
        # Highlight abnormal balances
        if acc.get('abnormal', False):
            for col in range(1, 10):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color='FFF3CD', end_color='FFF3CD', fill_type='solid'
                )
        
        row += 1
    
    # Totals
    if totals:
        ws.cell(row=row, column=1, value='TOTAL')
        ws.cell(row=row, column=1).font = Font(bold=True)
        
        ws.cell(row=row, column=4, value=format_currency(totals.get('total_opening_debit', Decimal('0.00'))))
        ws.cell(row=row, column=4).font = Font(bold=True)
        ws.cell(row=row, column=5, value=format_currency(totals.get('total_opening_credit', Decimal('0.00'))))
        ws.cell(row=row, column=5).font = Font(bold=True)
        
        ws.cell(row=row, column=6, value=format_currency(totals.get('total_period_debit', Decimal('0.00'))))
        ws.cell(row=row, column=6).font = Font(bold=True)
        ws.cell(row=row, column=7, value=format_currency(totals.get('total_period_credit', Decimal('0.00'))))
        ws.cell(row=row, column=7).font = Font(bold=True)
        
        ws.cell(row=row, column=8, value=format_currency(totals.get('total_closing_debit', Decimal('0.00'))))
        ws.cell(row=row, column=8).font = Font(bold=True)
        ws.cell(row=row, column=9, value=format_currency(totals.get('total_closing_credit', Decimal('0.00'))))
        ws.cell(row=row, column=9).font = Font(bold=True)
        
        # Add border to totals row
        for col in range(1, 10):
            ws.cell(row=row, column=col).border = Border(top=Side(style='double'))
    
    auto_width_columns(ws)
    
    response = create_excel_response(f'trial_balance_movements_{start_date}_to_{end_date}.xlsx')
    wb.save(response)
    return response


# ============ PROFIT & LOSS EXPORT ============

def export_profit_loss(revenue_accounts, expense_accounts, start_date, end_date, company_name=''):
    """Export Profit & Loss to Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Profit & Loss'
    
    # Title
    style_title_row(ws, 1, f'Profit & Loss Statement', 3)
    ws.cell(row=2, column=1, value=f'Period: {start_date} to {end_date}')
    if company_name:
        ws.cell(row=3, column=1, value=company_name)
    
    row = 5
    
    # Revenue Section
    ws.cell(row=row, column=1, value='REVENUE')
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    
    total_revenue = Decimal('0.00')
    for acc in revenue_accounts:
        ws.cell(row=row, column=1, value=acc.get('code', ''))
        ws.cell(row=row, column=2, value=acc.get('name', ''))
        amount = acc.get('balance', acc.get('amount', Decimal('0.00')))
        ws.cell(row=row, column=3, value=format_currency(abs(amount) if amount else 0))
        total_revenue += abs(amount) if amount else Decimal('0.00')
        row += 1
    
    ws.cell(row=row, column=2, value='Total Revenue')
    ws.cell(row=row, column=2).font = Font(bold=True)
    ws.cell(row=row, column=3, value=format_currency(total_revenue))
    ws.cell(row=row, column=3).font = Font(bold=True)
    row += 2
    
    # Expense Section
    ws.cell(row=row, column=1, value='EXPENSES')
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    
    total_expenses = Decimal('0.00')
    for acc in expense_accounts:
        ws.cell(row=row, column=1, value=acc.get('code', ''))
        ws.cell(row=row, column=2, value=acc.get('name', ''))
        amount = acc.get('balance', acc.get('amount', Decimal('0.00')))
        ws.cell(row=row, column=3, value=format_currency(abs(amount) if amount else 0))
        total_expenses += abs(amount) if amount else Decimal('0.00')
        row += 1
    
    ws.cell(row=row, column=2, value='Total Expenses')
    ws.cell(row=row, column=2).font = Font(bold=True)
    ws.cell(row=row, column=3, value=format_currency(total_expenses))
    ws.cell(row=row, column=3).font = Font(bold=True)
    row += 2
    
    # Net Profit/Loss
    net = total_revenue - total_expenses
    ws.cell(row=row, column=2, value='NET PROFIT / (LOSS)')
    ws.cell(row=row, column=2).font = Font(bold=True, size=12)
    ws.cell(row=row, column=3, value=format_currency(net))
    ws.cell(row=row, column=3).font = Font(bold=True, size=12)
    
    auto_width_columns(ws)
    
    response = create_excel_response(f'profit_loss_{start_date}_to_{end_date}.xlsx')
    wb.save(response)
    return response


# ============ BALANCE SHEET EXPORT ============

def export_balance_sheet(assets, liabilities, equity, end_date, start_date=None, company_name=''):
    """Export Balance Sheet to Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Balance Sheet'
    
    # Title - show period if start_date provided
    if start_date:
        title = f'Balance Sheet ({start_date} to {end_date})'
    else:
        title = f'Balance Sheet as of {end_date}'
    style_title_row(ws, 1, title, 3)
    if company_name:
        ws.cell(row=2, column=1, value=company_name)
    
    row = 4
    
    # Assets
    ws.cell(row=row, column=1, value='ASSETS')
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    
    total_assets = Decimal('0.00')
    for acc in assets:
        ws.cell(row=row, column=1, value=acc.get('code', ''))
        ws.cell(row=row, column=2, value=acc.get('name', ''))
        amount = acc.get('balance', Decimal('0.00'))
        ws.cell(row=row, column=3, value=format_currency(amount))
        total_assets += amount or Decimal('0.00')
        row += 1
    
    ws.cell(row=row, column=2, value='Total Assets')
    ws.cell(row=row, column=2).font = Font(bold=True)
    ws.cell(row=row, column=3, value=format_currency(total_assets))
    ws.cell(row=row, column=3).font = Font(bold=True)
    row += 2
    
    # Liabilities
    ws.cell(row=row, column=1, value='LIABILITIES')
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    
    total_liabilities = Decimal('0.00')
    for acc in liabilities:
        ws.cell(row=row, column=1, value=acc.get('code', ''))
        ws.cell(row=row, column=2, value=acc.get('name', ''))
        amount = acc.get('balance', Decimal('0.00'))
        ws.cell(row=row, column=3, value=format_currency(abs(amount) if amount else 0))
        total_liabilities += abs(amount) if amount else Decimal('0.00')
        row += 1
    
    ws.cell(row=row, column=2, value='Total Liabilities')
    ws.cell(row=row, column=2).font = Font(bold=True)
    ws.cell(row=row, column=3, value=format_currency(total_liabilities))
    ws.cell(row=row, column=3).font = Font(bold=True)
    row += 2
    
    # Equity
    ws.cell(row=row, column=1, value='EQUITY')
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    
    total_equity = Decimal('0.00')
    for acc in equity:
        ws.cell(row=row, column=1, value=acc.get('code', ''))
        ws.cell(row=row, column=2, value=acc.get('name', ''))
        amount = acc.get('balance', Decimal('0.00'))
        ws.cell(row=row, column=3, value=format_currency(abs(amount) if amount else 0))
        total_equity += abs(amount) if amount else Decimal('0.00')
        row += 1
    
    ws.cell(row=row, column=2, value='Total Equity')
    ws.cell(row=row, column=2).font = Font(bold=True)
    ws.cell(row=row, column=3, value=format_currency(total_equity))
    ws.cell(row=row, column=3).font = Font(bold=True)
    row += 2
    
    # Total L + E
    ws.cell(row=row, column=2, value='Total Liabilities & Equity')
    ws.cell(row=row, column=2).font = Font(bold=True, size=12)
    ws.cell(row=row, column=3, value=format_currency(total_liabilities + total_equity))
    ws.cell(row=row, column=3).font = Font(bold=True, size=12)
    
    auto_width_columns(ws)
    
    response = create_excel_response(f'balance_sheet_{end_date}.xlsx')
    wb.save(response)
    return response


# ============ GENERAL LEDGER EXPORT ============

def export_general_ledger(transactions, account_name, start_date, end_date,
                          opening_balance=None, period_debit=None,
                          period_credit=None, closing_balance=None):
    """Export General Ledger to Excel with opening balance, period totals, and closing."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'General Ledger'

    style_title_row(ws, 1, f'General Ledger - {account_name}', 7)
    ws.cell(row=2, column=1, value=f'Period: {start_date} to {end_date}')

    headers = ['Date', 'Entry #', 'Reference', 'Description', 'Debit', 'Credit', 'Balance']
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=header)
    style_header_row(ws, 4, len(headers))

    row = 5
    opening_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    bold = Font(bold=True)

    # Opening balance row
    if opening_balance is not None:
        ws.cell(row=row, column=1, value='')
        ws.cell(row=row, column=4, value='Opening Balance').font = bold
        ws.cell(row=row, column=7, value=format_currency(opening_balance)).font = bold
        for c in range(1, 8):
            ws.cell(row=row, column=c).fill = opening_fill
        row += 1

    for txn in transactions:
        d = txn.get('date', '')
        ws.cell(row=row, column=1, value=d.strftime('%d/%m/%Y') if hasattr(d, 'strftime') else str(d))
        ws.cell(row=row, column=2, value=txn.get('entry_number', ''))
        ws.cell(row=row, column=3, value=txn.get('reference', ''))
        ws.cell(row=row, column=4, value=txn.get('description', ''))
        ws.cell(row=row, column=5, value=format_currency(txn.get('debit', 0)))
        ws.cell(row=row, column=6, value=format_currency(txn.get('credit', 0)))
        ws.cell(row=row, column=7, value=format_currency(txn.get('balance', 0)))
        row += 1

    # Period totals row
    totals_fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
    if period_debit is not None:
        ws.cell(row=row, column=4, value='Period Totals').font = bold
        ws.cell(row=row, column=5, value=format_currency(period_debit)).font = bold
        ws.cell(row=row, column=6, value=format_currency(period_credit)).font = bold
        for c in range(1, 8):
            ws.cell(row=row, column=c).fill = totals_fill
        row += 1

    # Closing balance row
    closing_fill = PatternFill(start_color='263238', end_color='263238', fill_type='solid')
    closing_font = Font(bold=True, color='FFFFFF')
    if closing_balance is not None:
        ws.cell(row=row, column=4, value='Closing Balance').font = closing_font
        ws.cell(row=row, column=7, value=format_currency(closing_balance)).font = closing_font
        for c in range(1, 8):
            ws.cell(row=row, column=c).fill = closing_fill

    auto_width_columns(ws)

    response = create_excel_response(f'general_ledger_{start_date}_to_{end_date}.xlsx')
    wb.save(response)
    return response


# ============ JOURNAL REGISTER EXPORT ============

def export_journal_register(entries, start_date, end_date):
    """Export Journal Register to Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Journal Register'
    
    # Title
    style_title_row(ws, 1, f'Journal Register', 8)
    ws.cell(row=2, column=1, value=f'Period: {start_date} to {end_date}')
    
    # Headers
    headers = ['Entry #', 'Date', 'Reference', 'Description', 'Source', 'Debit', 'Credit', 'Status']
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=header)
    style_header_row(ws, 4, len(headers))
    
    # Data
    row = 5
    for entry in entries:
        ws.cell(row=row, column=1, value=entry.entry_number)
        ws.cell(row=row, column=2, value=entry.date.strftime('%Y-%m-%d') if entry.date else '')
        ws.cell(row=row, column=3, value=entry.reference)
        ws.cell(row=row, column=4, value=entry.description)
        ws.cell(row=row, column=5, value=entry.get_source_module_display() if hasattr(entry, 'get_source_module_display') else entry.source_module)
        ws.cell(row=row, column=6, value=format_currency(entry.total_debit))
        ws.cell(row=row, column=7, value=format_currency(entry.total_credit))
        ws.cell(row=row, column=8, value=entry.get_status_display() if hasattr(entry, 'get_status_display') else entry.status)
        row += 1
    
    auto_width_columns(ws)
    
    response = create_excel_response(f'journal_register_{start_date}_to_{end_date}.xlsx')
    wb.save(response)
    return response


# ============ AR AGING EXPORT ============

def export_ar_aging(customers, as_of_date):
    """Export AR Aging to Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'AR Aging'
    
    # Title
    style_title_row(ws, 1, f'Accounts Receivable Aging as of {as_of_date}', 7)
    
    # Headers
    headers = ['Customer', 'Current', '1-30 Days', '31-60 Days', '61-90 Days', 'Over 90 Days', 'Total']
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
    style_header_row(ws, 3, len(headers))
    
    # Data
    row = 4
    totals = {'current': 0, '1_30': 0, '31_60': 0, '61_90': 0, 'over_90': 0, 'total': 0}
    
    for cust in customers:
        ws.cell(row=row, column=1, value=cust.get('name', ''))
        ws.cell(row=row, column=2, value=format_currency(cust.get('current', 0)))
        ws.cell(row=row, column=3, value=format_currency(cust.get('1_30', cust.get('days_1_30', 0))))
        ws.cell(row=row, column=4, value=format_currency(cust.get('31_60', cust.get('days_31_60', 0))))
        ws.cell(row=row, column=5, value=format_currency(cust.get('61_90', cust.get('days_61_90', 0))))
        ws.cell(row=row, column=6, value=format_currency(cust.get('over_90', cust.get('days_over_90', 0))))
        ws.cell(row=row, column=7, value=format_currency(cust.get('total', 0)))
        
        totals['current'] += cust.get('current', 0) or 0
        totals['1_30'] += cust.get('1_30', cust.get('days_1_30', 0)) or 0
        totals['31_60'] += cust.get('31_60', cust.get('days_31_60', 0)) or 0
        totals['61_90'] += cust.get('61_90', cust.get('days_61_90', 0)) or 0
        totals['over_90'] += cust.get('over_90', cust.get('days_over_90', 0)) or 0
        totals['total'] += cust.get('total', 0) or 0
        row += 1
    
    # Totals row
    ws.cell(row=row, column=1, value='TOTAL')
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2, value=format_currency(totals['current']))
    ws.cell(row=row, column=3, value=format_currency(totals['1_30']))
    ws.cell(row=row, column=4, value=format_currency(totals['31_60']))
    ws.cell(row=row, column=5, value=format_currency(totals['61_90']))
    ws.cell(row=row, column=6, value=format_currency(totals['over_90']))
    ws.cell(row=row, column=7, value=format_currency(totals['total']))
    for col in range(1, 8):
        ws.cell(row=row, column=col).font = Font(bold=True)
    
    auto_width_columns(ws)
    
    response = create_excel_response(f'ar_aging_{as_of_date}.xlsx')
    wb.save(response)
    return response


# ============ AP AGING EXPORT ============

def export_ap_aging(vendors, as_of_date):
    """Export AP Aging to Excel with GL reconciliation."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'AP Aging'

    style_title_row(ws, 1, f'Accounts Payable Aging as of {as_of_date}', 7)

    headers = ['Reference', 'Current', '1-30 Days', '31-60 Days', '61-90 Days', 'Over 90 Days', 'Total']
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
    style_header_row(ws, 3, len(headers))

    row = 4
    totals = {'current': 0, '1_30': 0, '31_60': 0, '61_90': 0, 'over_90': 0, 'total': 0}

    for vendor in vendors:
        ws.cell(row=row, column=1, value=vendor.get('name', ''))
        ws.cell(row=row, column=2, value=format_currency(vendor.get('current', 0)))
        ws.cell(row=row, column=3, value=format_currency(vendor.get('1_30', vendor.get('days_1_30', 0))))
        ws.cell(row=row, column=4, value=format_currency(vendor.get('31_60', vendor.get('days_31_60', 0))))
        ws.cell(row=row, column=5, value=format_currency(vendor.get('61_90', vendor.get('days_61_90', 0))))
        ws.cell(row=row, column=6, value=format_currency(vendor.get('over_90', vendor.get('days_over_90', 0))))
        ws.cell(row=row, column=7, value=format_currency(vendor.get('total', 0)))

        totals['current'] += vendor.get('current', 0) or 0
        totals['1_30'] += vendor.get('1_30', vendor.get('days_1_30', 0)) or 0
        totals['31_60'] += vendor.get('31_60', vendor.get('days_31_60', 0)) or 0
        totals['61_90'] += vendor.get('61_90', vendor.get('days_61_90', 0)) or 0
        totals['over_90'] += vendor.get('over_90', vendor.get('days_over_90', 0)) or 0
        totals['total'] += vendor.get('total', 0) or 0
        row += 1

    ws.cell(row=row, column=1, value='TOTAL')
    ws.cell(row=row, column=1).font = Font(bold=True)
    for col in range(2, 8):
        ws.cell(row=row, column=col).font = Font(bold=True)
    ws.cell(row=row, column=2, value=format_currency(totals['current']))
    ws.cell(row=row, column=3, value=format_currency(totals['1_30']))
    ws.cell(row=row, column=4, value=format_currency(totals['31_60']))
    ws.cell(row=row, column=5, value=format_currency(totals['61_90']))
    ws.cell(row=row, column=6, value=format_currency(totals['over_90']))
    ws.cell(row=row, column=7, value=format_currency(totals['total']))

    auto_width_columns(ws)

    # --- Sheet 2: GL Reconciliation ---
    ws2 = wb.create_sheet('GL Reconciliation')
    style_title_row(ws2, 1, 'AP GL Reconciliation', 4)
    ws2.cell(row=2, column=1, value=f'As of: {as_of_date}')

    from apps.finance.models import Account, JournalEntryLine
    from django.db.models import Sum
    from django.db.models.functions import Coalesce
    from decimal import Decimal

    ap_account = Account.objects.filter(code='2000', is_active=True).first()
    if ap_account:
        agg = JournalEntryLine.objects.filter(
            account=ap_account, journal_entry__status='posted',
        ).aggregate(
            d=Coalesce(Sum('debit'), Decimal('0')),
            c=Coalesce(Sum('credit'), Decimal('0')),
        )
        gl_bal = agg['c'] - agg['d']
        report_total = Decimal(str(totals['total']))

        r = 4
        rec_headers = ['Description', 'Amount (AED)', 'Status']
        for col, h in enumerate(rec_headers, 1):
            c = ws2.cell(row=r, column=col, value=h)
            c.font = Font(bold=True)
            c.fill = PatternFill(start_color='DDDDDD', fill_type='solid')
        r += 1

        ok_fill = PatternFill(start_color='C6EFCE', fill_type='solid')
        err_fill = PatternFill(start_color='FFC7CE', fill_type='solid')

        ws2.cell(row=r, column=1, value='AP GL Balance (2000)')
        ws2.cell(row=r, column=2, value=format_currency(gl_bal))
        r += 1
        ws2.cell(row=r, column=1, value='AP Aging Report Total')
        ws2.cell(row=r, column=2, value=format_currency(report_total))
        r += 1

        diff = abs(gl_bal - report_total)
        match = diff < Decimal('0.01')
        ws2.cell(row=r, column=1, value='Difference')
        ws2.cell(row=r, column=1).font = Font(bold=True)
        ws2.cell(row=r, column=2, value=format_currency(diff))
        status_cell = ws2.cell(row=r, column=3, value='RECONCILED' if match else 'MISMATCH')
        status_cell.fill = ok_fill if match else err_fill
        status_cell.font = Font(bold=True)

        r += 2
        ws2.cell(row=r, column=1, value='Unpaid Vendor Bills')
        ws2.cell(row=r, column=1).font = Font(bold=True)
        r += 1

        from apps.purchase.models import VendorBill
        for b in VendorBill.objects.filter(status__in=['posted', 'partial']).order_by('bill_date'):
            out = b.total_amount - b.paid_amount
            if out > 0:
                ws2.cell(row=r, column=1, value=f'{b.bill_number} (due: {b.due_date})')
                ws2.cell(row=r, column=2, value=format_currency(out))
                overdue = (date.today() - b.due_date).days if b.due_date else 0
                if overdue > 0:
                    ws2.cell(row=r, column=3, value=f'OVERDUE {overdue} days')
                    ws2.cell(row=r, column=3).font = Font(color='FF0000')
                r += 1

    auto_width_columns(ws2)

    response = create_excel_response(f'ap_aging_{as_of_date}.xlsx')
    wb.save(response)
    return response


# ============ VAT REPORT EXPORT ============

def export_vat_report(data, start_date, end_date):
    """Export VAT Report to Excel with GL reconciliation."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'VAT Report'

    is_submitted = data.get('is_submitted', False)
    source_label = 'Filed VAT Return' if is_submitted else 'Calculated from GL'
    return_number = data.get('vat_return_number', '')
    return_status = data.get('vat_return_status', 'draft')

    style_title_row(ws, 1, 'UAE VAT Return Report (FTA Format)', 4)
    ws.cell(row=2, column=1, value=f'Period: {start_date} to {end_date}')
    ws.cell(row=3, column=1, value=f'Data Source: {source_label}')
    if return_number:
        ws.cell(row=3, column=3, value=f'Return #: {return_number}  Status: {return_status}')

    row = 5

    header_fill = PatternFill(start_color='E2EFDA', fill_type='solid')
    for col, h in enumerate(['Description', 'Taxable Amount (AED)', 'VAT Amount (AED)'], 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(bold=True)
        c.fill = header_fill
    row += 1

    ws.cell(row=row, column=1, value='OUTPUT VAT (Sales)')
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1

    ws.cell(row=row, column=1, value='Standard Rated Supplies (Box 1)')
    ws.cell(row=row, column=2, value=format_currency(data.get('standard_sales', 0)))
    ws.cell(row=row, column=3, value=format_currency(data.get('output_vat', 0)))
    row += 1

    ws.cell(row=row, column=1, value='Zero Rated Supplies (Box 3)')
    ws.cell(row=row, column=2, value=format_currency(data.get('zero_rated_sales', 0)))
    ws.cell(row=row, column=3, value=format_currency(0))
    row += 1

    ws.cell(row=row, column=1, value='Exempt Supplies (Box 4)')
    ws.cell(row=row, column=2, value=format_currency(data.get('exempt_sales', 0)))
    ws.cell(row=row, column=3, value=format_currency(0))
    row += 1

    out_of_scope = data.get('out_of_scope_sales', 0)
    if out_of_scope:
        ws.cell(row=row, column=1, value='Out of Scope Supplies')
        ws.cell(row=row, column=2, value=format_currency(out_of_scope))
        row += 1

    row += 1
    ws.cell(row=row, column=1, value='INPUT VAT (Purchases)')
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1

    ws.cell(row=row, column=1, value='Standard Rated Expenses (Box 9)')
    ws.cell(row=row, column=2, value=format_currency(data.get('standard_purchases', 0)))
    ws.cell(row=row, column=3, value=format_currency(data.get('input_vat', 0)))
    row += 2

    adjustments = data.get('adjustments', 0) or 0
    if adjustments:
        ws.cell(row=row, column=1, value='Adjustments')
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=3, value=format_currency(adjustments))
        row += 1

    net_vat = data.get('net_vat', 0) or ((data.get('output_vat', 0) or 0) - (data.get('input_vat', 0) or 0))
    ws.cell(row=row, column=1, value='NET VAT PAYABLE / (REFUNDABLE)')
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    ws.cell(row=row, column=3, value=format_currency(net_vat))
    ws.cell(row=row, column=3).font = Font(bold=True, size=12)

    auto_width_columns(ws)

    # --- Sheet 2: GL Reconciliation ---
    ws2 = wb.create_sheet('GL Reconciliation')
    style_title_row(ws2, 1, 'VAT GL Reconciliation', 5)
    ws2.cell(row=2, column=1, value=f'Period: {start_date} to {end_date}')

    r = 4
    rec_headers = ['Account', 'GL Balance', 'Report Amount', 'Difference', 'Status']
    for col, h in enumerate(rec_headers, 1):
        c = ws2.cell(row=r, column=col, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill(start_color='DDDDDD', fill_type='solid')
    r += 1

    from apps.finance.models import Account, JournalEntryLine
    from django.db.models import Sum, Q
    from django.db.models.functions import Coalesce
    from decimal import Decimal

    vat_out_accs = Account.objects.filter(code='2110', is_active=True)
    vat_in_accs = Account.objects.filter(code='1300', is_active=True)
    vat_net_accs = Account.objects.filter(code='2120', is_active=True)

    def _gl_bal(accounts, normal='credit'):
        agg = JournalEntryLine.objects.filter(
            account__in=accounts, journal_entry__status='posted',
        ).exclude(
            journal_entry__reference__startswith='TEST-CF-'
        ).aggregate(
            d=Coalesce(Sum('debit'), Decimal('0')),
            c=Coalesce(Sum('credit'), Decimal('0')),
        )
        return agg['c'] - agg['d'] if normal == 'credit' else agg['d'] - agg['c']

    output_gl = _gl_bal(vat_out_accs, 'credit')
    input_gl = _gl_bal(vat_in_accs, 'debit')
    net_gl = _gl_bal(vat_net_accs, 'credit')

    report_output = data.get('output_vat', 0) or Decimal('0')
    report_input = data.get('input_vat', 0) or Decimal('0')
    report_net = net_vat if net_vat else Decimal('0')

    ok_fill = PatternFill(start_color='C6EFCE', fill_type='solid')
    err_fill = PatternFill(start_color='FFC7CE', fill_type='solid')

    for label, gl_val, rpt_val in [
        ('2110 VAT Output', output_gl, report_output),
        ('1300 VAT Input', input_gl, report_input),
        ('2120 VAT Net Payable', net_gl, report_net),
    ]:
        diff = abs(Decimal(str(gl_val)) - Decimal(str(rpt_val)))
        match = diff < Decimal('0.01')
        ws2.cell(row=r, column=1, value=label)
        ws2.cell(row=r, column=2, value=format_currency(gl_val))
        ws2.cell(row=r, column=3, value=format_currency(rpt_val))
        ws2.cell(row=r, column=4, value=format_currency(diff))
        status_cell = ws2.cell(row=r, column=5, value='MATCH' if match else 'MISMATCH')
        status_cell.fill = ok_fill if match else err_fill
        r += 1

    r += 2
    ws2.cell(row=r, column=1, value='Settlement Status')
    ws2.cell(row=r, column=1).font = Font(bold=True)
    r += 1
    settled_dr = JournalEntryLine.objects.filter(
        account__in=vat_net_accs, journal_entry__status='posted', debit__gt=0
    ).aggregate(d=Coalesce(Sum('debit'), Decimal('0')))['d']
    remaining = net_gl - settled_dr if net_gl > settled_dr else Decimal('0')
    ws2.cell(row=r, column=1, value='Total Liability')
    ws2.cell(row=r, column=2, value=format_currency(net_gl))
    r += 1
    ws2.cell(row=r, column=1, value='Amount Settled')
    ws2.cell(row=r, column=2, value=format_currency(settled_dr))
    r += 1
    ws2.cell(row=r, column=1, value='Outstanding')
    ws2.cell(row=r, column=2, value=format_currency(remaining))
    ws2.cell(row=r, column=2).font = Font(bold=True, color='FF0000' if remaining > 0 else '008000')

    auto_width_columns(ws2)

    response = create_excel_response(f'vat_report_{start_date}_to_{end_date}.xlsx')
    wb.save(response)
    return response


# ============ BUDGET VS ACTUAL EXPORT ============

def export_budget_vs_actual(data, budget_name, period):
    """Export Budget vs Actual to Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Budget vs Actual'
    
    # Title
    style_title_row(ws, 1, f'Budget vs Actual Report', 5)
    ws.cell(row=2, column=1, value=f'Budget: {budget_name}')
    ws.cell(row=3, column=1, value=f'Period: {period}')
    
    # Headers
    headers = ['Account', 'Budget', 'Actual', 'Variance', 'Variance %']
    for col, header in enumerate(headers, 1):
        ws.cell(row=5, column=col, value=header)
    style_header_row(ws, 5, len(headers))
    
    # Data
    row = 6
    for item in data:
        ws.cell(row=row, column=1, value=item.get('account', ''))
        ws.cell(row=row, column=2, value=format_currency(item.get('budget', 0)))
        ws.cell(row=row, column=3, value=format_currency(item.get('actual', 0)))
        ws.cell(row=row, column=4, value=format_currency(item.get('variance', 0)))
        ws.cell(row=row, column=5, value=f"{item.get('variance_pct', 0):.1f}%")
        row += 1
    
    auto_width_columns(ws)
    
    response = create_excel_response(f'budget_vs_actual_{period}.xlsx')
    wb.save(response)
    return response


# ============ BANK LEDGER EXPORT ============

def export_bank_ledger(transactions, bank_name, start_date, end_date):
    """Export Bank Ledger to Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Bank Ledger'
    
    # Title
    style_title_row(ws, 1, f'Bank Ledger - {bank_name}', 6)
    ws.cell(row=2, column=1, value=f'Period: {start_date} to {end_date}')
    
    # Headers
    headers = ['Date', 'Reference', 'Description', 'Debit', 'Credit', 'Balance']
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=header)
    style_header_row(ws, 4, len(headers))
    
    # Data
    row = 5
    for txn in transactions:
        ws.cell(row=row, column=1, value=txn.get('date', ''))
        ws.cell(row=row, column=2, value=txn.get('reference', ''))
        ws.cell(row=row, column=3, value=txn.get('description', ''))
        ws.cell(row=row, column=4, value=format_currency(txn.get('debit', 0)))
        ws.cell(row=row, column=5, value=format_currency(txn.get('credit', 0)))
        ws.cell(row=row, column=6, value=format_currency(txn.get('balance', 0)))
        row += 1
    
    auto_width_columns(ws)
    
    response = create_excel_response(f'bank_ledger_{start_date}_to_{end_date}.xlsx')
    wb.save(response)
    return response


# ============ CASH FLOW EXPORT ============

def export_cash_flow(operating, investing, financing, start_date, end_date, 
                     opening_balance=Decimal('0.00'), closing_balance=Decimal('0.00'),
                     opening_detail=None, closing_detail=None, company_name='',
                     excluded_adjustments=Decimal('0.00')):
    """
    Export Cash Flow Statement to Excel - IFRS/GAAP Compliant.
    
    Includes:
    - Beginning and Ending Cash Balance (Reconciliation)
    - Detailed line items for each activity
    - Net Change validation
    - Supplemental disclosures section
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Cash Flow Statement'
    
    # Styles
    header_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=11)
    total_font = Font(bold=True)
    total_fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
    reconciliation_fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
    
    # Title
    ws.merge_cells('A1:C1')
    ws.cell(row=1, column=1, value='CASH FLOW STATEMENT')
    ws.cell(row=1, column=1).font = header_font
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    
    if company_name:
        ws.merge_cells('A2:C2')
        ws.cell(row=2, column=1, value=company_name)
        ws.cell(row=2, column=1).alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A3:C3')
    ws.cell(row=3, column=1, value=f'For the Period: {start_date} to {end_date}')
    ws.cell(row=3, column=1).alignment = Alignment(horizontal='center')
    
    ws.cell(row=4, column=1, value='(Direct Method - IFRS Compliant)')
    ws.cell(row=4, column=1).font = Font(italic=True)
    
    row = 6
    
    # ========================================
    # BEGINNING CASH BALANCE
    # ========================================
    ws.cell(row=row, column=1, value='BEGINNING CASH BALANCE')
    ws.cell(row=row, column=1).font = total_font
    ws.cell(row=row, column=2, value=float(opening_balance))
    ws.cell(row=row, column=2).number_format = '#,##0.00'
    ws.cell(row=row, column=2).font = total_font
    for col in range(1, 3):
        ws.cell(row=row, column=col).fill = reconciliation_fill
    row += 2
    
    # ========================================
    # OPERATING ACTIVITIES
    # ========================================
    ws.cell(row=row, column=1, value='CASH FLOWS FROM OPERATING ACTIVITIES')
    ws.cell(row=row, column=1).font = section_font
    row += 1
    
    total_operating = Decimal('0.00')
    if operating:
        for item in operating:
            desc = item.get('description', 'Other operating activity')
            amount = Decimal(str(item.get('amount', 0))) if item.get('amount') else Decimal('0.00')
            
            # Indent line items
            ws.cell(row=row, column=1, value=f'  {desc}')
            ws.cell(row=row, column=2, value=float(amount))
            ws.cell(row=row, column=2).number_format = '#,##0.00'
            total_operating += amount
            row += 1
    else:
        ws.cell(row=row, column=1, value='  No operating cash flows')
        ws.cell(row=row, column=1).font = Font(italic=True)
        row += 1
    
    # Operating Total
    ws.cell(row=row, column=1, value='Net Cash from Operating Activities')
    ws.cell(row=row, column=1).font = total_font
    ws.cell(row=row, column=2, value=float(total_operating))
    ws.cell(row=row, column=2).number_format = '#,##0.00'
    ws.cell(row=row, column=2).font = total_font
    for col in range(1, 3):
        ws.cell(row=row, column=col).fill = total_fill
    row += 2
    
    # ========================================
    # INVESTING ACTIVITIES
    # ========================================
    ws.cell(row=row, column=1, value='CASH FLOWS FROM INVESTING ACTIVITIES')
    ws.cell(row=row, column=1).font = section_font
    row += 1
    
    total_investing = Decimal('0.00')
    if investing:
        for item in investing:
            desc = item.get('description', 'Other investing activity')
            amount = Decimal(str(item.get('amount', 0))) if item.get('amount') else Decimal('0.00')
            
            ws.cell(row=row, column=1, value=f'  {desc}')
            ws.cell(row=row, column=2, value=float(amount))
            ws.cell(row=row, column=2).number_format = '#,##0.00'
            total_investing += amount
            row += 1
    else:
        # Show explicit zero line items for professional presentation (GAAP compliant)
        zero_investing_items = [
            'Purchase of property & equipment',
            'Proceeds from sale of assets',
            'Purchase of investments',
        ]
        for item_desc in zero_investing_items:
            ws.cell(row=row, column=1, value=f'  {item_desc}')
            ws.cell(row=row, column=2, value=0.00)
            ws.cell(row=row, column=2).number_format = '#,##0.00'
            ws.cell(row=row, column=1).font = Font(color='808080')  # Gray for no activity
            ws.cell(row=row, column=2).font = Font(color='808080')
            row += 1
    
    # Investing Total
    ws.cell(row=row, column=1, value='Net Cash from Investing Activities')
    ws.cell(row=row, column=1).font = total_font
    ws.cell(row=row, column=2, value=float(total_investing))
    ws.cell(row=row, column=2).number_format = '#,##0.00'
    ws.cell(row=row, column=2).font = total_font
    for col in range(1, 3):
        ws.cell(row=row, column=col).fill = total_fill
    row += 2
    
    # ========================================
    # FINANCING ACTIVITIES
    # ========================================
    ws.cell(row=row, column=1, value='CASH FLOWS FROM FINANCING ACTIVITIES')
    ws.cell(row=row, column=1).font = section_font
    row += 1
    
    total_financing = Decimal('0.00')
    if financing:
        for item in financing:
            desc = item.get('description', 'Other financing activity')
            amount = Decimal(str(item.get('amount', 0))) if item.get('amount') else Decimal('0.00')
            
            ws.cell(row=row, column=1, value=f'  {desc}')
            ws.cell(row=row, column=2, value=float(amount))
            ws.cell(row=row, column=2).number_format = '#,##0.00'
            total_financing += amount
            row += 1
    else:
        # Show explicit zero line items for professional presentation (GAAP compliant)
        zero_financing_items = [
            'Proceeds from borrowings',
            'Repayment of debt principal',
            'Owner contributions',
            'Owner drawings/dividends',
        ]
        for item_desc in zero_financing_items:
            ws.cell(row=row, column=1, value=f'  {item_desc}')
            ws.cell(row=row, column=2, value=0.00)
            ws.cell(row=row, column=2).number_format = '#,##0.00'
            ws.cell(row=row, column=1).font = Font(color='808080')  # Gray for no activity
            ws.cell(row=row, column=2).font = Font(color='808080')
            row += 1
    
    # Financing Total
    ws.cell(row=row, column=1, value='Net Cash from Financing Activities')
    ws.cell(row=row, column=1).font = total_font
    ws.cell(row=row, column=2, value=float(total_financing))
    ws.cell(row=row, column=2).number_format = '#,##0.00'
    ws.cell(row=row, column=2).font = total_font
    for col in range(1, 3):
        ws.cell(row=row, column=col).fill = total_fill
    row += 2
    
    # ========================================
    # ADJUSTMENTS (Reversals, corrections excluded from activities)
    # ========================================
    if excluded_adjustments != Decimal('0.00'):
        ws.cell(row=row, column=1, value='ADJUSTMENTS (Reversals/Corrections)')
        ws.cell(row=row, column=1).font = section_font
        row += 1
        
        ws.cell(row=row, column=1, value='  Accounting adjustments affecting cash')
        ws.cell(row=row, column=2, value=float(excluded_adjustments))
        ws.cell(row=row, column=2).number_format = '#,##0.00'
        row += 1
        
        ws.cell(row=row, column=1, value='Net Adjustments')
        ws.cell(row=row, column=1).font = total_font
        ws.cell(row=row, column=2, value=float(excluded_adjustments))
        ws.cell(row=row, column=2).number_format = '#,##0.00'
        ws.cell(row=row, column=2).font = total_font
        for col in range(1, 3):
            ws.cell(row=row, column=col).fill = total_fill
        row += 2
    
    # ========================================
    # NET CHANGE IN CASH (including adjustments)
    # ========================================
    net_change = total_operating + total_investing + total_financing
    total_net_change = net_change + excluded_adjustments  # Include adjustments for reconciliation
    
    ws.cell(row=row, column=1, value='NET INCREASE (DECREASE) IN CASH')
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    ws.cell(row=row, column=2, value=float(total_net_change))
    ws.cell(row=row, column=2).number_format = '#,##0.00'
    ws.cell(row=row, column=2).font = Font(bold=True, size=12)
    row += 2
    
    # ========================================
    # ENDING CASH BALANCE
    # ========================================
    ws.cell(row=row, column=1, value='ENDING CASH BALANCE')
    ws.cell(row=row, column=1).font = total_font
    ws.cell(row=row, column=2, value=float(closing_balance))
    ws.cell(row=row, column=2).number_format = '#,##0.00'
    ws.cell(row=row, column=2).font = total_font
    for col in range(1, 3):
        ws.cell(row=row, column=col).fill = reconciliation_fill
    row += 2
    
    # ========================================
    # RECONCILIATION CHECK
    # ========================================
    ws.cell(row=row, column=1, value='RECONCILIATION:')
    ws.cell(row=row, column=1).font = section_font
    row += 1
    
    expected_closing = opening_balance + total_net_change
    ws.cell(row=row, column=1, value=f'  Beginning Balance + Net Change')
    ws.cell(row=row, column=2, value=float(expected_closing))
    ws.cell(row=row, column=2).number_format = '#,##0.00'
    row += 1
    
    ws.cell(row=row, column=1, value=f'  Ending Cash Balance')
    ws.cell(row=row, column=2, value=float(closing_balance))
    ws.cell(row=row, column=2).number_format = '#,##0.00'
    row += 1
    
    variance = closing_balance - expected_closing
    reconciles = abs(variance) < Decimal('0.01')
    ws.cell(row=row, column=1, value=f'  Variance')
    ws.cell(row=row, column=2, value=float(variance))
    ws.cell(row=row, column=2).number_format = '#,##0.00'
    
    if reconciles:
        ws.cell(row=row, column=3, value='✓ RECONCILES')
        ws.cell(row=row, column=3).font = Font(bold=True, color='006400')
    else:
        ws.cell(row=row, column=3, value='✗ DOES NOT RECONCILE')
        ws.cell(row=row, column=3).font = Font(bold=True, color='FF0000')
    row += 2
    
    # ========================================
    # CASH ACCOUNT BREAKDOWN (if provided)
    # ========================================
    if opening_detail or closing_detail:
        ws.cell(row=row, column=1, value='CASH & BANK ACCOUNT BREAKDOWN')
        ws.cell(row=row, column=1).font = section_font
        row += 1
        
        # Header
        ws.cell(row=row, column=1, value='Account')
        ws.cell(row=row, column=2, value='Opening')
        ws.cell(row=row, column=3, value='Closing')
        for col in range(1, 4):
            ws.cell(row=row, column=col).font = total_font
        row += 1
        
        # Combine opening and closing details
        all_accounts = set()
        opening_dict = {d['account']: d['balance'] for d in (opening_detail or [])}
        closing_dict = {d['account']: d['balance'] for d in (closing_detail or [])}
        all_accounts.update(opening_dict.keys())
        all_accounts.update(closing_dict.keys())
        
        for acc in sorted(all_accounts):
            ws.cell(row=row, column=1, value=f'  {acc}')
            ws.cell(row=row, column=2, value=float(opening_dict.get(acc, 0)))
            ws.cell(row=row, column=2).number_format = '#,##0.00'
            ws.cell(row=row, column=3, value=float(closing_dict.get(acc, 0)))
            ws.cell(row=row, column=3).number_format = '#,##0.00'
            row += 1
        row += 1
    
    # ========================================
    # NOTES
    # ========================================
    ws.cell(row=row, column=1, value='NOTES:')
    ws.cell(row=row, column=1).font = section_font
    row += 1
    ws.cell(row=row, column=1, value='1. This statement is prepared using the Direct Method (IFRS/IAS 7 compliant)')
    row += 1
    ws.cell(row=row, column=1, value='2. Cash includes cash in hand and bank deposits')
    row += 1
    ws.cell(row=row, column=1, value='3. Opening balance entries are excluded from cash flow activities (they are positions, not transactions)')
    row += 1
    ws.cell(row=row, column=1, value=f'4. Report generated: {start_date} to {end_date}')
    
    # Set column widths
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    
    response = create_excel_response(f'cash_flow_statement_{start_date}_to_{end_date}.xlsx')
    wb.save(response)
    return response


# ============ CORPORATE TAX EXPORT ============

def export_corporate_tax(data):
    """Export Corporate Tax Report to Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Corporate Tax'
    
    # Title
    style_title_row(ws, 1, f'UAE Corporate Tax Computation', 3)
    ws.cell(row=2, column=1, value=f'Fiscal Year: {data.get("fiscal_year", "")}')
    ws.cell(row=3, column=1, value=f'Period: {data.get("start_date", "")} to {data.get("end_date", "")}')
    
    row = 5
    
    # Revenue & Expenses
    ws.cell(row=row, column=1, value='INCOME STATEMENT SUMMARY')
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    
    ws.cell(row=row, column=1, value='Total Revenue')
    ws.cell(row=row, column=2, value=format_currency(data.get('revenue', 0)))
    row += 1
    
    ws.cell(row=row, column=1, value='Total Expenses')
    ws.cell(row=row, column=2, value=format_currency(data.get('expenses', 0)))
    row += 1
    
    ws.cell(row=row, column=1, value='Accounting Profit')
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2, value=format_currency(data.get('accounting_profit', 0)))
    ws.cell(row=row, column=2).font = Font(bold=True)
    row += 2
    
    # Tax Computation
    ws.cell(row=row, column=1, value='TAX COMPUTATION')
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    
    ws.cell(row=row, column=1, value='Tax-Free Threshold')
    ws.cell(row=row, column=2, value=format_currency(data.get('tax_threshold', 375000)))
    row += 1
    
    ws.cell(row=row, column=1, value='Tax Rate')
    ws.cell(row=row, column=2, value=f'{data.get("tax_rate", 9)}%')
    row += 1
    
    ws.cell(row=row, column=1, value='Taxable Amount (Profit - Threshold)')
    ws.cell(row=row, column=2, value=format_currency(data.get('taxable_amount', 0)))
    row += 1
    
    ws.cell(row=row, column=1, value='TAX PAYABLE')
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    ws.cell(row=row, column=2, value=format_currency(data.get('tax_payable', 0)))
    ws.cell(row=row, column=2).font = Font(bold=True, size=12)
    row += 2
    
    # If there's a saved computation with adjustments
    computation = data.get('computation')
    if computation:
        ws.cell(row=row, column=1, value='SAVED COMPUTATION DETAILS')
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        
        ws.cell(row=row, column=1, value='Non-Deductible Expenses')
        ws.cell(row=row, column=2, value=format_currency(computation.non_deductible_expenses))
        row += 1
        
        ws.cell(row=row, column=1, value='Exempt Income')
        ws.cell(row=row, column=2, value=format_currency(computation.exempt_income))
        row += 1
        
        ws.cell(row=row, column=1, value='Other Adjustments')
        ws.cell(row=row, column=2, value=format_currency(computation.other_adjustments))
        row += 1
        
        ws.cell(row=row, column=1, value='Adjusted Taxable Income')
        ws.cell(row=row, column=2, value=format_currency(computation.taxable_income))
        row += 1
        
        ws.cell(row=row, column=1, value='Final Tax Payable')
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        ws.cell(row=row, column=2, value=format_currency(computation.tax_payable))
        ws.cell(row=row, column=2).font = Font(bold=True, size=12)
        row += 1
        
        ws.cell(row=row, column=1, value='Status')
        ws.cell(row=row, column=2, value=computation.get_status_display() if hasattr(computation, 'get_status_display') else computation.status)
    
    auto_width_columns(ws)
    
    fiscal_year = data.get('fiscal_year', 'unknown').replace(' ', '_')
    response = create_excel_response(f'corporate_tax_{fiscal_year}.xlsx')
    wb.save(response)
    return response


# ============ VAT AUDIT REPORT EXPORT ============

def export_vat_audit(start_date, end_date, transactions, box_totals):
    """Export VAT Audit Report to Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'VAT Audit Details'
    
    # Title
    style_title_row(ws, 1, f'UAE VAT Audit Report', 8)
    ws.cell(row=2, column=1, value=f'Period: {start_date} to {end_date}')
    ws.cell(row=3, column=1, value=f'Generated: {date.today().isoformat()}')
    
    # Box Summary
    row = 5
    ws.cell(row=row, column=1, value='VAT BOX SUMMARY')
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1
    
    # Header for summary
    summary_headers = ['Box', 'Description', 'Transactions', 'Net Amount']
    for col, header in enumerate(summary_headers, 1):
        ws.cell(row=row, column=col, value=header)
        ws.cell(row=row, column=col).font = Font(bold=True)
        ws.cell(row=row, column=col).fill = PatternFill(start_color='DDDDDD', fill_type='solid')
    row += 1
    
    box_descriptions = {
        'box1a': 'Standard Rated Supplies (Emirates)',
        'box3': 'Zero-rated Supplies',
        'box4': 'Exempt Supplies',
        'box6': 'Standard Rated Expenses',
        'box9': 'Output VAT Due',
        'box10': 'Input VAT Recoverable',
    }
    
    for box_key, description in box_descriptions.items():
        box_data = box_totals.get(box_key, {'count': 0, 'net': 0})
        ws.cell(row=row, column=1, value=box_key.upper())
        ws.cell(row=row, column=2, value=description)
        ws.cell(row=row, column=3, value=box_data.get('count', 0))
        ws.cell(row=row, column=4, value=format_currency(box_data.get('net', 0)))
        row += 1
    
    row += 2
    
    # Transaction Details
    ws.cell(row=row, column=1, value='TRANSACTION DETAILS')
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1
    
    # Headers
    headers = ['Date', 'Entry #', 'Reference', 'Description', 'Account', 'Debit', 'Credit', 'VAT Box']
    for col, header in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=header)
        ws.cell(row=row, column=col).font = Font(bold=True)
        ws.cell(row=row, column=col).fill = PatternFill(start_color='DDDDDD', fill_type='solid')
    row += 1
    
    # Data rows
    for txn in transactions:
        ws.cell(row=row, column=1, value=txn['date'].strftime('%d/%m/%Y') if hasattr(txn['date'], 'strftime') else str(txn['date']))
        ws.cell(row=row, column=2, value=txn.get('entry_number', ''))
        ws.cell(row=row, column=3, value=txn.get('reference', '') or '')
        ws.cell(row=row, column=4, value=(txn.get('description', '') or '')[:50])
        ws.cell(row=row, column=5, value=txn['account'].code if hasattr(txn.get('account'), 'code') else str(txn.get('account', '')))
        ws.cell(row=row, column=6, value=format_currency(txn.get('debit', 0)) if txn.get('debit', 0) > 0 else '')
        ws.cell(row=row, column=7, value=format_currency(txn.get('credit', 0)) if txn.get('credit', 0) > 0 else '')
        ws.cell(row=row, column=8, value=txn.get('vat_box', ''))
        row += 1
    
    auto_width_columns(ws)
    
    response = create_excel_response(f'vat_audit_{start_date}_to_{end_date}.xlsx')
    wb.save(response)
    return response


# ============ TAX RECONCILIATION BRIDGE EXPORT ============

def export_bank_vs_gl(comparison_data, as_of_date):
    """Export Bank vs GL Ledger Comparison to Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Bank vs GL'

    style_title_row(ws, 1, 'Bank vs GL Ledger Comparison', 7)
    ws.cell(row=2, column=1, value=f'As of: {as_of_date}')

    headers = [
        'Bank Account', 'Account Number', 'Bank Name',
        'GL Balance (AED)', 'Bank Balance (AED)', 'Difference (AED)',
        'Last Reconciled', 'Status',
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=header)
    style_header_row(ws, 4, len(headers))

    row = 5
    total_gl = Decimal('0.00')
    total_bank = Decimal('0.00')
    total_diff = Decimal('0.00')

    diff_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    match_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

    for data in comparison_data:
        bank = data['bank']
        gl_bal = data.get('gl_balance', Decimal('0.00')) or Decimal('0.00')
        bank_bal = data.get('bank_balance', Decimal('0.00')) or Decimal('0.00')
        diff = data.get('difference', Decimal('0.00')) or Decimal('0.00')

        ws.cell(row=row, column=1, value=bank.name)
        ws.cell(row=row, column=2, value=bank.account_number)
        ws.cell(row=row, column=3, value=bank.bank_name)
        ws.cell(row=row, column=4, value=format_currency(gl_bal))
        ws.cell(row=row, column=5, value=format_currency(bank_bal))

        diff_cell = ws.cell(row=row, column=6, value=format_currency(diff))
        if abs(diff) >= Decimal('0.01'):
            diff_cell.font = Font(bold=True, color='FF0000')

        last_recon = data.get('last_reconciled')
        ws.cell(
            row=row, column=7,
            value=last_recon.strftime('%d/%m/%Y') if last_recon else 'Never',
        )

        status = data.get('recon_status', '')
        status_label = {
            'reconciled': 'Reconciled',
            'difference': 'Difference',
            'not_reconciled': 'Not Reconciled',
        }.get(status, status)
        status_cell = ws.cell(row=row, column=8, value=status_label)
        if status == 'reconciled':
            status_cell.fill = match_fill
        elif status == 'difference':
            status_cell.fill = diff_fill

        total_gl += gl_bal
        total_bank += bank_bal
        total_diff += diff
        row += 1

    for col in range(1, len(headers) + 1):
        ws.cell(row=row, column=col).border = Border(top=Side(style='double'))
    ws.cell(row=row, column=1, value='TOTAL').font = Font(bold=True)
    ws.cell(row=row, column=4, value=format_currency(total_gl)).font = Font(bold=True)
    ws.cell(row=row, column=5, value=format_currency(total_bank)).font = Font(bold=True)
    ws.cell(row=row, column=6, value=format_currency(total_diff)).font = Font(bold=True)

    auto_width_columns(ws)

    response = create_excel_response(f'bank_vs_gl_{as_of_date}.xlsx')
    wb.save(response)
    return response


def export_tax_reconciliation(ct_bridge, vat_revenue_bridge, vat_liability_bridge, fiscal_year, vat_return):
    """Export Tax Reconciliation Bridge Report to Excel."""
    wb = Workbook()

    match_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    mismatch_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    bold = Font(bold=True)

    def _header(ws, row, cols):
        for i, col in enumerate(cols, 1):
            c = ws.cell(row=row, column=i, value=col)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal='center')

    def _val(v):
        return format_currency(v) if v is not None else ''

    # ── Sheet 1: Corporate Tax Bridge ──
    if ct_bridge:
        ws = wb.active
        ws.title = 'CT Bridge'
        style_title_row(ws, 1, f'Corporate Tax Bridge — {fiscal_year.name if fiscal_year else ""}', 3)
        ws.cell(row=2, column=1, value=f'Period: {fiscal_year.start_date} to {fiscal_year.end_date}' if fiscal_year else '')

        _header(ws, 4, ['Description', 'GL Amount (AED)', 'Tax Comp. (AED)'])
        rows = [
            ('Revenue', ct_bridge['gl_revenue'], ct_bridge['ct_comp'].revenue if ct_bridge.get('ct_comp') else None),
            ('Less: Expenses', ct_bridge['gl_expenses'], ct_bridge['ct_comp'].expenses if ct_bridge.get('ct_comp') else None),
            ('Accounting Profit', ct_bridge['gl_profit'], ct_bridge['ct_comp'].accounting_profit if ct_bridge.get('ct_comp') else None),
            ('+ Non-deductible expenses', '', ct_bridge['add_backs']),
            ('- Exempt income', '', ct_bridge['exempt_income']),
            ('+/- Other adjustments', '', ct_bridge['other_adj']),
            ('Taxable Income', '', ct_bridge['taxable_income']),
            ('Less: Threshold', '', ct_bridge['threshold']),
            ('Amount Taxed @ 9%', '', ct_bridge['above_threshold']),
            ('Corporate Tax Payable', ct_bridge['computed_tax'], ct_bridge['stored_tax']),
        ]
        r = 5
        for desc, gl, tc in rows:
            ws.cell(row=r, column=1, value=desc)
            ws.cell(row=r, column=2, value=_val(gl))
            ws.cell(row=r, column=3, value=_val(tc))
            if desc in ('Accounting Profit', 'Taxable Income', 'Corporate Tax Payable'):
                for col in range(1, 4):
                    ws.cell(row=r, column=col).font = bold
            r += 1

        status_cell = ws.cell(row=r + 1, column=1, value='Reconciliation Status')
        status_cell.font = bold
        match_cell = ws.cell(row=r + 1, column=2, value='MATCH' if ct_bridge['ct_match'] else 'MISMATCH')
        match_cell.fill = match_fill if ct_bridge['ct_match'] else mismatch_fill
        match_cell.font = bold
        auto_width_columns(ws)
    else:
        ws = wb.active
        ws.title = 'CT Bridge'
        ws.cell(row=1, column=1, value='No fiscal year selected')

    # ── Sheet 2: VAT Revenue Bridge ──
    if vat_revenue_bridge:
        ws2 = wb.create_sheet('VAT Revenue Bridge')
        vr_label = f'{vat_return.return_number}' if vat_return else ''
        style_title_row(ws2, 1, f'VAT Revenue Bridge — {vr_label}', 3)
        ws2.cell(row=2, column=1,
                 value=f'Period: {vat_return.period_start} to {vat_return.period_end}' if vat_return else '')

        _header(ws2, 4, ['Source', 'Amount (AED)', 'Status'])
        items = [
            ('GL Revenue (posted journals)', vat_revenue_bridge['gl_revenue']),
            ('', ''),
            ('Box 1 — Standard Rated', vat_revenue_bridge['box1']),
            ('Box 3 — Zero Rated', vat_revenue_bridge['box3']),
            ('Box 4 — Exempt', vat_revenue_bridge['box4']),
            ('Out of Scope', vat_revenue_bridge['out_of_scope']),
            ('Total Invoice Items', vat_revenue_bridge['box_total']),
            ('', ''),
            ('Difference (GL − Boxes)', vat_revenue_bridge['difference']),
        ]
        r = 5
        for desc, val in items:
            ws2.cell(row=r, column=1, value=desc)
            ws2.cell(row=r, column=2, value=_val(val) if val != '' else '')
            r += 1

        status = ws2.cell(row=r, column=1, value='Status')
        status.font = bold
        m = ws2.cell(row=r, column=2, value='MATCH' if vat_revenue_bridge['match'] else 'MISMATCH')
        m.fill = match_fill if vat_revenue_bridge['match'] else mismatch_fill
        m.font = bold
        auto_width_columns(ws2)

    # ── Sheet 3: VAT Liability Bridge ──
    if vat_liability_bridge:
        ws3 = wb.create_sheet('VAT Liability Bridge')
        style_title_row(ws3, 1, f'VAT Liability Bridge — {vat_return.return_number if vat_return else ""}', 5)

        _header(ws3, 3, ['Account', 'GL Balance', 'Expected', 'Difference', 'Status'])
        rows = [
            ('Output VAT', 'gl_output', 'output_expected', 'output_diff', 'output_match'),
            ('Input VAT', 'gl_input', 'input_expected', 'input_diff', 'input_match'),
            ('VAT Net Payable', 'gl_vat_payable', 'payable_expected', 'payable_diff', 'payable_match'),
        ]
        r = 4
        for label, gl_k, exp_k, diff_k, match_k in rows:
            ws3.cell(row=r, column=1, value=label)
            ws3.cell(row=r, column=2, value=_val(vat_liability_bridge[gl_k]))
            ws3.cell(row=r, column=3, value=_val(vat_liability_bridge[exp_k]))
            ws3.cell(row=r, column=4, value=_val(vat_liability_bridge[diff_k]))
            status_val = 'MATCH' if vat_liability_bridge[match_k] else 'MISMATCH'
            sc = ws3.cell(row=r, column=5, value=status_val)
            sc.fill = match_fill if vat_liability_bridge[match_k] else mismatch_fill
            sc.font = bold
            r += 1
        auto_width_columns(ws3)

    fy_name = fiscal_year.name.replace(' ', '_') if fiscal_year else 'none'
    response = create_excel_response(f'tax_reconciliation_{fy_name}.xlsx')
    wb.save(response)
    return response


# ============ ASSET REGISTER EXPORT ============

def export_depreciation_report(records, totals, from_date, to_date):
    """Export Depreciation Report to Excel with reconciliation."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Depreciation Report'

    style_title_row(ws, 1, 'Depreciation Report', 9)
    ws.cell(row=2, column=1, value=f'Period: {from_date} to {to_date}')
    ws.cell(row=3, column=1, value=f'Generated: {date.today().isoformat()}')

    headers = [
        'Date', 'Asset #', 'Asset Name', 'Category',
        'Cost (AED)', 'Period Depr. (AED)', 'Accum. Depr. (AED)',
        'Book Value (AED)', 'Journal Ref',
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=5, column=col, value=h)
    style_header_row(ws, 5, len(headers))

    row = 6
    sum_cost = Decimal('0')
    sum_depr = Decimal('0')
    sum_accum = Decimal('0')
    sum_nbv = Decimal('0')

    for rec in records:
        dep_date = rec.get('date')
        cost = rec.get('cost', Decimal('0'))
        depr = rec.get('depreciation_amount', Decimal('0'))
        accum = rec.get('accumulated_depreciation', Decimal('0'))
        nbv = rec.get('book_value', Decimal('0'))

        ws.cell(row=row, column=1, value=dep_date.strftime('%d/%m/%Y') if hasattr(dep_date, 'strftime') else str(dep_date))
        ws.cell(row=row, column=2, value=rec.get('asset_number', ''))
        ws.cell(row=row, column=3, value=rec.get('asset_name', ''))
        ws.cell(row=row, column=4, value=rec.get('category', ''))
        ws.cell(row=row, column=5, value=format_currency(cost))
        ws.cell(row=row, column=6, value=format_currency(depr))
        ws.cell(row=row, column=7, value=format_currency(accum))
        ws.cell(row=row, column=8, value=format_currency(nbv))
        ws.cell(row=row, column=9, value=rec.get('journal_ref', '-'))

        sum_cost += cost
        sum_depr += depr
        sum_accum += accum
        sum_nbv += nbv
        row += 1

    for col in range(1, len(headers) + 1):
        ws.cell(row=row, column=col).border = Border(top=Side(style='double'))
    ws.cell(row=row, column=4, value='TOTAL').font = Font(bold=True)
    ws.cell(row=row, column=5, value=format_currency(sum_cost)).font = Font(bold=True)
    ws.cell(row=row, column=6, value=format_currency(sum_depr)).font = Font(bold=True)
    ws.cell(row=row, column=7, value=format_currency(sum_accum)).font = Font(bold=True)
    ws.cell(row=row, column=8, value=format_currency(sum_nbv)).font = Font(bold=True)

    row += 2
    ws.cell(row=row, column=1, value='RECONCILIATION CHECK')
    ws.cell(row=row, column=1).font = Font(bold=True, size=11)
    row += 1

    rec_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    for label, val in [
        ('Total Asset Cost', totals.get('total_cost', Decimal('0'))),
        ('Total Accumulated Depreciation', totals.get('total_accumulated', Decimal('0'))),
        ('Total Net Book Value (Cost − Accum. Depr.)', totals.get('total_book_value', Decimal('0'))),
    ]:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=format_currency(val)).font = Font(bold=True)
        for c in range(1, 3):
            ws.cell(row=row, column=c).fill = rec_fill
        row += 1

    row += 1
    ws.cell(row=row, column=1, value=f'Total Depreciation for Period: AED {float(totals.get("total_depreciation", 0)):,.2f}')
    ws.cell(row=row, column=1).font = Font(bold=True, size=11)

    auto_width_columns(ws)

    response = create_excel_response(f'depreciation_report_{from_date}_to_{to_date}.xlsx')
    wb.save(response)
    return response


def export_asset_register(assets, gl_reconciliation=None, as_of_date=None):
    """Export Fixed Asset Register with GL reconciliation to Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Asset Register'

    title = 'Fixed Asset Register'
    if as_of_date:
        title += f' as of {as_of_date}'
    style_title_row(ws, 1, title, 10)

    headers = [
        'Asset #', 'Name', 'Category', 'Status',
        'Acquisition Date', 'Cost', 'Accum Depreciation',
        'Book Value', 'Method', 'Useful Life (Yrs)',
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    style_header_row(ws, 3, len(headers))

    row = 4
    totals = {'cost': Decimal('0'), 'accum': Decimal('0'), 'nbv': Decimal('0')}
    for a in assets:
        cost = a.get('cost', Decimal('0'))
        accum = a.get('accum_depreciation', Decimal('0'))
        nbv = a.get('book_value', cost - accum)
        totals['cost'] += cost
        totals['accum'] += accum
        totals['nbv'] += nbv

        ws.cell(row=row, column=1, value=a.get('asset_number', ''))
        ws.cell(row=row, column=2, value=a.get('name', ''))
        ws.cell(row=row, column=3, value=a.get('category', ''))
        ws.cell(row=row, column=4, value=a.get('status', ''))
        ws.cell(row=row, column=5, value=str(a.get('acquisition_date', '')))
        ws.cell(row=row, column=6, value=format_currency(cost))
        ws.cell(row=row, column=7, value=format_currency(accum))
        ws.cell(row=row, column=8, value=format_currency(nbv))
        ws.cell(row=row, column=9, value=a.get('method', ''))
        ws.cell(row=row, column=10, value=a.get('useful_life', ''))

        has_je = a.get('has_journal', False)
        if not has_je:
            for c in range(1, 11):
                ws.cell(row=row, column=c).fill = PatternFill(
                    start_color='FFD7D7', end_color='FFD7D7', fill_type='solid'
                )
        row += 1

    ws.cell(row=row, column=5, value='TOTAL')
    ws.cell(row=row, column=5).font = Font(bold=True)
    ws.cell(row=row, column=6, value=format_currency(totals['cost']))
    ws.cell(row=row, column=7, value=format_currency(totals['accum']))
    ws.cell(row=row, column=8, value=format_currency(totals['nbv']))
    for c in range(5, 9):
        ws.cell(row=row, column=c).font = Font(bold=True)

    if gl_reconciliation:
        row += 2
        ws2 = wb.create_sheet('GL Reconciliation')
        style_title_row(ws2, 1, 'Asset Register vs GL Reconciliation', 4)

        rec_headers = ['', 'Asset Register', 'GL Balance', 'Difference']
        for col, h in enumerate(rec_headers, 1):
            ws2.cell(row=3, column=col, value=h)
        style_header_row(ws2, 3, len(rec_headers))

        r = 4
        for label, reg_key, gl_key in [
            ('Gross Cost', 'register_cost', 'gl_cost'),
            ('Accum Depreciation', 'register_accum', 'gl_accum'),
            ('Net Book Value', 'register_nbv', 'gl_nbv'),
        ]:
            reg_val = gl_reconciliation.get(reg_key, Decimal('0'))
            gl_val = gl_reconciliation.get(gl_key, Decimal('0'))
            diff = reg_val - gl_val
            ws2.cell(row=r, column=1, value=label)
            ws2.cell(row=r, column=2, value=format_currency(reg_val))
            ws2.cell(row=r, column=3, value=format_currency(gl_val))
            ws2.cell(row=r, column=4, value=format_currency(diff))

            match_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            mismatch_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            ws2.cell(row=r, column=4).fill = match_fill if abs(diff) < Decimal('0.01') else mismatch_fill
            r += 1

        auto_width_columns(ws2)

    auto_width_columns(ws)

    response = create_excel_response(f'asset_register_{as_of_date or "all"}.xlsx')
    wb.save(response)
    return response

