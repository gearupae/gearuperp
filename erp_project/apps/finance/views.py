"""
Finance Views - UAE VAT & Corporate Tax Compliant
Chart of Accounts, Journal Entries, Payments, Reports
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import get_user_model
from django.contrib import messages
from django.views.generic import ListView, CreateView, UpdateView, DetailView, TemplateView
from django.urls import reverse_lazy
from django.db.models import Q, Sum, F
from django.core.exceptions import ValidationError
from django.core.paginator import Paginator
from django.http import JsonResponse
from datetime import date, timedelta
from decimal import Decimal

User = get_user_model()

from .models import (
    Account, AccountType, FiscalYear, AccountingPeriod, JournalEntry, JournalEntryLine, 
    TaxCode, Payment, BankAccount, ExpenseClaim, ExpenseItem, VATReturn, CorporateTaxComputation,
    Budget, BudgetLine, BankTransfer, BankReconciliation, BankStatement, BankStatementLine,
    ReconciliationItem, OpeningBalanceEntry, OpeningBalanceLine, WriteOff, ExchangeRate,
    AccountMapping, AccountingSettings
)
from .forms import (
    AccountForm, FiscalYearForm, AccountingPeriodForm, JournalEntryForm, JournalEntryLineFormSet, 
    PaymentForm, TaxCodeForm, BankAccountForm, ExpenseClaimForm, ExpenseItemFormSet,
    VATReturnForm, CorporateTaxForm, BudgetForm, BudgetLineFormSet, BankTransferForm,
    BankReconciliationForm, BankStatementForm, BankStatementLineFormSet,
    OpeningBalanceEntryForm, OpeningBalanceLineFormSet, WriteOffForm, ExchangeRateForm
)
from django import forms
from apps.core.mixins import PermissionRequiredMixin, CreatePermissionMixin, UpdatePermissionMixin
from apps.core.utils import PermissionChecker


# ============ CHART OF ACCOUNTS VIEWS ============

class AccountListView(PermissionRequiredMixin, ListView):
    model = Account
    template_name = 'finance/account_list.html'
    context_object_name = 'accounts'
    module_name = 'finance'
    permission_type = 'view'
    
    def get_queryset(self):
        queryset = Account.objects.filter(is_active=True)
        
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(code__icontains=search) | Q(name__icontains=search)
            )
        
        account_type = self.request.GET.get('type')
        if account_type:
            queryset = queryset.filter(account_type=account_type)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Chart of Accounts'
        context['account_types'] = AccountType.choices
        context['form'] = AccountForm()
        context['can_create'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'create')
        context['can_edit'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'edit')
        context['can_delete'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'delete')
        
        # Flag abnormal balances
        for account in context['accounts']:
            account.abnormal = account.has_abnormal_balance
        
        return context
    
    def post(self, request, *args, **kwargs):
        if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'create')):
            messages.error(request, 'Permission denied.')
            return redirect('finance:account_list')
        
        form = AccountForm(request.POST)
        if form.is_valid():
            account = form.save()
            messages.success(request, f'Account {account.code} created.')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
        return redirect('finance:account_list')


class AccountUpdateView(UpdatePermissionMixin, UpdateView):
    model = Account
    form_class = AccountForm
    template_name = 'finance/account_form.html'
    success_url = reverse_lazy('finance:account_list')
    module_name = 'finance'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Edit Account: {self.object.code}'
        return context
    
    def form_valid(self, form):
        messages.success(self.request, f'Account {form.instance.code} updated.')
        return super().form_valid(form)


@login_required
def account_delete(request, pk):
    """Soft delete - system accounts cannot be deleted."""
    account = get_object_or_404(Account, pk=pk)
    if account.is_system:
        messages.error(request, 'System accounts cannot be deleted.')
        return redirect('finance:account_list')
    
    if request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'delete'):
        account.is_active = False
        account.save()
        messages.success(request, f'Account {account.code} deleted.')
    else:
        messages.error(request, 'Permission denied.')
    return redirect('finance:account_list')


# ============ JOURNAL ENTRY VIEWS ============

class JournalEntryListView(PermissionRequiredMixin, ListView):
    model = JournalEntry
    template_name = 'finance/journal_list.html'
    context_object_name = 'entries'
    module_name = 'finance'
    permission_type = 'view'
    paginate_by = 25
    
    def get_queryset(self):
        queryset = JournalEntry.objects.filter(is_active=True)
        
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(entry_number__icontains=search) | Q(reference__icontains=search)
            )
        
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Journal Entries'
        context['status_choices'] = JournalEntry.STATUS_CHOICES
        context['can_create'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'create')
        context['can_edit'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'edit')
        context['today'] = date.today().isoformat()
        return context


class JournalEntryCreateView(CreatePermissionMixin, CreateView):
    model = JournalEntry
    form_class = JournalEntryForm
    template_name = 'finance/journal_form.html'
    success_url = reverse_lazy('finance:journal_list')
    module_name = 'finance'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Create Journal Entry'
        context['today'] = date.today().isoformat()
        if self.request.POST:
            context['lines_formset'] = JournalEntryLineFormSet(self.request.POST)
        else:
            context['lines_formset'] = JournalEntryLineFormSet()
        context['accounts'] = Account.objects.filter(is_active=True).order_by('code')
        return context
    
    def form_valid(self, form):
        context = self.get_context_data()
        lines_formset = context['lines_formset']
        
        if lines_formset.is_valid():
            self.object = form.save()
            lines_formset.instance = self.object
            lines_formset.save()
            self.object.calculate_totals()
            
            if not self.object.is_balanced:
                messages.warning(self.request, f'Journal Entry {self.object.entry_number} created but is UNBALANCED. Please correct before posting.')
            else:
                messages.success(self.request, f'Journal Entry {self.object.entry_number} created.')
            return redirect(self.success_url)
        else:
            return self.render_to_response(self.get_context_data(form=form))


class JournalEntryUpdateView(UpdatePermissionMixin, UpdateView):
    model = JournalEntry
    form_class = JournalEntryForm
    template_name = 'finance/journal_form.html'
    success_url = reverse_lazy('finance:journal_list')
    module_name = 'finance'
    
    def get_object(self, queryset=None):
        obj = super().get_object(queryset)
        # Use the new is_editable property for SAP/Oracle compliant rules
        if not obj.is_editable:
            reason = obj.edit_restriction_reason or 'This journal entry cannot be edited.'
            messages.error(self.request, reason)
            return None
        return obj
    
    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        if self.object is None:
            return redirect('finance:journal_list')
        return super().get(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Edit Journal Entry: {self.object.entry_number}'
        context['today'] = date.today().isoformat()
        if self.request.POST:
            context['lines_formset'] = JournalEntryLineFormSet(self.request.POST, instance=self.object)
        else:
            context['lines_formset'] = JournalEntryLineFormSet(instance=self.object)
        context['accounts'] = Account.objects.filter(is_active=True).order_by('code')
        return context
    
    def form_valid(self, form):
        context = self.get_context_data()
        lines_formset = context['lines_formset']
        
        if lines_formset.is_valid():
            self.object = form.save()
            lines_formset.save()
            self.object.calculate_totals()
            messages.success(self.request, f'Journal Entry {self.object.entry_number} updated.')
            return redirect(self.success_url)
        else:
            return self.render_to_response(self.get_context_data(form=form))


class JournalEntryDetailView(PermissionRequiredMixin, DetailView):
    model = JournalEntry
    template_name = 'finance/journal_detail.html'
    context_object_name = 'entry'
    module_name = 'finance'
    permission_type = 'view'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Journal Entry: {self.object.entry_number}'
        
        has_edit_permission = (
            self.request.user.is_superuser or 
            PermissionChecker.has_permission(self.request.user, 'finance', 'edit')
        )
        
        # Use new SAP/Oracle compliant properties
        context['can_edit'] = has_edit_permission and self.object.is_editable
        context['can_delete'] = has_edit_permission and self.object.is_deletable
        context['can_post'] = has_edit_permission and self.object.is_editable and self.object.is_balanced and self.object.line_count >= 2
        context['can_reverse'] = has_edit_permission and self.object.is_reversible
        
        # Show reason why actions are blocked
        context['edit_restriction_reason'] = self.object.edit_restriction_reason
        context['is_system_generated'] = self.object.is_system_generated
        context['is_locked'] = self.object.is_locked
        
        return context


@login_required
def journal_post(request, pk):
    """Post a journal entry - validates balance, min lines, leaf accounts, period."""
    entry = get_object_or_404(JournalEntry, pk=pk)
    
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:journal_list')
    
    if entry.status != 'draft':
        messages.error(request, 'Only draft entries can be posted.')
        return redirect('finance:journal_detail', pk=pk)
    
    # Validate before posting
    errors = entry.validate_for_posting()
    if errors:
        for error in errors:
            messages.error(request, error)
        return redirect('finance:journal_detail', pk=pk)
    
    try:
        entry.post(user=request.user)
        messages.success(request, f'Journal Entry {entry.entry_number} posted successfully.')
    except ValidationError as e:
        for error in e.messages:
            messages.error(request, error)
    except Exception as e:
        messages.error(request, str(e))
    
    return redirect('finance:journal_detail', pk=pk)


@login_required
def journal_reverse(request, pk):
    """
    Reverse a posted journal entry - creates auto-reversal entry.
    This is the ONLY way to correct posted transactions (SAP/Oracle compliant).
    """
    entry = get_object_or_404(JournalEntry, pk=pk)
    
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:journal_list')
    
    if not entry.is_reversible:
        if entry.status != 'posted':
            messages.error(request, 'Only posted entries can be reversed.')
        elif entry.period and entry.period.is_locked:
            messages.error(request, f'Cannot reverse - accounting period {entry.period.name} is locked.')
        elif entry.fiscal_year and entry.fiscal_year.is_closed:
            messages.error(request, f'Cannot reverse - fiscal year {entry.fiscal_year.name} is closed.')
        else:
            messages.error(request, 'This journal entry cannot be reversed.')
        return redirect('finance:journal_detail', pk=pk)
    
    reason = request.POST.get('reason', 'User requested reversal')
    
    try:
        reversal = entry.reverse(user=request.user, reason=reason)
        messages.success(request, f'Journal Entry {entry.entry_number} reversed. Reversal entry: {reversal.entry_number}')
        return redirect('finance:journal_detail', pk=reversal.pk)
    except ValidationError as e:
        for error in e.messages:
            messages.error(request, error)
    except Exception as e:
        messages.error(request, str(e))
    
    return redirect('finance:journal_detail', pk=pk)


# ============ PAYMENT VIEWS ============

class PaymentListView(PermissionRequiredMixin, ListView):
    model = Payment
    template_name = 'finance/payment_list.html'
    context_object_name = 'payments'
    module_name = 'finance'
    permission_type = 'view'
    paginate_by = 25
    
    def get_queryset(self):
        queryset = Payment.objects.filter(is_active=True)
        
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(payment_number__icontains=search) |
                Q(party_name__icontains=search) |
                Q(reference__icontains=search)
            )
        
        payment_type = self.request.GET.get('type')
        if payment_type:
            queryset = queryset.filter(payment_type=payment_type)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Payments'
        context['type_choices'] = Payment.PAYMENT_TYPE_CHOICES
        context['can_create'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'create')
        context['can_edit'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'edit')
        context['can_delete'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'delete')
        context['today'] = date.today().isoformat()
        
        # Summary
        payments = self.get_queryset()
        context['total_received'] = payments.filter(payment_type='received').aggregate(Sum('amount'))['amount__sum'] or 0
        context['total_made'] = payments.filter(payment_type='made').aggregate(Sum('amount'))['amount__sum'] or 0
        return context


class PaymentCreateView(CreatePermissionMixin, CreateView):
    model = Payment
    form_class = PaymentForm
    template_name = 'finance/payment_form.html'
    success_url = reverse_lazy('finance:payment_list')
    module_name = 'finance'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Create Payment'
        context['today'] = date.today().isoformat()
        return context
    
    def form_valid(self, form):
        form.instance.party_type = 'customer' if form.instance.payment_type == 'received' else 'vendor'
        form.instance.party_id = 0
        messages.success(self.request, 'Payment created.')
        return super().form_valid(form)


class PaymentUpdateView(UpdatePermissionMixin, UpdateView):
    model = Payment
    form_class = PaymentForm
    template_name = 'finance/payment_form.html'
    success_url = reverse_lazy('finance:payment_list')
    module_name = 'finance'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Edit Payment: {self.object.payment_number}'
        context['today'] = date.today().isoformat()
        return context
    
    def form_valid(self, form):
        messages.success(self.request, f'Payment {form.instance.payment_number} updated.')
        return super().form_valid(form)


@login_required
def payment_cancel(request, pk):
    """Cancel a payment - creates auto-reversal journal."""
    payment = get_object_or_404(Payment, pk=pk)
    
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:payment_list')
    
    if payment.status == 'cancelled':
        messages.error(request, 'Payment is already cancelled.')
        return redirect('finance:payment_list')
    
    from django.utils import timezone
    payment.status = 'cancelled'
    payment.cancelled_date = timezone.now()
    payment.cancellation_reason = request.POST.get('reason', 'User requested cancellation')
    payment.save()
    
    # If there's a linked journal entry, reverse it
    if payment.journal_entry and payment.journal_entry.status == 'posted':
        try:
            reversal = payment.journal_entry.reverse(user=request.user, reason=f'Payment {payment.payment_number} cancelled')
            payment.reversal_entry = reversal
            payment.save()
        except Exception as e:
            messages.warning(request, f'Could not reverse journal entry: {e}')
    
    messages.success(request, f'Payment {payment.payment_number} cancelled.')
    return redirect('finance:payment_list')


@login_required
def payment_delete(request, pk):
    """Delete a draft payment (soft delete)."""
    payment = get_object_or_404(Payment, pk=pk)
    
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'delete')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:payment_list')
    
    if payment.status != 'draft':
        messages.error(request, 'Only draft payments can be deleted.')
        return redirect('finance:payment_list')
    
    # Soft delete
    payment.is_active = False
    payment.save()
    
    messages.success(request, f'Payment {payment.payment_number} deleted.')
    return redirect('finance:payment_list')


# ============ TAX CODE VIEWS ============

class TaxCodeListView(PermissionRequiredMixin, ListView):
    model = TaxCode
    template_name = 'finance/taxcode_list.html'
    context_object_name = 'taxcodes'
    module_name = 'finance'
    permission_type = 'view'
    
    def get_queryset(self):
        return TaxCode.objects.filter(is_active=True)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Tax Codes'
        context['form'] = TaxCodeForm()
        context['can_create'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'create')
        context['can_edit'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'edit')
        return context
    
    def post(self, request, *args, **kwargs):
        if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'create')):
            messages.error(request, 'Permission denied.')
            return redirect('finance:taxcode_list')
        
        form = TaxCodeForm(request.POST)
        if form.is_valid():
            taxcode = form.save()
            messages.success(request, f'Tax Code {taxcode.code} created.')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
        return redirect('finance:taxcode_list')


# ============ FINANCIAL REPORTS ============

@login_required
def trial_balance(request):
    """Trial Balance Report - Total Debit must equal Total Credit."""
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'view')):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard')
    
    as_of_date = request.GET.get('date', date.today().isoformat())
    export_format = request.GET.get('format', '')
    
    accounts = Account.objects.filter(is_active=True).order_by('code')
    
    trial_data = []
    total_debit = Decimal('0.00')
    total_credit = Decimal('0.00')
    
    for account in accounts:
        balance = account.current_balance
        debit = Decimal('0.00')
        credit = Decimal('0.00')
        
        if account.debit_increases:
            if balance >= 0:
                debit = balance
            else:
                credit = abs(balance)
        else:
            if balance >= 0:
                credit = balance
            else:
                debit = abs(balance)
        
        if debit != 0 or credit != 0:
            trial_data.append({
                'account': account,
                'code': account.code,
                'name': account.name,
                'debit': debit,
                'credit': credit,
                'abnormal': account.has_abnormal_balance,
            })
            total_debit += debit
            total_credit += credit
    
    is_balanced = total_debit == total_credit
    
    # Excel Export
    if export_format == 'excel':
        from .excel_exports import export_trial_balance
        return export_trial_balance(trial_data, as_of_date)
    
    return render(request, 'finance/trial_balance.html', {
        'title': 'Trial Balance',
        'trial_data': trial_data,
        'total_debit': total_debit,
        'total_credit': total_credit,
        'is_balanced': is_balanced,
        'as_of_date': as_of_date,
    })


@login_required
def profit_loss(request):
    """
    Profit & Loss Statement (Income Statement).
    SINGLE SOURCE OF TRUTH: Reads only from JournalEntryLine for Income and Expense accounts.
    """
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'view')):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard')
    
    # Get date range
    end_date = request.GET.get('end_date', date.today().isoformat())
    start_date = request.GET.get('start_date', date(date.today().year, 1, 1).isoformat())
    
    # Income accounts - calculate balance from journal lines
    income_accounts = Account.objects.filter(
        is_active=True, 
        account_type=AccountType.INCOME
    ).order_by('code')
    
    income_data = []
    total_income = Decimal('0.00')
    for acc in income_accounts:
        # Income accounts: Credits increase, Debits decrease (Credit - Debit = Balance)
        lines = JournalEntryLine.objects.filter(
            account=acc,
            journal_entry__status='posted',
            journal_entry__date__gte=start_date,
            journal_entry__date__lte=end_date,
        ).aggregate(
            total_debit=Sum('debit'),
            total_credit=Sum('credit')
        )
        debit = lines['total_debit'] or Decimal('0.00')
        credit = lines['total_credit'] or Decimal('0.00')
        balance = credit - debit  # Income is increased by credits
        
        if balance != 0:
            income_data.append({'account': acc, 'amount': balance})
            total_income += balance
    
    # Expense accounts - calculate balance from journal lines
    expense_accounts = Account.objects.filter(
        is_active=True, 
        account_type=AccountType.EXPENSE
    ).order_by('code')
    
    expense_data = []
    total_expenses = Decimal('0.00')
    for acc in expense_accounts:
        # Expense accounts: Debits increase, Credits decrease (Debit - Credit = Balance)
        lines = JournalEntryLine.objects.filter(
            account=acc,
            journal_entry__status='posted',
            journal_entry__date__gte=start_date,
            journal_entry__date__lte=end_date,
        ).aggregate(
            total_debit=Sum('debit'),
            total_credit=Sum('credit')
        )
        debit = lines['total_debit'] or Decimal('0.00')
        credit = lines['total_credit'] or Decimal('0.00')
        balance = debit - credit  # Expense is increased by debits
        
        if balance != 0:
            expense_data.append({'account': acc, 'amount': balance})
            total_expenses += balance
    
    # Calculate profit
    net_profit_before_tax = total_income - total_expenses
    
    # Corporate tax calculation (9% on profit > AED 375,000)
    tax_threshold = Decimal('375000.00')
    tax_rate = Decimal('0.09')
    
    if net_profit_before_tax > tax_threshold:
        corporate_tax = (net_profit_before_tax - tax_threshold) * tax_rate
    else:
        corporate_tax = Decimal('0.00')
    
    net_profit_after_tax = net_profit_before_tax - corporate_tax
    
    # Excel Export
    export_format = request.GET.get('format', '')
    if export_format == 'excel':
        from .excel_exports import export_profit_loss
        # Prepare data for export
        revenue_export = [{'code': d['account'].code, 'name': d['account'].name, 'balance': d['amount']} for d in income_data]
        expense_export = [{'code': d['account'].code, 'name': d['account'].name, 'balance': d['amount']} for d in expense_data]
        return export_profit_loss(revenue_export, expense_export, start_date, end_date)
    
    return render(request, 'finance/profit_loss.html', {
        'title': 'Profit & Loss Statement',
        'income_data': income_data,
        'expense_data': expense_data,
        'total_income': total_income,
        'total_expenses': total_expenses,
        'net_profit_before_tax': net_profit_before_tax,
        'corporate_tax': corporate_tax,
        'net_profit_after_tax': net_profit_after_tax,
        'start_date': start_date,
        'end_date': end_date,
    })


@login_required
def balance_sheet(request):
    """
    Balance Sheet - Assets = Liabilities + Equity.
    SINGLE SOURCE OF TRUTH: Reads only from JournalEntryLine for all account types.
    """
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'view')):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard')
    
    as_of_date = request.GET.get('date', date.today().isoformat())
    
    def get_account_balance(account, up_to_date):
        """Calculate account balance from journal lines up to a specific date."""
        lines = JournalEntryLine.objects.filter(
            account=account,
            journal_entry__status='posted',
            journal_entry__date__lte=up_to_date,
        ).aggregate(
            total_debit=Sum('debit'),
            total_credit=Sum('credit')
        )
        debit = lines['total_debit'] or Decimal('0.00')
        credit = lines['total_credit'] or Decimal('0.00')
        
        # Account type determines balance calculation
        if account.debit_increases:
            # Assets and Expenses: Debit - Credit
            return account.opening_balance + (debit - credit)
        else:
            # Liabilities, Equity, Income: Credit - Debit
            return account.opening_balance + (credit - debit)
    
    # Assets
    asset_accounts = Account.objects.filter(
        is_active=True, 
        account_type=AccountType.ASSET
    ).order_by('code')
    
    asset_data = []
    total_assets = Decimal('0.00')
    for acc in asset_accounts:
        balance = get_account_balance(acc, as_of_date)
        if balance != 0:
            asset_data.append({'account': acc, 'amount': balance})
            total_assets += balance
    
    # Liabilities
    liability_accounts = Account.objects.filter(
        is_active=True, 
        account_type=AccountType.LIABILITY
    ).order_by('code')
    
    liability_data = []
    total_liabilities = Decimal('0.00')
    for acc in liability_accounts:
        balance = get_account_balance(acc, as_of_date)
        if balance != 0:
            liability_data.append({'account': acc, 'amount': balance})
            total_liabilities += balance
    
    # Equity
    equity_accounts = Account.objects.filter(
        is_active=True, 
        account_type=AccountType.EQUITY
    ).order_by('code')
    
    equity_data = []
    total_equity = Decimal('0.00')
    for acc in equity_accounts:
        balance = get_account_balance(acc, as_of_date)
        if balance != 0:
            equity_data.append({'account': acc, 'amount': balance})
            total_equity += balance
    
    # Calculate retained earnings (current year P&L from journal lines)
    income_accounts = Account.objects.filter(
        is_active=True, account_type=AccountType.INCOME
    )
    expense_accounts = Account.objects.filter(
        is_active=True, account_type=AccountType.EXPENSE
    )
    
    income_lines = JournalEntryLine.objects.filter(
        account__in=income_accounts,
        journal_entry__status='posted',
        journal_entry__date__lte=as_of_date,
    ).aggregate(
        total_debit=Sum('debit'),
        total_credit=Sum('credit')
    )
    income_total = (income_lines['total_credit'] or Decimal('0.00')) - (income_lines['total_debit'] or Decimal('0.00'))
    
    expense_lines = JournalEntryLine.objects.filter(
        account__in=expense_accounts,
        journal_entry__status='posted',
        journal_entry__date__lte=as_of_date,
    ).aggregate(
        total_debit=Sum('debit'),
        total_credit=Sum('credit')
    )
    expense_total = (expense_lines['total_debit'] or Decimal('0.00')) - (expense_lines['total_credit'] or Decimal('0.00'))
    
    current_year_profit = income_total - expense_total
    total_equity += current_year_profit
    
    # Balance check
    total_liabilities_equity = total_liabilities + total_equity
    is_balanced = abs(total_assets - total_liabilities_equity) < Decimal('0.01')  # Allow small rounding difference
    
    # Excel Export
    export_format = request.GET.get('format', '')
    if export_format == 'excel':
        from .excel_exports import export_balance_sheet
        assets_export = [{'code': d['account'].code, 'name': d['account'].name, 'balance': d['amount']} for d in asset_data]
        liabilities_export = [{'code': d['account'].code, 'name': d['account'].name, 'balance': d['amount']} for d in liability_data]
        equity_export = [{'code': d['account'].code, 'name': d['account'].name, 'balance': d['amount']} for d in equity_data]
        if current_year_profit != 0:
            equity_export.append({'code': '', 'name': 'Current Year Profit/Loss', 'balance': current_year_profit})
        return export_balance_sheet(assets_export, liabilities_export, equity_export, as_of_date)
    
    return render(request, 'finance/balance_sheet.html', {
        'title': 'Balance Sheet',
        'asset_data': asset_data,
        'liability_data': liability_data,
        'equity_data': equity_data,
        'total_assets': total_assets,
        'total_liabilities': total_liabilities,
        'total_equity': total_equity,
        'current_year_profit': current_year_profit,
        'total_liabilities_equity': total_liabilities_equity,
        'is_balanced': is_balanced,
        'as_of_date': as_of_date,
    })


@login_required
def general_ledger(request):
    """General Ledger - All transactions for an account."""
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'view')):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard')
    
    account_id = request.GET.get('account')
    start_date = request.GET.get('start_date', date(date.today().year, 1, 1).isoformat())
    end_date = request.GET.get('end_date', date.today().isoformat())
    
    accounts = Account.objects.filter(is_active=True).order_by('code')
    selected_account = None
    transactions = []
    running_balance = Decimal('0.00')
    
    if account_id:
        selected_account = get_object_or_404(Account, pk=account_id)
        running_balance = selected_account.opening_balance
        
        lines = JournalEntryLine.objects.filter(
            account=selected_account,
            journal_entry__status='posted',
            journal_entry__date__gte=start_date,
            journal_entry__date__lte=end_date,
        ).select_related('journal_entry').order_by('journal_entry__date', 'id')
        
        for line in lines:
            if selected_account.debit_increases:
                running_balance += line.debit - line.credit
            else:
                running_balance += line.credit - line.debit
            
            transactions.append({
                'date': line.journal_entry.date,
                'journal_pk': line.journal_entry.pk,
                'entry_number': line.journal_entry.entry_number,
                'reference': line.journal_entry.reference,
                'description': line.description or line.journal_entry.description,
                'debit': line.debit,
                'credit': line.credit,
                'balance': running_balance,
            })
    
    # Excel Export
    export_format = request.GET.get('format', '')
    if export_format == 'excel' and selected_account:
        from .excel_exports import export_general_ledger
        return export_general_ledger(transactions, selected_account.name, start_date, end_date)
    
    return render(request, 'finance/general_ledger.html', {
        'title': 'General Ledger',
        'accounts': accounts,
        'selected_account': selected_account,
        'transactions': transactions,
        'opening_balance': selected_account.opening_balance if selected_account else Decimal('0.00'),
        'closing_balance': running_balance,
        'start_date': start_date,
        'end_date': end_date,
    })


@login_required
def vat_report(request):
    """
    UAE VAT Return Report (FTA format).
    SINGLE SOURCE OF TRUTH: VAT calculated ONLY from JournalEntryLine (VAT accounts).
    """
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'view')):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard')
    
    # Get date range
    start_date = request.GET.get('start_date', date(date.today().year, (date.today().month - 1) // 3 * 3 + 1, 1).isoformat())
    end_date = request.GET.get('end_date', date.today().isoformat())
    
    # Get VAT returns
    vat_returns = VATReturn.objects.filter(is_active=True).order_by('-period_start')
    
    # Get VAT Payable account (Output VAT - typically 2100)
    vat_payable_account = Account.objects.filter(
        code__startswith='21', account_type=AccountType.LIABILITY, is_active=True
    ).first()
    
    # Get VAT Recoverable account (Input VAT - typically 1300)
    vat_recoverable_account = Account.objects.filter(
        code__startswith='13', account_type=AccountType.ASSET, is_active=True
    ).first()
    
    # Get Sales accounts (Income - typically 4xxx)
    sales_accounts = Account.objects.filter(
        account_type=AccountType.INCOME, is_active=True
    )
    
    # Get Expense accounts (Expense - typically 5xxx)
    expense_accounts = Account.objects.filter(
        account_type=AccountType.EXPENSE, is_active=True
    )
    
    # Calculate Output VAT from VAT Payable journal lines
    output_vat_lines = JournalEntryLine.objects.filter(
        account=vat_payable_account,
        journal_entry__status='posted',
        journal_entry__date__gte=start_date,
        journal_entry__date__lte=end_date,
    ) if vat_payable_account else JournalEntryLine.objects.none()
    
    current_output_vat = output_vat_lines.aggregate(
        total=Sum('credit')
    )['total'] or Decimal('0.00')
    
    # Calculate Input VAT from VAT Recoverable journal lines
    input_vat_lines = JournalEntryLine.objects.filter(
        account=vat_recoverable_account,
        journal_entry__status='posted',
        journal_entry__date__gte=start_date,
        journal_entry__date__lte=end_date,
    ) if vat_recoverable_account else JournalEntryLine.objects.none()
    
    current_input_vat = input_vat_lines.aggregate(
        total=Sum('debit')
    )['total'] or Decimal('0.00')
    
    # Calculate Sales from Income account journal lines (Credits = Sales)
    sales_lines = JournalEntryLine.objects.filter(
        account__in=sales_accounts,
        journal_entry__status='posted',
        journal_entry__date__gte=start_date,
        journal_entry__date__lte=end_date,
    )
    
    current_sales = sales_lines.aggregate(
        total=Sum('credit')
    )['total'] or Decimal('0.00')
    
    # Calculate Purchases from Expense account journal lines (Debits = Expenses)
    expense_lines = JournalEntryLine.objects.filter(
        account__in=expense_accounts,
        journal_entry__status='posted',
        journal_entry__date__gte=start_date,
        journal_entry__date__lte=end_date,
    )
    
    current_purchases = expense_lines.aggregate(
        total=Sum('debit')
    )['total'] or Decimal('0.00')
    
    # Standard Rated = Amounts where VAT was charged (5%)
    # This is a simplified calculation - assumes all sales/purchases are standard rated
    standard_rated_supplies = current_sales
    standard_rated_vat = current_output_vat
    
    standard_rated_expenses = current_purchases
    
    # Net VAT
    current_net_vat = current_output_vat - current_input_vat
    
    # Excel Export
    export_format = request.GET.get('format', '')
    if export_format == 'excel':
        from .excel_exports import export_vat_report
        vat_data = {
            'standard_sales': current_sales,
            'output_vat': current_output_vat,
            'zero_rated_sales': Decimal('0.00'),
            'exempt_sales': Decimal('0.00'),
            'standard_purchases': current_purchases,
            'input_vat': current_input_vat,
        }
        return export_vat_report(vat_data, start_date, end_date)
    
    return render(request, 'finance/vat_report.html', {
        'title': 'VAT Report (FTA)',
        'vat_returns': vat_returns,
        
        # Box 1: Standard Rated Supplies
        'standard_rated_supplies': standard_rated_supplies,
        'standard_rated_vat': standard_rated_vat,
        
        # Box 2-4: Zero rated, Exempt, Out of scope (not yet implemented)
        'zero_rated_supplies': Decimal('0.00'),
        'exempt_supplies': Decimal('0.00'),
        'out_of_scope': Decimal('0.00'),
        
        # Box 6-7: Standard Rated Expenses
        'standard_rated_expenses': standard_rated_expenses,
        'input_vat': current_input_vat,
        
        # Totals
        'total_sales': current_sales,
        'total_output_vat': current_output_vat,
        'total_purchases': current_purchases,
        'total_input_vat': current_input_vat,
        
        # Box 10: Net VAT
        'net_vat': current_net_vat,
        'is_refund': current_net_vat < 0,
        
        # Date range
        'start_date': start_date,
        'end_date': end_date,
    })


@login_required
def corporate_tax_report(request):
    """
    UAE Corporate Tax Computation Report.
    SINGLE SOURCE OF TRUTH: Reads only from JournalEntryLine for P&L accounts.
    
    UAE Corporate Tax Law (Federal Decree-Law No. 47 of 2022):
    - Tax Rate: 9% on profit exceeding AED 375,000
    - Accounting Profit → Adjustments → Taxable Income → Tax
    """
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'view')):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard')
    
    # Get fiscal year
    fiscal_year_id = request.GET.get('fiscal_year')
    fiscal_years = FiscalYear.objects.filter(is_active=True).order_by('-start_date')
    
    # Default to current year
    if fiscal_year_id:
        selected_fiscal_year = FiscalYear.objects.filter(pk=fiscal_year_id).first()
    else:
        selected_fiscal_year = fiscal_years.first()
    
    # Get date range from fiscal year
    if selected_fiscal_year:
        start_date = selected_fiscal_year.start_date.isoformat()
        end_date = selected_fiscal_year.end_date.isoformat()
    else:
        start_date = date(date.today().year, 1, 1).isoformat()
        end_date = date.today().isoformat()
    
    # Get existing computations
    tax_computations = CorporateTaxComputation.objects.filter(is_active=True).select_related(
        'fiscal_year', 'journal_entry', 'payment_journal_entry'
    ).order_by('-fiscal_year__start_date')
    
    # Calculate current year tax from Journal Lines (SINGLE SOURCE OF TRUTH)
    income_accounts = Account.objects.filter(is_active=True, account_type=AccountType.INCOME)
    expense_accounts = Account.objects.filter(is_active=True, account_type=AccountType.EXPENSE)
    
    # Revenue from journal lines (Credits to Income accounts)
    income_lines = JournalEntryLine.objects.filter(
        account__in=income_accounts,
        journal_entry__status='posted',
        journal_entry__date__gte=start_date,
        journal_entry__date__lte=end_date,
    ).aggregate(
        total_debit=Sum('debit'),
        total_credit=Sum('credit')
    )
    current_revenue = (income_lines['total_credit'] or Decimal('0.00')) - (income_lines['total_debit'] or Decimal('0.00'))
    
    # Expenses from journal lines (Debits to Expense accounts)
    expense_lines = JournalEntryLine.objects.filter(
        account__in=expense_accounts,
        journal_entry__status='posted',
        journal_entry__date__gte=start_date,
        journal_entry__date__lte=end_date,
    ).aggregate(
        total_debit=Sum('debit'),
        total_credit=Sum('credit')
    )
    current_expenses = (expense_lines['total_debit'] or Decimal('0.00')) - (expense_lines['total_credit'] or Decimal('0.00'))
    
    # Accounting profit (before adjustments)
    accounting_profit = current_revenue - current_expenses
    
    # UAE Corporate Tax (9% on profit > AED 375,000)
    tax_threshold = Decimal('375000.00')
    tax_rate = Decimal('9.00')
    
    # For quick calculation (without adjustments) - display only
    if accounting_profit > tax_threshold:
        taxable_amount = accounting_profit - tax_threshold
        tax_payable_estimate = (taxable_amount * tax_rate / 100).quantize(Decimal('0.01'))
    else:
        taxable_amount = Decimal('0.00')
        tax_payable_estimate = Decimal('0.00')
    
    # Get existing computation for selected fiscal year
    existing_computation = None
    if selected_fiscal_year:
        existing_computation = CorporateTaxComputation.objects.filter(
            fiscal_year=selected_fiscal_year, is_active=True
        ).first()
    
    # Excel Export
    export_format = request.GET.get('format', '')
    if export_format == 'excel' and selected_fiscal_year:
        from .excel_exports import export_corporate_tax
        tax_data = {
            'fiscal_year': selected_fiscal_year.name,
            'start_date': start_date,
            'end_date': end_date,
            'revenue': current_revenue,
            'expenses': current_expenses,
            'accounting_profit': accounting_profit,
            'tax_threshold': tax_threshold,
            'tax_rate': tax_rate,
            'taxable_amount': taxable_amount,
            'tax_payable': tax_payable_estimate,
            'computation': existing_computation,
        }
        return export_corporate_tax(tax_data)
    
    return render(request, 'finance/corporate_tax_report.html', {
        'title': 'Corporate Tax Report',
        'fiscal_years': fiscal_years,
        'selected_fiscal_year': selected_fiscal_year,
        'tax_computations': tax_computations,
        'existing_computation': existing_computation,
        'current_revenue': current_revenue,
        'current_expenses': current_expenses,
        'accounting_profit': accounting_profit,
        'tax_threshold': tax_threshold,
        'tax_rate': tax_rate,
        'taxable_amount': taxable_amount,
        'tax_payable_estimate': tax_payable_estimate,
        'start_date': start_date,
        'end_date': end_date,
        'can_create': request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'create'),
        'can_post': request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit'),
    })


@login_required
def corporate_tax_create(request):
    """
    Create Corporate Tax Computation with adjustments.
    """
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'create')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:corporate_tax_report')
    
    if request.method == 'POST':
        fiscal_year_id = request.POST.get('fiscal_year')
        fiscal_year = get_object_or_404(FiscalYear, pk=fiscal_year_id)
        
        # Check if already exists
        existing = CorporateTaxComputation.objects.filter(fiscal_year=fiscal_year, is_active=True).first()
        if existing:
            messages.error(request, f'Tax computation already exists for {fiscal_year.name}.')
            return redirect('finance:corporate_tax_report')
        
        # Get values from form
        try:
            revenue = Decimal(request.POST.get('revenue', '0'))
            expenses = Decimal(request.POST.get('expenses', '0'))
            non_deductible = Decimal(request.POST.get('non_deductible_expenses', '0'))
            exempt_income = Decimal(request.POST.get('exempt_income', '0'))
            other_adjustments = Decimal(request.POST.get('other_adjustments', '0'))
            notes = request.POST.get('notes', '')
        except:
            messages.error(request, 'Invalid amounts.')
            return redirect('finance:corporate_tax_report')
        
        # Create computation
        computation = CorporateTaxComputation.objects.create(
            fiscal_year=fiscal_year,
            revenue=revenue,
            expenses=expenses,
            non_deductible_expenses=non_deductible,
            exempt_income=exempt_income,
            other_adjustments=other_adjustments,
            notes=notes,
        )
        computation.calculate()
        
        messages.success(request, f'Corporate Tax computation created for {fiscal_year.name}. Tax Payable: AED {computation.tax_payable:,.2f}')
        return redirect('finance:corporate_tax_report')
    
    # GET - redirect to report
    return redirect('finance:corporate_tax_report')


@login_required
def corporate_tax_post_provision(request, pk):
    """
    Post Corporate Tax Provision Journal.
    Dr Corporate Tax Expense (P&L)
    Cr Corporate Tax Payable (Liability)
    """
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:corporate_tax_report')
    
    computation = get_object_or_404(CorporateTaxComputation, pk=pk)
    
    try:
        journal = computation.post_provision(user=request.user)
        messages.success(request, f'Tax provision posted. Journal: {journal.entry_number}')
    except ValidationError as e:
        messages.error(request, str(e))
    except Exception as e:
        messages.error(request, f'Error posting provision: {e}')
    
    return redirect('finance:corporate_tax_report')


@login_required
def corporate_tax_pay(request, pk):
    """
    Pay Corporate Tax.
    Dr Corporate Tax Payable (clears liability)
    Cr Bank
    """
    computation = get_object_or_404(CorporateTaxComputation, pk=pk)
    
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:corporate_tax_report')
    
    if request.method == 'POST':
        bank_account_id = request.POST.get('bank_account')
        payment_date = request.POST.get('payment_date')
        reference = request.POST.get('reference', '')
        
        bank_account = BankAccount.objects.filter(pk=bank_account_id, is_active=True).first()
        if not bank_account:
            messages.error(request, 'Invalid bank account.')
            return redirect('finance:corporate_tax_report')
        
        from datetime import datetime
        try:
            if payment_date:
                payment_date = datetime.strptime(payment_date, '%Y-%m-%d').date()
            else:
                payment_date = date.today()
        except ValueError:
            payment_date = date.today()
        
        try:
            journal = computation.post_payment(
                bank_account=bank_account,
                payment_date=payment_date,
                reference=reference,
                user=request.user
            )
            messages.success(request, f'Tax payment of AED {computation.paid_amount:,.2f} recorded. Journal: {journal.entry_number}')
        except ValidationError as e:
            messages.error(request, str(e))
        except Exception as e:
            messages.error(request, f'Error processing payment: {e}')
        
        return redirect('finance:corporate_tax_report')
    
    # GET - Show payment form
    bank_accounts = BankAccount.objects.filter(is_active=True)
    context = {
        'title': f'Pay Corporate Tax - {computation.fiscal_year.name}',
        'computation': computation,
        'bank_accounts': bank_accounts,
        'today': date.today().strftime('%Y-%m-%d'),
    }
    return render(request, 'finance/corporate_tax_pay.html', context)


@login_required
def journal_register(request):
    """
    Journal Register - Comprehensive read-only control & audit report.
    Lists ALL journal entries from all sources with full filtering capabilities.
    """
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'view')):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard')
    
    # Date filters
    start_date = request.GET.get('start_date', date(date.today().year, 1, 1).isoformat())
    end_date = request.GET.get('end_date', date.today().isoformat())
    
    # Base queryset with related data for performance
    entries = JournalEntry.objects.filter(
        is_active=True,
        date__gte=start_date,
        date__lte=end_date,
    ).select_related(
        'fiscal_year', 'period', 'posted_by', 'created_by', 'reversal_of'
    ).prefetch_related('lines', 'lines__account')
    
    # Filter by fiscal year
    fiscal_year_id = request.GET.get('fiscal_year')
    if fiscal_year_id:
        entries = entries.filter(fiscal_year_id=fiscal_year_id)
    
    # Filter by period
    period_id = request.GET.get('period')
    if period_id:
        entries = entries.filter(period_id=period_id)
    
    # Filter by status
    status = request.GET.get('status')
    if status:
        entries = entries.filter(status=status)
    
    # Filter by entry type (source module)
    entry_type = request.GET.get('entry_type')
    if entry_type:
        entries = entries.filter(entry_type=entry_type)
    
    # Filter by source module (using actual source_module field)
    source_module = request.GET.get('source_module')
    if source_module:
        # Direct filter on source_module field
        entries = entries.filter(source_module=source_module)
    
    # Filter by account (journals affecting a specific account)
    account_id = request.GET.get('account')
    if account_id:
        entries = entries.filter(lines__account_id=account_id).distinct()
    
    # Filter by created_by
    created_by = request.GET.get('created_by')
    if created_by:
        entries = entries.filter(created_by_id=created_by)
    
    # Search filter (reference, description, entry number)
    search = request.GET.get('search')
    if search:
        entries = entries.filter(
            Q(entry_number__icontains=search) |
            Q(reference__icontains=search) |
            Q(description__icontains=search)
        )
    
    # Amount range filters
    min_amount = request.GET.get('min_amount')
    if min_amount:
        try:
            entries = entries.filter(total_debit__gte=Decimal(min_amount))
        except:
            pass
    
    max_amount = request.GET.get('max_amount')
    if max_amount:
        try:
            entries = entries.filter(total_debit__lte=Decimal(max_amount))
        except:
            pass
    
    # Reversal only filter
    reversal_only = request.GET.get('reversal_only')
    if reversal_only == '1':
        entries = entries.filter(Q(entry_type='reversal') | Q(reversal_of__isnull=False))
    
    # Sorting
    sort_by = request.GET.get('sort', '-date')
    valid_sorts = ['date', '-date', 'entry_number', '-entry_number', 'total_debit', '-total_debit']
    if sort_by in valid_sorts:
        entries = entries.order_by(sort_by, '-created_at')
    else:
        entries = entries.order_by('-date', '-created_at')
    
    # Pagination
    paginator = Paginator(entries, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Calculate summary statistics
    all_entries = JournalEntry.objects.filter(
        is_active=True,
        date__gte=start_date,
        date__lte=end_date,
    )
    
    summary = {
        'total_entries': entries.count(),
        'total_debit': entries.aggregate(Sum('total_debit'))['total_debit__sum'] or Decimal('0.00'),
        'total_credit': entries.aggregate(Sum('total_credit'))['total_credit__sum'] or Decimal('0.00'),
        'posted_count': entries.filter(status='posted').count(),
        'draft_count': entries.filter(status='draft').count(),
        'reversed_count': entries.filter(status='reversed').count(),
    }
    
    # Prepare filter options
    fiscal_years = FiscalYear.objects.filter(is_active=True).order_by('-start_date')
    periods = AccountingPeriod.objects.filter(is_active=True).select_related('fiscal_year').order_by('-start_date')
    accounts = Account.objects.filter(is_active=True).order_by('code')
    users = User.objects.filter(is_active=True).order_by('first_name', 'last_name')
    
    status_choices = JournalEntry.STATUS_CHOICES
    entry_type_choices = JournalEntry.ENTRY_TYPE_CHOICES
    
    # Source module choices - match JournalEntry.SOURCE_MODULE_CHOICES
    source_module_choices = [
        ('manual', 'Manual Entry'),
        ('sales', 'Sales Invoice'),
        ('purchase', 'Vendor Bill'),
        ('payment', 'Payment'),
        ('bank_transfer', 'Bank Transfer'),
        ('expense_claim', 'Expense Claim'),
        ('payroll', 'Payroll'),
        ('inventory', 'Inventory'),
        ('fixed_asset', 'Fixed Asset'),
        ('project', 'Project'),
        ('pdc', 'PDC Cheque'),
        ('property', 'Property/Rent'),
        ('vat', 'VAT Adjustment'),
        ('corporate_tax', 'Corporate Tax'),
        ('petty_cash', 'Petty Cash'),
        ('opening_balance', 'Opening Balance'),
        ('year_end', 'Year-End Closing'),
    ]
    
    # Determine source module for each entry (for display) - using actual source_module field
    def get_source_info(entry):
        """Get source module info from entry's source_module field."""
        source_classes = {
            'manual': ('Manual Entry', 'bg-secondary'),
            'sales': ('Sales Invoice', 'bg-success'),
            'purchase': ('Vendor Bill', 'bg-primary'),
            'payment': ('Payment', 'bg-info'),
            'bank_transfer': ('Bank Transfer', 'bg-secondary'),
            'expense_claim': ('Expense Claim', 'bg-warning'),
            'payroll': ('Payroll', 'bg-purple'),
            'inventory': ('Inventory', 'bg-teal'),
            'fixed_asset': ('Fixed Asset', 'bg-dark'),
            'project': ('Project', 'bg-orange'),
            'pdc': ('PDC Cheque', 'bg-pink'),
            'property': ('Property/Rent', 'bg-cyan'),
            'vat': ('VAT Adjustment', 'bg-danger'),
            'corporate_tax': ('Corporate Tax', 'bg-dark'),
            'petty_cash': ('Petty Cash', 'bg-warning'),
            'opening_balance': ('Opening Balance', 'bg-info'),
            'year_end': ('Year-End Closing', 'bg-dark'),
        }
        
        source_module = entry.source_module or 'manual'
        label, css_class = source_classes.get(source_module, ('Unknown', 'bg-light'))
        
        # Override for special entry types
        if entry.entry_type == 'reversal' or entry.reversal_of:
            return (source_module, f'{label} (Reversal)', 'bg-warning')
        if entry.entry_type == 'opening':
            return ('opening_balance', 'Opening Balance', 'bg-info')
        
        return (source_module, label, css_class)
    
    # Add source info to entries
    entries_with_source = []
    for entry in page_obj:
        source_key, source_label, source_class = get_source_info(entry)
        entries_with_source.append({
            'entry': entry,
            'source_key': source_key,
            'source_label': source_label,
            'source_class': source_class,
        })
    
    context = {
        'title': 'Journal Register',
        'entries': entries_with_source,
        'page_obj': page_obj,
        'summary': summary,
        'start_date': start_date,
        'end_date': end_date,
        'fiscal_years': fiscal_years,
        'periods': periods,
        'accounts': accounts,
        'users': users,
        'status_choices': status_choices,
        'entry_type_choices': entry_type_choices,
        'source_module_choices': source_module_choices,
    }
    
    # Export functionality
    export_format = request.GET.get('export')
    if export_format:
        return journal_register_export(request, entries, export_format, start_date, end_date)
    
    return render(request, 'finance/journal_register.html', context)


def journal_register_export(request, entries, export_format, start_date, end_date):
    """Export journal register to various formats."""
    import csv
    from django.http import HttpResponse
    
    if export_format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="journal_register_{start_date}_to_{end_date}.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'Entry Number', 'Date', 'Fiscal Year', 'Period', 'Reference', 
            'Description', 'Total Debit', 'Total Credit', 'Status', 
            'Entry Type', 'Created By', 'Created At', 'Posted By', 'Posted At'
        ])
        
        for entry in entries:
            writer.writerow([
                entry.entry_number,
                entry.date,
                entry.fiscal_year.name if entry.fiscal_year else '',
                entry.period.name if entry.period else '',
                entry.reference,
                entry.description,
                entry.total_debit,
                entry.total_credit,
                entry.get_status_display(),
                entry.get_entry_type_display(),
                entry.created_by.get_full_name() if entry.created_by else '',
                entry.created_at,
                entry.posted_by.get_full_name() if entry.posted_by else '',
                entry.posted_date,
            ])
        
        return response
    
    elif export_format == 'excel':
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.utils import get_column_letter
            from io import BytesIO
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Journal Register'
            
            # Title
            ws.merge_cells('A1:N1')
            ws['A1'] = f'Journal Register: {start_date} to {end_date}'
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Headers
            headers = [
                'Entry Number', 'Date', 'Fiscal Year', 'Period', 'Reference', 
                'Description', 'Total Debit', 'Total Credit', 'Status', 
                'Entry Type', 'Created By', 'Created At', 'Posted By', 'Posted At'
            ]
            
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Data
            for row_num, entry in enumerate(entries, 4):
                ws.cell(row=row_num, column=1, value=entry.entry_number)
                ws.cell(row=row_num, column=2, value=entry.date)
                ws.cell(row=row_num, column=3, value=entry.fiscal_year.name if entry.fiscal_year else '')
                ws.cell(row=row_num, column=4, value=entry.period.name if entry.period else '')
                ws.cell(row=row_num, column=5, value=entry.reference)
                ws.cell(row=row_num, column=6, value=entry.description)
                ws.cell(row=row_num, column=7, value=float(entry.total_debit))
                ws.cell(row=row_num, column=8, value=float(entry.total_credit))
                ws.cell(row=row_num, column=9, value=entry.get_status_display())
                ws.cell(row=row_num, column=10, value=entry.get_entry_type_display())
                ws.cell(row=row_num, column=11, value=entry.created_by.get_full_name() if entry.created_by else '')
                ws.cell(row=row_num, column=12, value=entry.created_at.strftime('%Y-%m-%d %H:%M') if entry.created_at else '')
                ws.cell(row=row_num, column=13, value=entry.posted_by.get_full_name() if entry.posted_by else '')
                ws.cell(row=row_num, column=14, value=entry.posted_date.strftime('%Y-%m-%d %H:%M') if entry.posted_date else '')
            
            # Auto-width columns
            for col in range(1, 15):
                ws.column_dimensions[get_column_letter(col)].width = 15
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="journal_register_{start_date}_to_{end_date}.xlsx"'
            return response
            
        except ImportError:
            messages.error(request, 'Excel export requires openpyxl library.')
            return redirect('finance:journal_register')
    
    return redirect('finance:journal_register')


@login_required
def journal_register_detail(request, pk):
    """
    Journal Register Detail - Read-only drill-down view of a journal entry.
    Shows full journal details with audit trail.
    """
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'view')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:journal_register')
    
    entry = get_object_or_404(
        JournalEntry.objects.select_related(
            'fiscal_year', 'period', 'posted_by', 'created_by', 'reversal_of'
        ).prefetch_related('lines', 'lines__account', 'reversed_by'),
        pk=pk
    )
    
    # Determine source module
    ref = entry.reference.upper() if entry.reference else ''
    
    if entry.entry_type == 'opening':
        source_info = {'module': 'Opening Balance', 'class': 'bg-info'}
    elif entry.entry_type == 'reversal' or entry.reversal_of:
        source_info = {'module': 'Reversal', 'class': 'bg-warning'}
    elif ref.startswith('INV') or 'INVOICE' in ref:
        source_info = {'module': 'Sales Invoice', 'class': 'bg-success'}
    elif ref.startswith('BILL') or 'BILL' in ref:
        source_info = {'module': 'Purchase Bill', 'class': 'bg-primary'}
    elif ref.startswith('PR') or ref.startswith('PM') or 'PAYMENT' in ref:
        source_info = {'module': 'Payment', 'class': 'bg-purple'}
    elif ref.startswith('TRANSFER') or 'TRANSFER' in ref:
        source_info = {'module': 'Bank Transfer', 'class': 'bg-secondary'}
    elif ref.startswith('EXPENSE') or 'EXPENSE' in ref:
        source_info = {'module': 'Expense Claim', 'class': 'bg-orange'}
    elif ref.startswith('VAT') or 'VAT' in ref:
        source_info = {'module': 'VAT Adjustment', 'class': 'bg-danger'}
    elif ref.startswith('ADJ') or entry.entry_type == 'adjustment':
        source_info = {'module': 'Adjustment', 'class': 'bg-dark'}
    else:
        source_info = {'module': 'Manual Journal', 'class': 'bg-light text-dark'}
    
    # Get linked records
    linked_records = []
    
    # Check for linked invoice
    if hasattr(entry, 'invoice') and entry.invoice.exists():
        for inv in entry.invoice.all():
            linked_records.append({
                'type': 'Sales Invoice',
                'number': inv.invoice_number,
                'url': f'/sales/invoices/{inv.pk}/',
            })
    
    # Check for linked payments
    if hasattr(entry, 'payments') and entry.payments.exists():
        for payment in entry.payments.all():
            linked_records.append({
                'type': 'Payment',
                'number': payment.payment_number,
                'url': f'/finance/payments/{payment.pk}/edit/',
            })
    
    # Check for linked bank transfers
    if hasattr(entry, 'bank_transfers') and entry.bank_transfers.exists():
        for transfer in entry.bank_transfers.all():
            linked_records.append({
                'type': 'Bank Transfer',
                'number': transfer.transfer_number,
                'url': f'/finance/bank-transfers/',
            })
    
    # Check for linked expense claims
    if hasattr(entry, 'expense_claims') and entry.expense_claims.exists():
        for claim in entry.expense_claims.all():
            linked_records.append({
                'type': 'Expense Claim',
                'number': claim.claim_number,
                'url': f'/finance/expense-claims/{claim.pk}/',
            })
    
    # Check for opening balance entry
    if hasattr(entry, 'opening_balance_entry') and entry.opening_balance_entry.exists():
        for ob in entry.opening_balance_entry.all():
            linked_records.append({
                'type': 'Opening Balance',
                'number': ob.entry_number,
                'url': f'/finance/opening-balances/{ob.pk}/',
            })
    
    # Period lock status
    is_period_locked = entry.period.is_locked if entry.period else False
    
    context = {
        'title': f'Journal Entry: {entry.entry_number}',
        'entry': entry,
        'source_info': source_info,
        'linked_records': linked_records,
        'is_period_locked': is_period_locked,
        'can_edit': (
            entry.status == 'draft' and 
            not is_period_locked and
            (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit'))
        ),
    }
    
    return render(request, 'finance/journal_register_detail.html', context)


# ============ FISCAL YEAR & PERIOD VIEWS ============

class FiscalYearListView(PermissionRequiredMixin, ListView):
    model = FiscalYear
    template_name = 'finance/fiscalyear_list.html'
    context_object_name = 'fiscal_years'
    module_name = 'finance'
    permission_type = 'view'
    
    def get_queryset(self):
        return FiscalYear.objects.filter(is_active=True)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Fiscal Years'
        context['form'] = FiscalYearForm()
        context['can_create'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'create')
        context['can_edit'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'edit')
        context['today'] = date.today().isoformat()
        return context
    
    def post(self, request, *args, **kwargs):
        if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'create')):
            messages.error(request, 'Permission denied.')
            return redirect('finance:fiscalyear_list')
        
        form = FiscalYearForm(request.POST)
        if form.is_valid():
            fy = form.save()
            messages.success(request, f'Fiscal Year {fy.name} created.')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
        return redirect('finance:fiscalyear_list')


@login_required
def fiscalyear_close(request, pk):
    """Close a fiscal year."""
    fy = get_object_or_404(FiscalYear, pk=pk)
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:fiscalyear_list')
    
    if fy.is_closed:
        messages.error(request, 'Fiscal year is already closed.')
        return redirect('finance:fiscalyear_list')
    
    fy.close(request.user)
    messages.success(request, f'Fiscal Year {fy.name} closed.')
    return redirect('finance:fiscalyear_list')


class AccountingPeriodListView(PermissionRequiredMixin, ListView):
    model = AccountingPeriod
    template_name = 'finance/period_list.html'
    context_object_name = 'periods'
    module_name = 'finance'
    permission_type = 'view'
    
    def get_queryset(self):
        return AccountingPeriod.objects.filter(is_active=True).select_related('fiscal_year')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Accounting Periods'
        context['form'] = AccountingPeriodForm()
        context['can_create'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'create')
        context['can_edit'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'edit')
        context['today'] = date.today().isoformat()
        return context
    
    def post(self, request, *args, **kwargs):
        if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'create')):
            messages.error(request, 'Permission denied.')
            return redirect('finance:period_list')
        
        form = AccountingPeriodForm(request.POST)
        if form.is_valid():
            period = form.save()
            messages.success(request, f'Period {period.name} created.')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
        return redirect('finance:period_list')


@login_required
def period_lock(request, pk):
    """Lock/unlock an accounting period."""
    from django.utils import timezone
    period = get_object_or_404(AccountingPeriod, pk=pk)
    
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:period_list')
    
    period.is_locked = not period.is_locked
    if period.is_locked:
        period.locked_date = timezone.now()
        period.locked_by = request.user
    else:
        period.locked_date = None
        period.locked_by = None
    period.save()
    
    status = 'locked' if period.is_locked else 'unlocked'
    messages.success(request, f'Period {period.name} {status}.')
    return redirect('finance:period_list')


# ============ BANK ACCOUNT VIEWS ============

class BankAccountListView(PermissionRequiredMixin, ListView):
    model = BankAccount
    template_name = 'finance/bankaccount_list.html'
    context_object_name = 'bank_accounts'
    module_name = 'finance'
    permission_type = 'view'
    
    def get_queryset(self):
        return BankAccount.objects.filter(is_active=True).select_related('gl_account')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Bank Accounts'
        context['form'] = BankAccountForm()
        context['can_create'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'create')
        context['can_edit'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'edit')
        
        # Update balances from GL
        for ba in context['bank_accounts']:
            ba.update_balance()
        
        return context
    
    def post(self, request, *args, **kwargs):
        if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'create')):
            messages.error(request, 'Permission denied.')
            return redirect('finance:bankaccount_list')
        
        form = BankAccountForm(request.POST)
        if form.is_valid():
            ba = form.save()
            messages.success(request, f'Bank Account {ba.name} created.')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
        return redirect('finance:bankaccount_list')


class BankAccountUpdateView(UpdatePermissionMixin, UpdateView):
    model = BankAccount
    form_class = BankAccountForm
    template_name = 'finance/bankaccount_form.html'
    success_url = reverse_lazy('finance:bankaccount_list')
    module_name = 'finance'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Edit Bank Account: {self.object.name}'
        return context
    
    def form_valid(self, form):
        messages.success(self.request, f'Bank Account {form.instance.name} updated.')
        return super().form_valid(form)


# ============ BANK TRANSFER VIEWS ============

class BankTransferListView(PermissionRequiredMixin, ListView):
    model = BankTransfer
    template_name = 'finance/banktransfer_list.html'
    context_object_name = 'transfers'
    module_name = 'finance'
    permission_type = 'view'
    paginate_by = 25
    
    def get_queryset(self):
        return BankTransfer.objects.filter(is_active=True).select_related('from_bank', 'to_bank')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Bank Transfers'
        context['can_create'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'create')
        context['can_edit'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'edit')
        context['today'] = date.today().isoformat()
        return context


class BankTransferCreateView(CreatePermissionMixin, CreateView):
    model = BankTransfer
    form_class = BankTransferForm
    template_name = 'finance/banktransfer_form.html'
    success_url = reverse_lazy('finance:banktransfer_list')
    module_name = 'finance'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Create Bank Transfer'
        context['today'] = date.today().isoformat()
        return context
    
    def form_valid(self, form):
        messages.success(self.request, 'Bank Transfer created.')
        return super().form_valid(form)


@login_required
def banktransfer_confirm(request, pk):
    """Confirm a bank transfer and create journal entry."""
    transfer = get_object_or_404(BankTransfer, pk=pk)
    
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:banktransfer_list')
    
    if transfer.status != 'draft':
        messages.error(request, 'Only draft transfers can be confirmed.')
        return redirect('finance:banktransfer_list')
    
    try:
        transfer.confirm(request.user)
        messages.success(request, f'Bank Transfer {transfer.transfer_number} confirmed.')
    except Exception as e:
        messages.error(request, str(e))
    
    return redirect('finance:banktransfer_list')


# ============ EXPENSE CLAIM VIEWS ============

class ExpenseClaimListView(PermissionRequiredMixin, ListView):
    model = ExpenseClaim
    template_name = 'finance/expenseclaim_list.html'
    context_object_name = 'claims'
    module_name = 'finance'
    permission_type = 'view'
    paginate_by = 25
    
    def get_queryset(self):
        queryset = ExpenseClaim.objects.filter(is_active=True).select_related('employee')
        
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Expense Claims'
        context['status_choices'] = ExpenseClaim.STATUS_CHOICES
        context['can_create'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'create')
        context['can_edit'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'edit')
        context['can_approve'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'approve')
        context['today'] = date.today().isoformat()
        return context


class ExpenseClaimCreateView(CreatePermissionMixin, CreateView):
    model = ExpenseClaim
    form_class = ExpenseClaimForm
    template_name = 'finance/expenseclaim_form.html'
    success_url = reverse_lazy('finance:expenseclaim_list')
    module_name = 'finance'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Create Expense Claim'
        context['today'] = date.today().isoformat()
        if self.request.POST:
            context['items_formset'] = ExpenseItemFormSet(self.request.POST, self.request.FILES)
        else:
            context['items_formset'] = ExpenseItemFormSet()
        return context
    
    def form_valid(self, form):
        context = self.get_context_data()
        items_formset = context['items_formset']
        
        if items_formset.is_valid():
            form.instance.employee = self.request.user
            self.object = form.save()
            items_formset.instance = self.object
            items_formset.save()
            self.object.calculate_totals()
            messages.success(self.request, f'Expense Claim {self.object.claim_number} created.')
            return redirect(self.success_url)
        else:
            return self.render_to_response(self.get_context_data(form=form))


class ExpenseClaimDetailView(PermissionRequiredMixin, DetailView):
    model = ExpenseClaim
    template_name = 'finance/expenseclaim_detail.html'
    context_object_name = 'claim'
    module_name = 'finance'
    permission_type = 'view'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Expense Claim: {self.object.claim_number}'
        context['can_approve'] = (
            self.request.user.is_superuser or 
            PermissionChecker.has_permission(self.request.user, 'finance', 'approve')
        ) and self.object.status == 'submitted'
        return context


@login_required
def expenseclaim_submit(request, pk):
    """Submit expense claim for approval."""
    claim = get_object_or_404(ExpenseClaim, pk=pk)
    
    if claim.status != 'draft':
        messages.error(request, 'Only draft claims can be submitted.')
        return redirect('finance:expenseclaim_list')
    
    claim.status = 'submitted'
    claim.save()
    messages.success(request, f'Expense Claim {claim.claim_number} submitted for approval.')
    return redirect('finance:expenseclaim_list')


@login_required
def expenseclaim_approve(request, pk):
    """Approve an expense claim."""
    from django.utils import timezone
    claim = get_object_or_404(ExpenseClaim, pk=pk)
    
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'approve')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:expenseclaim_list')
    
    if claim.status != 'submitted':
        messages.error(request, 'Only submitted claims can be approved.')
        return redirect('finance:expenseclaim_list')
    
    claim.status = 'approved'
    claim.approved_by = request.user
    claim.approved_date = timezone.now()
    claim.save()
    
    messages.success(request, f'Expense Claim {claim.claim_number} approved.')
    return redirect('finance:expenseclaim_list')


# ============ BUDGET VIEWS ============

class BudgetListView(PermissionRequiredMixin, ListView):
    model = Budget
    template_name = 'finance/budget_list.html'
    context_object_name = 'budgets'
    module_name = 'finance'
    permission_type = 'view'
    
    def get_queryset(self):
        return Budget.objects.filter(is_active=True).select_related('fiscal_year')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Budgets'
        context['can_create'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'create')
        context['can_edit'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'edit')
        return context


class BudgetCreateView(CreatePermissionMixin, CreateView):
    model = Budget
    form_class = BudgetForm
    template_name = 'finance/budget_form.html'
    success_url = reverse_lazy('finance:budget_list')
    module_name = 'finance'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Create Budget'
        if self.request.POST:
            context['lines_formset'] = BudgetLineFormSet(self.request.POST)
        else:
            context['lines_formset'] = BudgetLineFormSet()
        return context
    
    def form_valid(self, form):
        context = self.get_context_data()
        lines_formset = context['lines_formset']
        
        if lines_formset.is_valid():
            self.object = form.save()
            lines_formset.instance = self.object
            lines_formset.save()
            messages.success(self.request, f'Budget {self.object.name} created.')
            return redirect(self.success_url)
        else:
            return self.render_to_response(self.get_context_data(form=form))


class BudgetDetailView(PermissionRequiredMixin, DetailView):
    model = Budget
    template_name = 'finance/budget_detail.html'
    context_object_name = 'budget'
    module_name = 'finance'
    permission_type = 'view'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Budget: {self.object.name}'
        context['can_edit'] = (
            self.request.user.is_superuser or 
            PermissionChecker.has_permission(self.request.user, 'finance', 'edit')
        ) and self.object.status == 'draft'
        return context


# ============ VAT RETURN VIEWS ============

class VATReturnListView(PermissionRequiredMixin, ListView):
    model = VATReturn
    template_name = 'finance/vatreturn_list.html'
    context_object_name = 'vat_returns'
    module_name = 'finance'
    permission_type = 'view'
    
    def get_queryset(self):
        return VATReturn.objects.filter(is_active=True)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'VAT Returns'
        context['can_create'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'create')
        context['can_edit'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'edit')
        return context


class VATReturnCreateView(CreatePermissionMixin, CreateView):
    model = VATReturn
    form_class = VATReturnForm
    template_name = 'finance/vatreturn_form.html'
    success_url = reverse_lazy('finance:vatreturn_list')
    module_name = 'finance'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Create VAT Return'
        context['today'] = date.today().isoformat()
        return context
    
    def form_valid(self, form):
        self.object = form.save()
        self.object.calculate()
        messages.success(self.request, f'VAT Return {self.object.return_number} created.')
        return redirect(self.success_url)


class VATReturnDetailView(PermissionRequiredMixin, DetailView):
    model = VATReturn
    template_name = 'finance/vatreturn_detail.html'
    context_object_name = 'vat_return'
    module_name = 'finance'
    permission_type = 'view'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'VAT Return: {self.object.return_number}'
        context['can_edit'] = (
            self.request.user.is_superuser or 
            PermissionChecker.has_permission(self.request.user, 'finance', 'edit')
        ) and self.object.status == 'draft'
        return context


# ============ ADDITIONAL REPORTS ============

@login_required
def cash_flow(request):
    """Cash Flow Statement - Simplified."""
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'view')):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard')
    
    start_date = request.GET.get('start_date', date(date.today().year, 1, 1).isoformat())
    end_date = request.GET.get('end_date', date.today().isoformat())
    
    # Get all bank/cash accounts
    cash_accounts = Account.objects.filter(
        is_active=True,
        account_type=AccountType.ASSET,
        code__startswith='1'  # Assuming cash/bank accounts start with 1
    ).order_by('code')
    
    opening_cash = sum(acc.opening_balance for acc in cash_accounts)
    closing_cash = sum(acc.current_balance for acc in cash_accounts)
    
    # Operating activities (simplified - Income - Expenses)
    income = Account.objects.filter(
        is_active=True, account_type=AccountType.INCOME
    ).aggregate(total=Sum('balance'))['total'] or Decimal('0.00')
    
    expenses = Account.objects.filter(
        is_active=True, account_type=AccountType.EXPENSE
    ).aggregate(total=Sum('balance'))['total'] or Decimal('0.00')
    
    operating_activities = abs(income) - expenses
    
    # Net change in cash
    net_change = closing_cash - opening_cash
    
    # Prepare data for display and export
    operating = [
        {'description': 'Net Income', 'amount': abs(income)},
        {'description': 'Less: Total Expenses', 'amount': -expenses},
    ]
    investing = []  # Simplified - no investing activities tracked
    financing = []  # Simplified - no financing activities tracked
    
    # Excel Export
    export_format = request.GET.get('format', '')
    if export_format == 'excel':
        from .excel_exports import export_cash_flow
        return export_cash_flow(operating, investing, financing, start_date, end_date)
    
    return render(request, 'finance/cash_flow.html', {
        'title': 'Cash Flow Statement',
        'opening_cash': opening_cash,
        'closing_cash': closing_cash,
        'operating_activities': operating_activities,
        'net_change': net_change,
        'start_date': start_date,
        'end_date': end_date,
    })


@login_required
def ar_aging(request):
    """
    Accounts Receivable Aging Report.
    SINGLE SOURCE OF TRUTH: Reads only from JournalEntryLine (AR account).
    """
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'view')):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard')
    
    # Support date filter
    as_of_date_str = request.GET.get('date', '')
    if as_of_date_str:
        try:
            today = date.fromisoformat(as_of_date_str)
        except ValueError:
            today = date.today()
    else:
        today = date.today()
    
    # Get AR account (typically 1200 or similar)
    ar_account = Account.objects.filter(
        code__startswith='12', account_type=AccountType.ASSET, is_active=True
    ).first()
    
    if not ar_account:
        messages.warning(request, 'Accounts Receivable account not found in Chart of Accounts.')
        return render(request, 'finance/ar_aging.html', {
            'title': 'Accounts Receivable Aging',
            'aging_data': {'current': [], '1_30': [], '31_60': [], '61_90': [], 'over_90': []},
            'totals': {'current': Decimal('0.00'), '1_30': Decimal('0.00'), '31_60': Decimal('0.00'), '61_90': Decimal('0.00'), 'over_90': Decimal('0.00')},
            'grand_total': Decimal('0.00'),
            'as_of_date': today,
        })
    
    # Get all open AR items (invoices) from journal lines
    # Open items = Journal lines where there's a debit to AR without matching credit (payment)
    ar_lines = JournalEntryLine.objects.filter(
        account=ar_account,
        journal_entry__status='posted',
        debit__gt=0,
    ).select_related('journal_entry').order_by('journal_entry__date')
    
    # Build aging data from AR journal lines
    aging_data = {
        'current': [],
        '1_30': [],
        '31_60': [],
        '61_90': [],
        'over_90': [],
    }
    
    totals = {
        'current': Decimal('0.00'),
        '1_30': Decimal('0.00'),
        '31_60': Decimal('0.00'),
        '61_90': Decimal('0.00'),
        'over_90': Decimal('0.00'),
    }
    
    # Group by reference (invoice number) and calculate net outstanding
    invoice_balances = {}
    for line in ar_lines:
        ref = line.journal_entry.reference
        if ref not in invoice_balances:
            invoice_balances[ref] = {
                'date': line.journal_entry.date,
                'reference': ref,
                'description': line.description or line.journal_entry.description,
                'debit': Decimal('0.00'),
                'credit': Decimal('0.00'),
            }
        invoice_balances[ref]['debit'] += line.debit
    
    # Get credits (payments) to AR
    ar_credits = JournalEntryLine.objects.filter(
        account=ar_account,
        journal_entry__status='posted',
        credit__gt=0,
    ).select_related('journal_entry')
    
    for line in ar_credits:
        ref = line.journal_entry.reference
        if ref in invoice_balances:
            invoice_balances[ref]['credit'] += line.credit
    
    # Calculate aging
    for ref, data in invoice_balances.items():
        outstanding = data['debit'] - data['credit']
        if outstanding <= 0:
            continue  # Fully paid
        
        days_old = (today - data['date']).days
        
        item_data = {
            'reference': ref,
            'description': data['description'],
            'date': data['date'],
            'total': data['debit'],
            'paid': data['credit'],
            'outstanding': outstanding,
            'days_old': days_old,
        }
        
        if days_old <= 30:
            aging_data['current'].append(item_data)
            totals['current'] += outstanding
        elif days_old <= 60:
            aging_data['1_30'].append(item_data)
            totals['1_30'] += outstanding
        elif days_old <= 90:
            aging_data['31_60'].append(item_data)
            totals['31_60'] += outstanding
        elif days_old <= 120:
            aging_data['61_90'].append(item_data)
            totals['61_90'] += outstanding
        else:
            aging_data['over_90'].append(item_data)
            totals['over_90'] += outstanding
    
    grand_total = sum(totals.values())
    
    # Excel Export
    export_format = request.GET.get('format', '')
    if export_format == 'excel':
        from .excel_exports import export_ar_aging
        # Flatten data for export
        customer_data = []
        for bucket, items in aging_data.items():
            for item in items:
                customer_data.append({
                    'name': item['reference'],
                    'current': item['outstanding'] if bucket == 'current' else 0,
                    'days_1_30': item['outstanding'] if bucket == '1_30' else 0,
                    'days_31_60': item['outstanding'] if bucket == '31_60' else 0,
                    'days_61_90': item['outstanding'] if bucket == '61_90' else 0,
                    'days_over_90': item['outstanding'] if bucket == 'over_90' else 0,
                    'total': item['outstanding'],
                })
        return export_ar_aging(customer_data, today.isoformat())
    
    return render(request, 'finance/ar_aging.html', {
        'title': 'Accounts Receivable Aging',
        'aging_data': aging_data,
        'totals': totals,
        'grand_total': grand_total,
        'as_of_date': today,
        'ar_account': ar_account,
    })


@login_required
def ap_aging(request):
    """
    Accounts Payable Aging Report.
    SINGLE SOURCE OF TRUTH: Reads only from JournalEntryLine (AP account).
    """
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'view')):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard')
    
    # Support date filter
    as_of_date_str = request.GET.get('date', '')
    if as_of_date_str:
        try:
            today = date.fromisoformat(as_of_date_str)
        except ValueError:
            today = date.today()
    else:
        today = date.today()
    
    # Get AP account (typically 2000 or similar)
    ap_account = Account.objects.filter(
        code__startswith='20', account_type=AccountType.LIABILITY, is_active=True
    ).first()
    
    if not ap_account:
        messages.warning(request, 'Accounts Payable account not found in Chart of Accounts.')
        return render(request, 'finance/ap_aging.html', {
            'title': 'Accounts Payable Aging',
            'aging_data': {'current': [], '1_30': [], '31_60': [], '61_90': [], 'over_90': []},
            'totals': {'current': Decimal('0.00'), '1_30': Decimal('0.00'), '31_60': Decimal('0.00'), '61_90': Decimal('0.00'), 'over_90': Decimal('0.00')},
            'grand_total': Decimal('0.00'),
            'as_of_date': today,
        })
    
    # Get all open AP items (bills) from journal lines
    # Open items = Journal lines where there's a credit to AP without matching debit (payment)
    ap_lines = JournalEntryLine.objects.filter(
        account=ap_account,
        journal_entry__status='posted',
        credit__gt=0,
    ).select_related('journal_entry').order_by('journal_entry__date')
    
    # Build aging data from AP journal lines
    aging_data = {
        'current': [],
        '1_30': [],
        '31_60': [],
        '61_90': [],
        'over_90': [],
    }
    
    totals = {
        'current': Decimal('0.00'),
        '1_30': Decimal('0.00'),
        '31_60': Decimal('0.00'),
        '61_90': Decimal('0.00'),
        'over_90': Decimal('0.00'),
    }
    
    # Group by reference (bill number) and calculate net outstanding
    bill_balances = {}
    for line in ap_lines:
        ref = line.journal_entry.reference
        if ref not in bill_balances:
            bill_balances[ref] = {
                'date': line.journal_entry.date,
                'reference': ref,
                'description': line.description or line.journal_entry.description,
                'debit': Decimal('0.00'),
                'credit': Decimal('0.00'),
            }
        bill_balances[ref]['credit'] += line.credit
    
    # Get debits (payments) to AP
    ap_debits = JournalEntryLine.objects.filter(
        account=ap_account,
        journal_entry__status='posted',
        debit__gt=0,
    ).select_related('journal_entry')
    
    for line in ap_debits:
        ref = line.journal_entry.reference
        if ref in bill_balances:
            bill_balances[ref]['debit'] += line.debit
    
    # Calculate aging
    for ref, data in bill_balances.items():
        outstanding = data['credit'] - data['debit']
        if outstanding <= 0:
            continue  # Fully paid
        
        days_old = (today - data['date']).days
        
        item_data = {
            'reference': ref,
            'description': data['description'],
            'date': data['date'],
            'total': data['credit'],
            'paid': data['debit'],
            'outstanding': outstanding,
            'days_old': days_old,
        }
        
        if days_old <= 30:
            aging_data['current'].append(item_data)
            totals['current'] += outstanding
        elif days_old <= 60:
            aging_data['1_30'].append(item_data)
            totals['1_30'] += outstanding
        elif days_old <= 90:
            aging_data['31_60'].append(item_data)
            totals['31_60'] += outstanding
        elif days_old <= 120:
            aging_data['61_90'].append(item_data)
            totals['61_90'] += outstanding
        else:
            aging_data['over_90'].append(item_data)
            totals['over_90'] += outstanding
    
    grand_total = sum(totals.values())
    
    # Excel Export
    export_format = request.GET.get('format', '')
    if export_format == 'excel':
        from .excel_exports import export_ap_aging
        # Flatten data for export
        vendor_data = []
        for bucket, items in aging_data.items():
            for item in items:
                vendor_data.append({
                    'name': item['reference'],
                    'current': item['outstanding'] if bucket == 'current' else 0,
                    'days_1_30': item['outstanding'] if bucket == '1_30' else 0,
                    'days_31_60': item['outstanding'] if bucket == '31_60' else 0,
                    'days_61_90': item['outstanding'] if bucket == '61_90' else 0,
                    'days_over_90': item['outstanding'] if bucket == 'over_90' else 0,
                    'total': item['outstanding'],
                })
        return export_ap_aging(vendor_data, today.isoformat())
    
    return render(request, 'finance/ap_aging.html', {
        'title': 'Accounts Payable Aging',
        'aging_data': aging_data,
        'totals': totals,
        'grand_total': grand_total,
        'as_of_date': today,
        'ap_account': ap_account,
    })


@login_required
def bank_ledger(request):
    """Bank Ledger - Transactions for a specific bank account."""
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'view')):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard')
    
    bank_id = request.GET.get('bank')
    start_date = request.GET.get('start_date', date(date.today().year, 1, 1).isoformat())
    end_date = request.GET.get('end_date', date.today().isoformat())
    
    bank_accounts = BankAccount.objects.filter(is_active=True)
    selected_bank = None
    transactions = []
    running_balance = Decimal('0.00')
    
    if bank_id:
        selected_bank = get_object_or_404(BankAccount, pk=bank_id)
        gl_account = selected_bank.gl_account
        running_balance = gl_account.opening_balance
        
        lines = JournalEntryLine.objects.filter(
            account=gl_account,
            journal_entry__status='posted',
            journal_entry__date__gte=start_date,
            journal_entry__date__lte=end_date,
        ).select_related('journal_entry').order_by('journal_entry__date', 'id')
        
        for line in lines:
            running_balance += line.debit - line.credit
            
            transactions.append({
                'date': line.journal_entry.date,
                'journal_pk': line.journal_entry.pk,
                'entry_number': line.journal_entry.entry_number,
                'reference': line.journal_entry.reference,
                'description': line.description or line.journal_entry.description,
                'debit': line.debit,
                'credit': line.credit,
                'balance': running_balance,
            })
    
    # Excel Export
    export_format = request.GET.get('format', '')
    if export_format == 'excel' and selected_bank:
        from .excel_exports import export_bank_ledger
        return export_bank_ledger(transactions, selected_bank.name, start_date, end_date)
    
    return render(request, 'finance/bank_ledger.html', {
        'title': 'Bank Ledger',
        'bank_accounts': bank_accounts,
        'selected_bank': selected_bank,
        'transactions': transactions,
        'opening_balance': selected_bank.gl_account.opening_balance if selected_bank else Decimal('0.00'),
        'closing_balance': running_balance,
        'start_date': start_date,
        'end_date': end_date,
    })


@login_required
def budget_vs_actual(request):
    """Budget vs Actual Report."""
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'view')):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard')
    
    budget_id = request.GET.get('budget')
    budgets = Budget.objects.filter(is_active=True, status__in=['approved', 'locked'])
    
    selected_budget = None
    comparison_data = []
    total_budget = Decimal('0.00')
    total_actual = Decimal('0.00')
    
    if budget_id:
        selected_budget = get_object_or_404(Budget, pk=budget_id)
        
        for line in selected_budget.lines.all().select_related('account'):
            account = line.account
            actual = abs(account.balance) if account.balance else Decimal('0.00')
            budgeted = line.amount
            variance = budgeted - actual
            variance_pct = (variance / budgeted * 100) if budgeted else Decimal('0.00')
            
            comparison_data.append({
                'account': account,
                'budgeted': budgeted,
                'actual': actual,
                'variance': variance,
                'variance_pct': variance_pct,
                'is_over': actual > budgeted,
            })
            
            total_budget += budgeted
            total_actual += actual
    
    total_variance = total_budget - total_actual
    
    # Excel Export
    export_format = request.GET.get('format', '')
    if export_format == 'excel' and selected_budget:
        from .excel_exports import export_budget_vs_actual
        export_data = [{
            'account': f"{d['account'].code} - {d['account'].name}",
            'budget': d['budgeted'],
            'actual': d['actual'],
            'variance': d['variance'],
            'variance_pct': float(d['variance_pct']),
        } for d in comparison_data]
        return export_budget_vs_actual(export_data, selected_budget.name, selected_budget.fiscal_year.name)
    
    return render(request, 'finance/budget_vs_actual.html', {
        'title': 'Budget vs Actual',
        'budgets': budgets,
        'selected_budget': selected_budget,
        'comparison_data': comparison_data,
        'total_budget': total_budget,
        'total_actual': total_actual,
        'total_variance': total_variance,
    })


@login_required
def payment_post(request, pk):
    """
    Post a payment and create journal entry.
    Uses Account Mapping (SAP/Oracle-style Account Determination) for account selection.
    
    Payment Received: Dr Bank, Cr AR (clearing entry)
    Payment Made: Dr AP, Cr Bank (clearing entry)
    """
    payment = get_object_or_404(Payment, pk=pk)
    
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:payment_list')
    
    if payment.status != 'draft':
        messages.error(request, 'Only draft payments can be posted.')
        return redirect('finance:payment_list')
    
    if not payment.bank_account:
        messages.error(request, 'Bank account is required to post payment.')
        return redirect('finance:payment_list')
    
    # Create journal entry
    journal = JournalEntry.objects.create(
        date=payment.payment_date,
        reference=payment.payment_number,
        description=f"{payment.get_payment_type_display()}: {payment.party_name}",
        entry_type='standard',
        source_module='payment',
    )
    
    # Get AR/AP accounts using Account Mapping (SAP/Oracle standard)
    # Fallback to hardcoded codes for backward compatibility
    ar_account = AccountMapping.get_account_or_default('customer_receipt_ar_clear', '1200')
    if not ar_account:
        ar_account = Account.objects.filter(code__startswith='12', account_type='asset').first()
    
    ap_account = AccountMapping.get_account_or_default('vendor_payment_ap_clear', '2000')
    if not ap_account:
        ap_account = Account.objects.filter(code__startswith='20', account_type='liability').first()
    
    bank_account = payment.bank_account.gl_account
    
    if payment.payment_type == 'received':
        # Debit Bank, Credit AR (clears receivable)
        JournalEntryLine.objects.create(
            journal_entry=journal,
            account=bank_account,
            description=f"Payment from {payment.party_name}",
            debit=payment.amount,
        )
        if ar_account:
            JournalEntryLine.objects.create(
                journal_entry=journal,
                account=ar_account,
                description=f"Payment from {payment.party_name}",
                credit=payment.amount,
            )
        else:
            messages.warning(request, 'Accounts Receivable account not configured in Account Mapping.')
    else:
        # Debit AP (clears payable), Credit Bank
        if ap_account:
            JournalEntryLine.objects.create(
                journal_entry=journal,
                account=ap_account,
                description=f"Payment to {payment.party_name}",
                debit=payment.amount,
            )
        else:
            messages.warning(request, 'Accounts Payable account not configured in Account Mapping.')
        JournalEntryLine.objects.create(
            journal_entry=journal,
            account=bank_account,
            description=f"Payment to {payment.party_name}",
            credit=payment.amount,
        )
    
    journal.calculate_totals()
    
    try:
        journal.post(request.user)
        payment.journal_entry = journal
        payment.status = 'confirmed'
        payment.save()
        messages.success(request, f'Payment {payment.payment_number} posted successfully.')
    except Exception as e:
        journal.delete()
        messages.error(request, f'Failed to post payment: {e}')
    
    return redirect('finance:payment_list')


# ============ BANK STATEMENT VIEWS ============

class BankStatementListView(PermissionRequiredMixin, ListView):
    model = BankStatement
    template_name = 'finance/bankstatement_list.html'
    context_object_name = 'statements'
    module_name = 'finance'
    permission_type = 'view'
    paginate_by = 25
    
    def get_queryset(self):
        from .models import BankStatement
        queryset = BankStatement.objects.filter(is_active=True).select_related('bank_account')
        
        bank_id = self.request.GET.get('bank')
        if bank_id:
            queryset = queryset.filter(bank_account_id=bank_id)
        
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Bank Statements'
        context['bank_accounts'] = BankAccount.objects.filter(is_active=True)
        context['status_choices'] = BankStatement.STATUS_CHOICES
        context['can_create'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'create')
        context['can_edit'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'edit')
        return context


class BankStatementCreateView(CreatePermissionMixin, CreateView):
    model = BankStatement
    template_name = 'finance/bankstatement_form.html'
    success_url = reverse_lazy('finance:bankstatement_list')
    module_name = 'finance'
    fields = ['bank_account', 'statement_start_date', 'statement_end_date', 
              'opening_balance', 'closing_balance', 'notes']
    
    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        form.fields['bank_account'].widget.attrs['class'] = 'form-select'
        form.fields['bank_account'].queryset = BankAccount.objects.filter(is_active=True)
        for field_name in ['statement_start_date', 'statement_end_date']:
            form.fields[field_name].widget = forms.DateInput(attrs={'type': 'date', 'class': 'form-control'})
        for field_name in ['opening_balance', 'closing_balance']:
            form.fields[field_name].widget.attrs['class'] = 'form-control'
        form.fields['notes'].widget = forms.Textarea(attrs={'rows': 2, 'class': 'form-control'})
        return form
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Create Bank Statement'
        context['today'] = date.today().isoformat()
        return context
    
    def form_valid(self, form):
        messages.success(self.request, 'Bank Statement created successfully.')
        return super().form_valid(form)


@login_required
def bankstatement_template_download(request):
    """Download Excel template for bank statement import."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Bank Statement Lines"
    
    # Headers
    headers = [
        'Transaction Date (YYYY-MM-DD) *',
        'Description *',
        'Reference',
        'Debit Amount (Money Out)',
        'Credit Amount (Money In)',
        'Balance',
        'Value Date (YYYY-MM-DD)'
    ]
    
    # Style for headers
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = 22
    
    # Add sample data
    sample_data = [
        ['2026-01-05', 'Client Payment - INV-1001', 'TRF1001', '', '10500.00', '110500.00', '2026-01-05'],
        ['2026-01-07', 'Office Rent - January', 'CHQ789456', '6000.00', '', '104500.00', '2026-01-07'],
        ['2026-01-10', 'DEWA Utilities', 'DEWA012', '1200.00', '', '103300.00', '2026-01-10'],
        ['2026-01-12', 'Salary Transfer - January', 'SALJAN', '18000.00', '', '85300.00', '2026-01-12'],
        ['2026-01-15', 'Vendor Payment - Cloud Services', 'TRF2002', '3150.00', '', '82150.00', '2026-01-15'],
    ]
    
    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            if col_num in [4, 5, 6]:  # Amount columns
                cell.alignment = Alignment(horizontal='right')
    
    # Add instructions sheet
    ws2 = wb.create_sheet(title="Instructions")
    instructions = [
        ["Bank Statement Import Instructions"],
        [""],
        ["Required Fields (marked with *):"],
        ["- Transaction Date: Date in YYYY-MM-DD format (e.g., 2026-01-15)"],
        ["- Description: Transaction description from bank statement"],
        [""],
        ["Optional Fields:"],
        ["- Reference: Bank reference or cheque number"],
        ["- Debit Amount: Money going out (leave blank if credit)"],
        ["- Credit Amount: Money coming in (leave blank if debit)"],
        ["- Balance: Running balance after transaction"],
        ["- Value Date: Value date if different from transaction date"],
        [""],
        ["Notes:"],
        ["- Either Debit or Credit must have a value, not both"],
        ["- Delete sample rows before importing your actual data"],
        ["- Maximum 1000 rows per import"],
    ]
    
    for row_num, row_data in enumerate(instructions, 1):
        cell = ws2.cell(row=row_num, column=1, value=row_data[0] if row_data else "")
        if row_num == 1:
            cell.font = Font(bold=True, size=14)
        if "Required" in str(row_data) or "Optional" in str(row_data) or "Notes" in str(row_data):
            cell.font = Font(bold=True)
    
    ws2.column_dimensions['A'].width = 70
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=bank_statement_template.xlsx'
    wb.save(response)
    return response


@login_required
def bankstatement_import(request, pk):
    """Import bank statement lines from Excel file."""
    from openpyxl import load_workbook
    from decimal import Decimal, InvalidOperation
    
    statement = get_object_or_404(BankStatement, pk=pk)
    
    if request.method != 'POST':
        return redirect('finance:bankstatement_detail', pk=pk)
    
    excel_file = request.FILES.get('excel_file')
    if not excel_file:
        messages.error(request, 'Please select an Excel file to import.')
        return redirect('finance:bankstatement_detail', pk=pk)
    
    # Validate file extension
    if not excel_file.name.endswith(('.xlsx', '.xls')):
        messages.error(request, 'Please upload a valid Excel file (.xlsx or .xls).')
        return redirect('finance:bankstatement_detail', pk=pk)
    
    try:
        wb = load_workbook(excel_file, data_only=True)
        ws = wb.active
        
        imported_count = 0
        errors = []
        
        # Get existing max line number
        max_line = statement.lines.aggregate(max_line=models.Max('line_number'))['max_line'] or 0
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Skip empty rows
            if not row or not any(row):
                continue
            
            try:
                transaction_date = row[0]
                description = row[1]
                reference = row[2] or ''
                debit = row[3]
                credit = row[4]
                balance = row[5] if len(row) > 5 else None
                value_date = row[6] if len(row) > 6 else None
                
                # Validate required fields
                if not transaction_date:
                    errors.append(f"Row {row_num}: Transaction Date is required")
                    continue
                if not description:
                    errors.append(f"Row {row_num}: Description is required")
                    continue
                
                # Parse date
                if isinstance(transaction_date, str):
                    from datetime import datetime
                    transaction_date = datetime.strptime(transaction_date, '%Y-%m-%d').date()
                elif hasattr(transaction_date, 'date'):
                    transaction_date = transaction_date.date() if hasattr(transaction_date, 'date') else transaction_date
                
                # Parse value date
                if value_date:
                    if isinstance(value_date, str):
                        from datetime import datetime
                        value_date = datetime.strptime(value_date, '%Y-%m-%d').date()
                    elif hasattr(value_date, 'date'):
                        value_date = value_date.date() if hasattr(value_date, 'date') else value_date
                
                # Parse amounts
                debit_amount = Decimal(str(debit or 0).replace(',', ''))
                credit_amount = Decimal(str(credit or 0).replace(',', ''))
                balance_amount = Decimal(str(balance or 0).replace(',', '')) if balance else Decimal('0.00')
                
                if debit_amount == 0 and credit_amount == 0:
                    errors.append(f"Row {row_num}: Either Debit or Credit amount is required")
                    continue
                
                # Create statement line
                max_line += 1
                BankStatementLine.objects.create(
                    statement=statement,
                    line_number=max_line,
                    transaction_date=transaction_date,
                    value_date=value_date,
                    description=str(description)[:500],
                    reference=str(reference)[:200],
                    debit=debit_amount,
                    credit=credit_amount,
                    balance=balance_amount,
                    reconciliation_status='unmatched'
                )
                imported_count += 1
                
            except (ValueError, InvalidOperation) as e:
                errors.append(f"Row {row_num}: Invalid data format - {str(e)}")
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
        
        # Update statement totals
        statement.total_debits = statement.lines.aggregate(total=models.Sum('debit'))['total'] or Decimal('0.00')
        statement.total_credits = statement.lines.aggregate(total=models.Sum('credit'))['total'] or Decimal('0.00')
        statement.save()
        
        if imported_count > 0:
            messages.success(request, f'Successfully imported {imported_count} transaction(s).')
        
        if errors:
            error_msg = f'{len(errors)} error(s) during import. First 5: ' + '; '.join(errors[:5])
            messages.warning(request, error_msg)
        
        if imported_count == 0 and not errors:
            messages.warning(request, 'No data found in the Excel file. Make sure to use the template format.')
            
    except Exception as e:
        messages.error(request, f'Error reading Excel file: {str(e)}')
    
    return redirect('finance:bankstatement_detail', pk=pk)


class BankStatementDetailView(PermissionRequiredMixin, DetailView):
    """Bank Statement detail - Main reconciliation interface."""
    model = BankStatement
    template_name = 'finance/bankstatement_detail.html'
    context_object_name = 'statement'
    module_name = 'finance'
    permission_type = 'view'
    
    def get_context_data(self, **kwargs):
        from .forms import AdjustmentForm
        context = super().get_context_data(**kwargs)
        context['title'] = f'Bank Statement: {self.object.statement_number}'
        
        # Statement lines
        context['lines'] = self.object.lines.all().select_related(
            'matched_payment', 'matched_journal_line', 'adjustment_journal'
        )
        
        # Unmatched payments for this bank account
        context['unmatched_payments'] = Payment.objects.filter(
            bank_account=self.object.bank_account,
            status='confirmed',
        ).exclude(
            id__in=BankStatementLine.objects.filter(
                matched_payment__isnull=False
            ).values_list('matched_payment_id', flat=True)
        )
        
        # Unmatched journal entries for the bank GL account
        context['unmatched_journals'] = JournalEntryLine.objects.filter(
            account=self.object.bank_account.gl_account,
            journal_entry__status='posted',
        ).exclude(
            id__in=BankStatementLine.objects.filter(
                matched_journal_line__isnull=False
            ).values_list('matched_journal_line_id', flat=True)
        ).select_related('journal_entry')
        
        # Expense/Income accounts for adjustments
        context['expense_accounts'] = Account.objects.filter(
            is_active=True, account_type__in=['expense', 'income']
        ).order_by('account_type', 'code')
        
        # Permissions
        context['can_edit'] = (
            self.request.user.is_superuser or 
            PermissionChecker.has_permission(self.request.user, 'finance', 'edit')
        ) and self.object.status not in ['reconciled', 'locked']
        
        context['can_finalize'] = context['can_edit'] and self.object.unmatched_count == 0
        
        return context


@login_required
def bankstatement_import(request, pk):
    """Import bank statement lines from CSV."""
    import csv
    from io import TextIOWrapper
    from .models import BankStatement, BankStatementLine
    
    statement = get_object_or_404(BankStatement, pk=pk)
    
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:bankstatement_detail', pk=pk)
    
    if statement.status in ['reconciled', 'locked']:
        messages.error(request, 'Cannot import to reconciled/locked statements.')
        return redirect('finance:bankstatement_detail', pk=pk)
    
    if request.method == 'POST':
        csv_file = request.FILES.get('csv_file')
        if not csv_file:
            messages.error(request, 'Please upload a CSV file.')
            return redirect('finance:bankstatement_detail', pk=pk)
        
        try:
            # Clear existing lines
            statement.lines.all().delete()
            
            # Parse CSV
            decoded_file = TextIOWrapper(csv_file.file, encoding='utf-8-sig')
            reader = csv.DictReader(decoded_file)
            
            line_number = 1
            for row in reader:
                # Expected columns: Date, Description, Reference, Debit, Credit, Balance
                BankStatementLine.objects.create(
                    statement=statement,
                    line_number=line_number,
                    transaction_date=row.get('Date', row.get('date', date.today())),
                    description=row.get('Description', row.get('description', '')),
                    reference=row.get('Reference', row.get('reference', '')),
                    debit=Decimal(row.get('Debit', row.get('debit', '0')) or '0'),
                    credit=Decimal(row.get('Credit', row.get('credit', '0')) or '0'),
                    balance=Decimal(row.get('Balance', row.get('balance', '0')) or '0'),
                )
                line_number += 1
            
            statement.calculate_totals()
            statement.status = 'in_progress'
            statement.save()
            
            messages.success(request, f'{line_number - 1} lines imported successfully.')
        except Exception as e:
            messages.error(request, f'Import failed: {e}')
    
    return redirect('finance:bankstatement_detail', pk=pk)


@login_required
def bankstatement_add_line(request, pk):
    """Manually add a line to bank statement."""
    from .models import BankStatement, BankStatementLine
    
    statement = get_object_or_404(BankStatement, pk=pk)
    
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:bankstatement_detail', pk=pk)
    
    if statement.status in ['reconciled', 'locked']:
        messages.error(request, 'Cannot add lines to reconciled/locked statements.')
        return redirect('finance:bankstatement_detail', pk=pk)
    
    if request.method == 'POST':
        try:
            last_line = statement.lines.order_by('-line_number').first()
            next_line_number = (last_line.line_number + 1) if last_line else 1
            
            debit = Decimal(request.POST.get('debit', '0') or '0')
            credit = Decimal(request.POST.get('credit', '0') or '0')
            
            BankStatementLine.objects.create(
                statement=statement,
                line_number=next_line_number,
                transaction_date=request.POST.get('transaction_date'),
                description=request.POST.get('description', ''),
                reference=request.POST.get('reference', ''),
                debit=debit,
                credit=credit,
            )
            
            statement.calculate_totals()
            if statement.status == 'draft':
                statement.status = 'in_progress'
                statement.save()
            
            messages.success(request, 'Line added successfully.')
        except Exception as e:
            messages.error(request, f'Failed to add line: {e}')
    
    return redirect('finance:bankstatement_detail', pk=pk)


@login_required
def bankstatement_auto_match(request, pk):
    """Auto-match statement lines with accounting records."""
    from .models import BankStatement
    
    statement = get_object_or_404(BankStatement, pk=pk)
    
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:bankstatement_detail', pk=pk)
    
    if statement.status in ['reconciled', 'locked']:
        messages.error(request, 'Cannot modify reconciled/locked statements.')
        return redirect('finance:bankstatement_detail', pk=pk)
    
    date_tolerance = int(request.GET.get('tolerance', 3))
    matched_count = statement.auto_match(date_tolerance=date_tolerance)
    
    if matched_count > 0:
        messages.success(request, f'{matched_count} lines matched automatically.')
    else:
        messages.info(request, 'No automatic matches found.')
    
    return redirect('finance:bankstatement_detail', pk=pk)


@login_required
def bankstatement_manual_match(request, pk, line_id):
    """Manually match a statement line with a payment or journal entry."""
    from .models import BankStatement, BankStatementLine
    
    statement = get_object_or_404(BankStatement, pk=pk)
    line = get_object_or_404(BankStatementLine, pk=line_id, statement=statement)
    
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:bankstatement_detail', pk=pk)
    
    if statement.status in ['reconciled', 'locked']:
        messages.error(request, 'Cannot modify reconciled/locked statements.')
        return redirect('finance:bankstatement_detail', pk=pk)
    
    if request.method == 'POST':
        match_type = request.POST.get('match_type')
        
        try:
            if match_type == 'payment':
                payment_id = request.POST.get('payment_id')
                payment = get_object_or_404(Payment, pk=payment_id)
                line.match_with_payment(payment, request.user)
                messages.success(request, f'Line matched with payment {payment.payment_number}.')
            
            elif match_type == 'journal':
                journal_line_id = request.POST.get('journal_line_id')
                journal_line = get_object_or_404(JournalEntryLine, pk=journal_line_id)
                line.match_with_journal(journal_line, request.user)
                messages.success(request, f'Line matched with journal {journal_line.journal_entry.entry_number}.')
            
        except Exception as e:
            messages.error(request, f'Matching failed: {e}')
    
    return redirect('finance:bankstatement_detail', pk=pk)


@login_required
def bankstatement_unmatch(request, pk, line_id):
    """Unmatch a statement line."""
    from .models import BankStatement, BankStatementLine
    
    statement = get_object_or_404(BankStatement, pk=pk)
    line = get_object_or_404(BankStatementLine, pk=line_id, statement=statement)
    
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:bankstatement_detail', pk=pk)
    
    if statement.status in ['reconciled', 'locked']:
        messages.error(request, 'Cannot modify reconciled/locked statements.')
        return redirect('finance:bankstatement_detail', pk=pk)
    
    try:
        line.unmatch()
        messages.success(request, 'Line unmatched successfully.')
    except Exception as e:
        messages.error(request, f'Unmatch failed: {e}')
    
    return redirect('finance:bankstatement_detail', pk=pk)


@login_required
def bankstatement_adjustment(request, pk, line_id):
    """Create adjustment entry for unmatched bank item."""
    from .models import BankStatement, BankStatementLine
    
    statement = get_object_or_404(BankStatement, pk=pk)
    line = get_object_or_404(BankStatementLine, pk=line_id, statement=statement)
    
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:bankstatement_detail', pk=pk)
    
    if statement.status in ['reconciled', 'locked']:
        messages.error(request, 'Cannot modify reconciled/locked statements.')
        return redirect('finance:bankstatement_detail', pk=pk)
    
    if request.method == 'POST':
        adjustment_type = request.POST.get('adjustment_type')
        expense_account_id = request.POST.get('expense_account_id')
        
        try:
            expense_account = get_object_or_404(Account, pk=expense_account_id)
            journal = line.create_adjustment(adjustment_type, expense_account, request.user)
            messages.success(request, f'Adjustment journal {journal.entry_number} created.')
        except Exception as e:
            messages.error(request, f'Adjustment failed: {e}')
    
    return redirect('finance:bankstatement_detail', pk=pk)


@login_required
def bankstatement_finalize(request, pk):
    """Finalize bank statement reconciliation."""
    from .models import BankStatement
    
    statement = get_object_or_404(BankStatement, pk=pk)
    
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:bankstatement_detail', pk=pk)
    
    try:
        statement.finalize(request.user)
        messages.success(request, f'Statement {statement.statement_number} reconciled successfully.')
    except ValidationError as e:
        for error in e.messages:
            messages.error(request, error)
    except Exception as e:
        messages.error(request, f'Finalization failed: {e}')
    
    return redirect('finance:bankstatement_detail', pk=pk)


@login_required
def bankstatement_lock(request, pk):
    """Lock a reconciled bank statement."""
    from .models import BankStatement
    
    statement = get_object_or_404(BankStatement, pk=pk)
    
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'approve')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:bankstatement_detail', pk=pk)
    
    try:
        statement.lock(request.user)
        messages.success(request, f'Statement {statement.statement_number} locked.')
    except Exception as e:
        messages.error(request, f'Lock failed: {e}')
    
    return redirect('finance:bankstatement_detail', pk=pk)


# ============ BANK RECONCILIATION VIEWS ============

class BankReconciliationListView(PermissionRequiredMixin, ListView):
    model = BankReconciliation
    template_name = 'finance/bankreconciliation_list.html'
    context_object_name = 'reconciliations'
    module_name = 'finance'
    permission_type = 'view'
    paginate_by = 25
    
    def get_queryset(self):
        queryset = BankReconciliation.objects.filter(is_active=True).select_related('bank_account', 'bank_statement')
        
        bank_id = self.request.GET.get('bank')
        if bank_id:
            queryset = queryset.filter(bank_account_id=bank_id)
        
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Bank Reconciliations'
        context['bank_accounts'] = BankAccount.objects.filter(is_active=True)
        context['status_choices'] = BankReconciliation.STATUS_CHOICES
        context['can_create'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'create')
        context['can_edit'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'edit')
        context['can_approve'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'approve')
        return context


class BankReconciliationCreateView(CreatePermissionMixin, CreateView):
    model = BankReconciliation
    template_name = 'finance/bankreconciliation_form.html'
    success_url = reverse_lazy('finance:bankreconciliation_list')
    module_name = 'finance'
    fields = ['bank_account', 'bank_statement', 'reconciliation_date', 'period_start', 'period_end',
              'statement_opening_balance', 'statement_closing_balance', 'notes']
    
    def get_form(self, form_class=None):
        from .models import BankStatement
        form = super().get_form(form_class)
        for field_name in ['bank_account', 'bank_statement']:
            form.fields[field_name].widget.attrs['class'] = 'form-select'
        form.fields['bank_account'].queryset = BankAccount.objects.filter(is_active=True)
        form.fields['bank_statement'].queryset = BankStatement.objects.filter(
            is_active=True, status__in=['draft', 'in_progress', 'reconciled']
        )
        form.fields['bank_statement'].required = False
        for field_name in ['reconciliation_date', 'period_start', 'period_end']:
            form.fields[field_name].widget = forms.DateInput(attrs={'type': 'date', 'class': 'form-control'})
        for field_name in ['statement_opening_balance', 'statement_closing_balance']:
            form.fields[field_name].widget.attrs['class'] = 'form-control'
        form.fields['notes'].widget = forms.Textarea(attrs={'rows': 2, 'class': 'form-control'})
        return form
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Create Bank Reconciliation'
        context['today'] = date.today().isoformat()
        return context
    
    def form_valid(self, form):
        self.object = form.save()
        if self.object.bank_statement:
            self.object.calculate_from_statement()
        else:
            self.object.calculate()
        messages.success(self.request, 'Bank Reconciliation created successfully.')
        return redirect(self.success_url)


class BankReconciliationDetailView(PermissionRequiredMixin, DetailView):
    model = BankReconciliation
    template_name = 'finance/bankreconciliation_detail.html'
    context_object_name = 'reconciliation'
    module_name = 'finance'
    permission_type = 'view'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Reconciliation: {self.object.reconciliation_number}'
        context['items'] = self.object.items.all()
        
        # Permissions
        context['can_edit'] = (
            self.request.user.is_superuser or 
            PermissionChecker.has_permission(self.request.user, 'finance', 'edit')
        ) and self.object.status in ['draft', 'in_progress']
        
        context['can_complete'] = context['can_edit'] and self.object.is_reconciled
        context['can_approve'] = (
            self.request.user.is_superuser or 
            PermissionChecker.has_permission(self.request.user, 'finance', 'approve')
        ) and self.object.status == 'completed'
        
        return context


@login_required
def bankreconciliation_complete(request, pk):
    """Complete a bank reconciliation."""
    reconciliation = get_object_or_404(BankReconciliation, pk=pk)
    
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:bankreconciliation_detail', pk=pk)
    
    try:
        reconciliation.complete(request.user)
        messages.success(request, f'Reconciliation {reconciliation.reconciliation_number} completed.')
    except ValidationError as e:
        for error in e.messages:
            messages.error(request, error)
    except Exception as e:
        messages.error(request, f'Failed to complete: {e}')
    
    return redirect('finance:bankreconciliation_detail', pk=pk)


@login_required
def bankreconciliation_approve(request, pk):
    """Approve a completed bank reconciliation."""
    reconciliation = get_object_or_404(BankReconciliation, pk=pk)
    
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'approve')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:bankreconciliation_detail', pk=pk)
    
    try:
        reconciliation.approve(request.user)
        messages.success(request, f'Reconciliation {reconciliation.reconciliation_number} approved.')
    except ValidationError as e:
        for error in e.messages:
            messages.error(request, error)
    except Exception as e:
        messages.error(request, f'Failed to approve: {e}')
    
    return redirect('finance:bankreconciliation_detail', pk=pk)


# ============ RECONCILIATION REPORTS ============

@login_required
def reconciliation_statement_report(request):
    """Bank Reconciliation Statement Report."""
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'view')):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard')
    
    bank_id = request.GET.get('bank')
    as_of_date = request.GET.get('date', date.today().isoformat())
    
    bank_accounts = BankAccount.objects.filter(is_active=True)
    selected_bank = None
    statement_data = None
    
    if bank_id:
        selected_bank = get_object_or_404(BankAccount, pk=bank_id)
        
        # Get GL balance
        gl_account = selected_bank.gl_account
        gl_lines = JournalEntryLine.objects.filter(
            account=gl_account,
            journal_entry__status='posted',
            journal_entry__date__lte=as_of_date,
        ).aggregate(
            total_debit=Sum('debit'),
            total_credit=Sum('credit')
        )
        
        gl_balance = gl_account.opening_balance + (gl_lines['total_debit'] or Decimal('0.00')) - (gl_lines['total_credit'] or Decimal('0.00'))
        
        # Get latest bank statement
        from .models import BankStatement
        latest_statement = BankStatement.objects.filter(
            bank_account=selected_bank,
            statement_end_date__lte=as_of_date,
            status__in=['reconciled', 'locked']
        ).order_by('-statement_end_date').first()
        
        bank_balance = latest_statement.closing_balance if latest_statement else selected_bank.bank_statement_balance
        
        # Outstanding items (deposits in transit, outstanding checks)
        outstanding_deposits = Payment.objects.filter(
            bank_account=selected_bank,
            payment_type='received',
            status='confirmed',
            payment_date__lte=as_of_date,
        ).exclude(
            id__in=BankStatementLine.objects.filter(
                matched_payment__isnull=False,
                statement__status__in=['reconciled', 'locked']
            ).values_list('matched_payment_id', flat=True)
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        
        outstanding_checks = Payment.objects.filter(
            bank_account=selected_bank,
            payment_type='made',
            status='confirmed',
            payment_date__lte=as_of_date,
        ).exclude(
            id__in=BankStatementLine.objects.filter(
                matched_payment__isnull=False,
                statement__status__in=['reconciled', 'locked']
            ).values_list('matched_payment_id', flat=True)
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        
        adjusted_bank_balance = bank_balance + outstanding_deposits - outstanding_checks
        difference = gl_balance - adjusted_bank_balance
        
        statement_data = {
            'gl_balance': gl_balance,
            'bank_balance': bank_balance,
            'outstanding_deposits': outstanding_deposits,
            'outstanding_checks': outstanding_checks,
            'adjusted_bank_balance': adjusted_bank_balance,
            'difference': difference,
            'is_reconciled': abs(difference) < Decimal('0.01'),
        }
    
    return render(request, 'finance/reconciliation_statement_report.html', {
        'title': 'Bank Reconciliation Statement',
        'bank_accounts': bank_accounts,
        'selected_bank': selected_bank,
        'statement_data': statement_data,
        'as_of_date': as_of_date,
    })


@login_required
def unreconciled_transactions_report(request):
    """Unreconciled Transactions Report."""
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'view')):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard')
    
    bank_id = request.GET.get('bank')
    bank_accounts = BankAccount.objects.filter(is_active=True)
    selected_bank = None
    unreconciled_data = None
    
    if bank_id:
        selected_bank = get_object_or_404(BankAccount, pk=bank_id)
        
        # Unreconciled payments
        unreconciled_received = Payment.objects.filter(
            bank_account=selected_bank,
            payment_type='received',
            status='confirmed',
        ).exclude(
            id__in=BankStatementLine.objects.filter(
                matched_payment__isnull=False
            ).values_list('matched_payment_id', flat=True)
        )
        
        unreconciled_made = Payment.objects.filter(
            bank_account=selected_bank,
            payment_type='made',
            status='confirmed',
        ).exclude(
            id__in=BankStatementLine.objects.filter(
                matched_payment__isnull=False
            ).values_list('matched_payment_id', flat=True)
        )
        
        # Unreconciled bank statement lines
        from .models import BankStatement
        unreconciled_statement_lines = BankStatementLine.objects.filter(
            statement__bank_account=selected_bank,
            reconciliation_status='unmatched',
        ).select_related('statement')
        
        unreconciled_data = {
            'received': unreconciled_received,
            'made': unreconciled_made,
            'statement_lines': unreconciled_statement_lines,
            'total_received': unreconciled_received.aggregate(total=Sum('amount'))['total'] or Decimal('0.00'),
            'total_made': unreconciled_made.aggregate(total=Sum('amount'))['total'] or Decimal('0.00'),
        }
    
    return render(request, 'finance/unreconciled_transactions_report.html', {
        'title': 'Unreconciled Transactions',
        'bank_accounts': bank_accounts,
        'selected_bank': selected_bank,
        'unreconciled_data': unreconciled_data,
    })


@login_required
def reconciliation_adjustments_report(request):
    """Reconciliation Adjustments Report."""
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'view')):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard')
    
    start_date = request.GET.get('start_date', date(date.today().year, 1, 1).isoformat())
    end_date = request.GET.get('end_date', date.today().isoformat())
    bank_id = request.GET.get('bank')
    
    bank_accounts = BankAccount.objects.filter(is_active=True)
    
    # Get adjustment journal entries (tagged via reference starting with ADJ-)
    adjustments = JournalEntry.objects.filter(
        reference__startswith='ADJ-',
        status='posted',
        date__gte=start_date,
        date__lte=end_date,
    ).order_by('-date')
    
    if bank_id:
        selected_bank = get_object_or_404(BankAccount, pk=bank_id)
        adjustments = adjustments.filter(
            lines__account=selected_bank.gl_account
        ).distinct()
    else:
        selected_bank = None
    
    return render(request, 'finance/reconciliation_adjustments_report.html', {
        'title': 'Reconciliation Adjustments',
        'bank_accounts': bank_accounts,
        'selected_bank': selected_bank,
        'adjustments': adjustments,
        'start_date': start_date,
        'end_date': end_date,
    })


@login_required
def cleared_vs_uncleared_report(request):
    """Cleared vs Uncleared Transactions Report."""
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'view')):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard')
    
    bank_id = request.GET.get('bank')
    bank_accounts = BankAccount.objects.filter(is_active=True)
    selected_bank = None
    report_data = None
    
    if bank_id:
        selected_bank = get_object_or_404(BankAccount, pk=bank_id)
        
        # Cleared (reconciled) payments
        cleared_payments = Payment.objects.filter(
            bank_account=selected_bank,
            status='reconciled',
        )
        
        # Uncleared (confirmed but not reconciled) payments
        uncleared_payments = Payment.objects.filter(
            bank_account=selected_bank,
            status='confirmed',
        )
        
        report_data = {
            'cleared': cleared_payments,
            'uncleared': uncleared_payments,
            'cleared_total': cleared_payments.aggregate(total=Sum('amount'))['total'] or Decimal('0.00'),
            'uncleared_total': uncleared_payments.aggregate(total=Sum('amount'))['total'] or Decimal('0.00'),
        }
    
    return render(request, 'finance/cleared_vs_uncleared_report.html', {
        'title': 'Cleared vs Uncleared Transactions',
        'bank_accounts': bank_accounts,
        'selected_bank': selected_bank,
        'report_data': report_data,
    })


@login_required
def bank_vs_gl_report(request):
    """Bank Ledger vs GL Ledger Difference Report."""
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'view')):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard')
    
    as_of_date = request.GET.get('date', date.today().isoformat())
    
    bank_accounts = BankAccount.objects.filter(is_active=True)
    comparison_data = []
    
    for bank in bank_accounts:
        # GL balance
        gl_account = bank.gl_account
        gl_lines = JournalEntryLine.objects.filter(
            account=gl_account,
            journal_entry__status='posted',
            journal_entry__date__lte=as_of_date,
        ).aggregate(
            total_debit=Sum('debit'),
            total_credit=Sum('credit')
        )
        
        gl_balance = gl_account.opening_balance + (gl_lines['total_debit'] or Decimal('0.00')) - (gl_lines['total_credit'] or Decimal('0.00'))
        
        # Latest reconciled statement balance
        from .models import BankStatement
        latest_statement = BankStatement.objects.filter(
            bank_account=bank,
            status__in=['reconciled', 'locked']
        ).order_by('-statement_end_date').first()
        
        bank_balance = latest_statement.closing_balance if latest_statement else bank.bank_statement_balance
        
        difference = gl_balance - bank_balance
        
        comparison_data.append({
            'bank': bank,
            'gl_balance': gl_balance,
            'bank_balance': bank_balance,
            'difference': difference,
            'is_reconciled': abs(difference) < Decimal('0.01'),
            'last_reconciled': latest_statement.statement_end_date if latest_statement else None,
        })
    
    return render(request, 'finance/bank_vs_gl_report.html', {
        'title': 'Bank vs GL Ledger Comparison',
        'comparison_data': comparison_data,
        'as_of_date': as_of_date,
    })


# ============ OPENING BALANCE VIEWS ============

class OpeningBalanceListView(PermissionRequiredMixin, ListView):
    """List of all opening balance entries."""
    model = OpeningBalanceEntry
    template_name = 'finance/openingbalance_list.html'
    context_object_name = 'entries'
    module_name = 'finance'
    permission_type = 'view'
    
    def get_queryset(self):
        queryset = OpeningBalanceEntry.objects.filter(is_active=True).select_related(
            'fiscal_year', 'journal_entry', 'posted_by'
        )
        
        entry_type = self.request.GET.get('type')
        if entry_type:
            queryset = queryset.filter(entry_type=entry_type)
        
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
        
        return queryset.order_by('-entry_date', '-created_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Opening Balances'
        context['entry_types'] = OpeningBalanceEntry.ENTRY_TYPE_CHOICES
        return context


class OpeningBalanceCreateView(CreatePermissionMixin, CreateView):
    """Create new opening balance entry."""
    model = OpeningBalanceEntry
    form_class = OpeningBalanceEntryForm
    template_name = 'finance/openingbalance_form.html'
    success_url = reverse_lazy('finance:openingbalance_list')
    module_name = 'finance'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Create Opening Balance Entry'
        context['today'] = date.today().isoformat()
        if self.request.POST:
            context['formset'] = OpeningBalanceLineFormSet(self.request.POST)
        else:
            context['formset'] = OpeningBalanceLineFormSet()
        return context
    
    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['formset']
        
        if formset.is_valid():
            self.object = form.save(commit=False)
            self.object.created_by = self.request.user
            self.object.save()
            
            formset.instance = self.object
            formset.save()
            
            self.object.calculate_totals()
            
            messages.success(self.request, f'Opening Balance Entry {self.object.entry_number} created successfully.')
            return redirect(self.success_url)
        else:
            return self.render_to_response(self.get_context_data(form=form))


class OpeningBalanceDetailView(PermissionRequiredMixin, DetailView):
    """View opening balance entry details."""
    model = OpeningBalanceEntry
    template_name = 'finance/openingbalance_detail.html'
    context_object_name = 'entry'
    module_name = 'finance'
    permission_type = 'view'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Opening Balance: {self.object.entry_number}'
        context['lines'] = self.object.lines.all().select_related('account', 'customer', 'vendor', 'bank_account')
        return context


class OpeningBalanceUpdateView(UpdatePermissionMixin, UpdateView):
    """Update opening balance entry."""
    model = OpeningBalanceEntry
    form_class = OpeningBalanceEntryForm
    template_name = 'finance/openingbalance_form.html'
    success_url = reverse_lazy('finance:openingbalance_list')
    module_name = 'finance'
    
    def get_queryset(self):
        return OpeningBalanceEntry.objects.filter(status='draft')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Edit Opening Balance: {self.object.entry_number}'
        context['today'] = date.today().isoformat()
        if self.request.POST:
            context['formset'] = OpeningBalanceLineFormSet(self.request.POST, instance=self.object)
        else:
            context['formset'] = OpeningBalanceLineFormSet(instance=self.object)
        return context
    
    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['formset']
        
        if formset.is_valid():
            self.object = form.save(commit=False)
            self.object.updated_by = self.request.user
            self.object.save()
            
            formset.save()
            self.object.calculate_totals()
            
            messages.success(self.request, f'Opening Balance Entry {self.object.entry_number} updated successfully.')
            return redirect('finance:openingbalance_detail', pk=self.object.pk)
        else:
            return self.render_to_response(self.get_context_data(form=form))


@login_required
def openingbalance_post(request, pk):
    """Post an opening balance entry."""
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:openingbalance_list')
    
    entry = get_object_or_404(OpeningBalanceEntry, pk=pk)
    
    if request.method == 'POST':
        try:
            journal = entry.post(request.user)
            messages.success(request, f'Opening Balance Entry {entry.entry_number} posted successfully. Journal Entry: {journal.entry_number}')
        except ValidationError as e:
            messages.error(request, str(e))
    
    return redirect('finance:openingbalance_detail', pk=pk)


@login_required
def openingbalance_reverse(request, pk):
    """Reverse a posted opening balance entry."""
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:openingbalance_list')
    
    entry = get_object_or_404(OpeningBalanceEntry, pk=pk)
    
    if request.method == 'POST':
        try:
            entry.reverse(request.user)
            messages.success(request, f'Opening Balance Entry {entry.entry_number} reversed successfully.')
        except ValidationError as e:
            messages.error(request, str(e))
    
    return redirect('finance:openingbalance_detail', pk=pk)


# ============ WRITE-OFF VIEWS ============

class WriteOffListView(PermissionRequiredMixin, ListView):
    """List of all write-off entries."""
    model = WriteOff
    template_name = 'finance/writeoff_list.html'
    context_object_name = 'writeoffs'
    module_name = 'finance'
    permission_type = 'view'
    
    def get_queryset(self):
        queryset = WriteOff.objects.filter(is_active=True).select_related(
            'source_account', 'expense_account', 'customer', 'vendor', 'journal_entry'
        )
        
        writeoff_type = self.request.GET.get('type')
        if writeoff_type:
            queryset = queryset.filter(writeoff_type=writeoff_type)
        
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
        
        return queryset.order_by('-writeoff_date', '-created_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Write-Offs & Adjustments'
        context['writeoff_types'] = WriteOff.WRITEOFF_TYPE_CHOICES
        return context


class WriteOffCreateView(CreatePermissionMixin, CreateView):
    """Create new write-off entry."""
    model = WriteOff
    form_class = WriteOffForm
    template_name = 'finance/writeoff_form.html'
    success_url = reverse_lazy('finance:writeoff_list')
    module_name = 'finance'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Create Write-Off / Adjustment'
        context['today'] = date.today().isoformat()
        return context
    
    def form_valid(self, form):
        self.object = form.save(commit=False)
        self.object.created_by = self.request.user
        self.object.save()
        
        messages.success(self.request, f'Write-Off {self.object.writeoff_number} created successfully.')
        return redirect('finance:writeoff_detail', pk=self.object.pk)


class WriteOffDetailView(PermissionRequiredMixin, DetailView):
    """View write-off entry details."""
    model = WriteOff
    template_name = 'finance/writeoff_detail.html'
    context_object_name = 'writeoff'
    module_name = 'finance'
    permission_type = 'view'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Write-Off: {self.object.writeoff_number}'
        return context


class WriteOffUpdateView(UpdatePermissionMixin, UpdateView):
    """Update write-off entry."""
    model = WriteOff
    form_class = WriteOffForm
    template_name = 'finance/writeoff_form.html'
    module_name = 'finance'
    
    def get_queryset(self):
        return WriteOff.objects.filter(status='draft')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Edit Write-Off: {self.object.writeoff_number}'
        context['today'] = date.today().isoformat()
        return context
    
    def form_valid(self, form):
        self.object = form.save(commit=False)
        self.object.updated_by = self.request.user
        self.object.save()
        
        messages.success(self.request, f'Write-Off {self.object.writeoff_number} updated successfully.')
        return redirect('finance:writeoff_detail', pk=self.object.pk)


@login_required
def writeoff_approve(request, pk):
    """Approve a write-off entry."""
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'approve')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:writeoff_list')
    
    writeoff = get_object_or_404(WriteOff, pk=pk)
    
    if request.method == 'POST':
        try:
            writeoff.approve(request.user)
            messages.success(request, f'Write-Off {writeoff.writeoff_number} approved successfully.')
        except ValidationError as e:
            messages.error(request, str(e))
    
    return redirect('finance:writeoff_detail', pk=pk)


@login_required
def writeoff_post(request, pk):
    """Post a write-off entry."""
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:writeoff_list')
    
    writeoff = get_object_or_404(WriteOff, pk=pk)
    
    if request.method == 'POST':
        try:
            journal = writeoff.post(request.user)
            messages.success(request, f'Write-Off {writeoff.writeoff_number} posted successfully. Journal Entry: {journal.entry_number}')
        except ValidationError as e:
            messages.error(request, str(e))
    
    return redirect('finance:writeoff_detail', pk=pk)


@login_required
def writeoff_reverse(request, pk):
    """Reverse a posted write-off entry."""
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:writeoff_list')
    
    writeoff = get_object_or_404(WriteOff, pk=pk)
    
    if request.method == 'POST':
        try:
            writeoff.reverse(request.user)
            messages.success(request, f'Write-Off {writeoff.writeoff_number} reversed successfully.')
        except ValidationError as e:
            messages.error(request, str(e))
    
    return redirect('finance:writeoff_detail', pk=pk)


# ============ EXCHANGE RATE VIEWS ============

class ExchangeRateListView(PermissionRequiredMixin, ListView):
    """List of all exchange rates with inline creation form."""
    model = ExchangeRate
    template_name = 'finance/exchangerate_list.html'
    context_object_name = 'rates'
    module_name = 'finance'
    permission_type = 'view'
    
    def get_queryset(self):
        queryset = ExchangeRate.objects.filter(is_active=True)
        
        currency = self.request.GET.get('currency')
        if currency:
            queryset = queryset.filter(currency_code=currency.upper())
        
        return queryset.order_by('-rate_date', 'currency_code')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Exchange Rates'
        # Get unique currencies
        context['currencies'] = ExchangeRate.objects.filter(is_active=True).values_list(
            'currency_code', flat=True
        ).distinct().order_by('currency_code')
        # Add form for inline creation
        context['form'] = ExchangeRateForm()
        context['today'] = date.today().isoformat()
        return context
    
    def post(self, request, *args, **kwargs):
        """Handle inline form submission."""
        if not PermissionChecker.has_permission(request.user, 'finance', 'create'):
            messages.error(request, 'Permission denied.')
            return redirect('finance:exchangerate_list')
        
        form = ExchangeRateForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.created_by = request.user
            obj.save()
            messages.success(request, f'Exchange Rate for {obj.currency_code} added successfully.')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
        
        return redirect('finance:exchangerate_list')


class ExchangeRateCreateView(CreatePermissionMixin, CreateView):
    """Create new exchange rate - kept for backwards compatibility."""
    model = ExchangeRate
    form_class = ExchangeRateForm
    template_name = 'finance/exchangerate_form.html'
    success_url = reverse_lazy('finance:exchangerate_list')
    module_name = 'finance'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Add Exchange Rate'
        context['today'] = date.today().isoformat()
        return context
    
    def form_valid(self, form):
        self.object = form.save(commit=False)
        self.object.created_by = self.request.user
        self.object.save()
        
        messages.success(self.request, f'Exchange Rate for {self.object.currency_code} added successfully.')
        return redirect(self.success_url)


class ExchangeRateUpdateView(UpdatePermissionMixin, UpdateView):
    """Update exchange rate."""
    model = ExchangeRate
    form_class = ExchangeRateForm
    template_name = 'finance/exchangerate_form.html'
    success_url = reverse_lazy('finance:exchangerate_list')
    module_name = 'finance'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Edit Exchange Rate: {self.object.currency_code}'
        return context


# ============ VAT AUDIT REPORT ============

@login_required
def vat_audit_report(request):
    """
    VAT Audit Report - Line-level VAT details for FTA audit.
    Shows every transaction mapped to VAT box.
    """
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'view')):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard')
    
    # Date range
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if not start_date or not end_date:
        # Default to current month
        today = date.today()
        start_date = today.replace(day=1).isoformat()
        end_date = today.isoformat()
    
    # Get all VAT-related journal lines
    vat_tax_codes = TaxCode.objects.filter(is_active=True)
    vat_accounts = []
    for tc in vat_tax_codes:
        if tc.sales_account:
            vat_accounts.append(tc.sales_account.id)
        if tc.purchase_account:
            vat_accounts.append(tc.purchase_account.id)
    
    vat_accounts = list(set(vat_accounts))
    
    # Get all journal entries with VAT impact
    journal_lines = JournalEntryLine.objects.filter(
        journal_entry__status='posted',
        journal_entry__date__gte=start_date,
        journal_entry__date__lte=end_date,
    ).select_related('journal_entry', 'account').order_by('journal_entry__date', 'journal_entry__entry_number')
    
    # Group by VAT box
    vat_data = {
        'box1a': [],  # Standard rated supplies - Emirates
        'box1b': [],  # Standard rated supplies - GCC
        'box2': [],   # Tax refunds
        'box3': [],   # Zero-rated supplies
        'box4': [],   # Exempt supplies
        'box5': [],   # Total value of outputs
        'box6': [],   # Standard rated expenses
        'box7': [],   # Supplies subject to reverse charge
        'box8': [],   # Total value of inputs
        'box9': [],   # Output VAT due
        'box10': [],  # Input VAT recoverable
    }
    
    all_transactions = []
    
    for line in journal_lines:
        # Determine VAT box based on account
        vat_box = 'N/A'
        if line.account.account_type == 'income':
            # Output - Sales
            vat_box = 'Box 1a - Standard Supplies'
            vat_data['box1a'].append(line)
        elif line.account.account_type == 'expense':
            # Input - Expenses
            vat_box = 'Box 6 - Standard Expenses'
            vat_data['box6'].append(line)
        elif line.account.id in vat_accounts:
            # VAT accounts
            if line.credit > 0:
                vat_box = 'Box 9 - Output VAT'
                vat_data['box9'].append(line)
            else:
                vat_box = 'Box 10 - Input VAT'
                vat_data['box10'].append(line)
        
        all_transactions.append({
            'date': line.journal_entry.date,
            'entry_number': line.journal_entry.entry_number,
            'reference': line.journal_entry.reference,
            'description': line.description or line.journal_entry.description,
            'account': line.account,
            'debit': line.debit,
            'credit': line.credit,
            'vat_box': vat_box,
        })
    
    # Calculate totals by box
    box_totals = {}
    for box_name, lines in vat_data.items():
        total_debit = sum(l.debit for l in lines)
        total_credit = sum(l.credit for l in lines)
        box_totals[box_name] = {
            'count': len(lines),
            'debit': total_debit,
            'credit': total_credit,
            'net': total_debit - total_credit,
        }
    
    return render(request, 'finance/vat_audit_report.html', {
        'title': 'VAT Audit Report',
        'start_date': start_date,
        'end_date': end_date,
        'transactions': all_transactions,
        'box_totals': box_totals,
        'total_transactions': len(all_transactions),
    })


# ============ ACCOUNT MAPPING VIEWS ============
# SAP/Oracle-style Account Determination / Posting Profiles

@login_required
def account_mapping_list(request):
    """
    Account Mapping / Account Determination - Central configuration.
    One-time setup for all transaction types.
    """
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'view')):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard')
    
    can_edit = request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')
    
    # Group mappings by module
    modules = AccountMapping.MODULE_CHOICES
    mappings_by_module = {}
    
    for module_code, module_name in modules:
        mappings = AccountMapping.objects.filter(module=module_code).select_related('account')
        
        # Get all transaction types for this module
        module_types = [
            (code, label) for code, label in AccountMapping.TRANSACTION_TYPE_CHOICES
            if code.startswith(module_code) or 
               (module_code == 'general' and code in ['fx_gain', 'fx_loss', 'retained_earnings', 'opening_balance_equity', 'suspense', 'rounding']) or
               (module_code == 'banking' and code.startswith('bank_'))
        ]
        
        configured_types = {m.transaction_type: m for m in mappings}
        
        module_data = []
        for type_code, type_label in module_types:
            mapping = configured_types.get(type_code)
            module_data.append({
                'transaction_type': type_code,
                'label': type_label,
                'mapping': mapping,
                'account': mapping.account if mapping else None,
            })
        
        if module_data:
            mappings_by_module[module_code] = {
                'name': module_name,
                'items': module_data,
                'is_configured': AccountMapping.is_fully_configured(module_code),
            }
    
    # Get all active accounts for the dropdown
    accounts = Account.objects.filter(is_active=True).order_by('code')
    
    return render(request, 'finance/account_mapping_list.html', {
        'title': 'Account Mapping',
        'mappings_by_module': mappings_by_module,
        'accounts': accounts,
        'can_edit': can_edit,
    })


@login_required
def account_mapping_save(request):
    """
    Save account mapping via AJAX or form POST.
    """
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        return JsonResponse({'success': False, 'error': 'Permission denied.'})
    
    if request.method == 'POST':
        transaction_type = request.POST.get('transaction_type')
        account_id = request.POST.get('account_id')
        
        if not transaction_type:
            return JsonResponse({'success': False, 'error': 'Transaction type required.'})
        
        # Determine module from transaction type
        module = 'general'
        for mod_code, _ in AccountMapping.MODULE_CHOICES:
            if transaction_type.startswith(mod_code):
                module = mod_code
                break
        
        # Handle special cases
        if transaction_type.startswith('bank_'):
            module = 'banking'
        elif transaction_type in ['fx_gain', 'fx_loss', 'retained_earnings', 'opening_balance_equity', 'suspense', 'rounding']:
            module = 'general'
        
        if account_id:
            try:
                account = Account.objects.get(pk=account_id, is_active=True)
                mapping, created = AccountMapping.objects.update_or_create(
                    transaction_type=transaction_type,
                    defaults={
                        'module': module,
                        'account': account,
                    }
                )
                return JsonResponse({
                    'success': True, 
                    'message': f'Mapping saved: {account.code} - {account.name}',
                    'account_code': account.code,
                    'account_name': account.name,
                })
            except Account.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Account not found.'})
        else:
            # Remove mapping if account_id is empty
            AccountMapping.objects.filter(transaction_type=transaction_type).delete()
            return JsonResponse({'success': True, 'message': 'Mapping removed.'})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method.'})


@login_required
def accounting_settings(request):
    """
    View and edit global accounting settings.
    Controls auto-posting behavior per module.
    """
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard')
    
    settings_obj = AccountingSettings.get_settings()
    
    if request.method == 'POST':
        # Update settings
        settings_obj.auto_post_sales_invoice = request.POST.get('auto_post_sales_invoice') == 'on'
        settings_obj.auto_post_vendor_bill = request.POST.get('auto_post_vendor_bill') == 'on'
        settings_obj.auto_post_expense_claim = request.POST.get('auto_post_expense_claim') == 'on'
        settings_obj.auto_post_payroll = request.POST.get('auto_post_payroll') == 'on'
        settings_obj.auto_post_payment = request.POST.get('auto_post_payment') == 'on'
        settings_obj.auto_post_bank_transfer = request.POST.get('auto_post_bank_transfer') == 'on'
        settings_obj.require_approval_before_posting = request.POST.get('require_approval_before_posting') == 'on'
        settings_obj.allow_posting_to_closed_period = request.POST.get('allow_posting_to_closed_period') == 'on'
        settings_obj.round_to_fils = request.POST.get('round_to_fils') == 'on'
        
        # VAT rate
        try:
            vat_rate = Decimal(request.POST.get('default_vat_rate', '5.00'))
            settings_obj.default_vat_rate = vat_rate
        except:
            pass
        
        settings_obj.vat_registration_number = request.POST.get('vat_registration_number', '')
        
        settings_obj.save()
        messages.success(request, 'Accounting settings updated successfully.')
        return redirect('finance:accounting_settings')
    
    return render(request, 'finance/accounting_settings.html', {
        'title': 'Accounting Settings',
        'settings': settings_obj,
    })
