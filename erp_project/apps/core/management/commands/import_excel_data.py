"""
Management command to import data from Excel file.
Imports data into Company Settings, Chart of Accounts, Customers, Vendors, 
Employees, Sales Invoices, Purchase Bills, Expense Claims, and Journal Entries.
"""
import os
from datetime import datetime, date
from decimal import Decimal
from django.core.management.base import BaseCommand
from django.db import transaction
from django.contrib.auth import get_user_model

import openpyxl

User = get_user_model()


class Command(BaseCommand):
    help = 'Import data from Excel file into the ERP system'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            default='Excel Upload/Finance Module.xlsx',
            help='Path to the Excel file'
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Run without saving to database'
        )

    def handle(self, *args, **options):
        file_path = options['file']
        self.dry_run = options['dry_run']
        
        if not os.path.exists(file_path):
            self.stderr.write(self.style.ERROR(f'File not found: {file_path}'))
            return
        
        self.stdout.write(self.style.SUCCESS(f'Loading workbook: {file_path}'))
        self.wb = openpyxl.load_workbook(file_path)
        
        # Get or create admin user for tracking
        self.admin_user = User.objects.filter(is_superuser=True).first()
        if not self.admin_user:
            self.admin_user = User.objects.first()
        
        # Store mappings for cross-referencing
        self.customer_map = {}  # customer_code -> Customer
        self.vendor_map = {}    # vendor_code -> Vendor
        self.employee_map = {}  # employee_code -> Employee
        self.account_map = {}   # account_code -> Account
        
        try:
            with transaction.atomic():
                # Import in order of dependencies
                self.import_company_settings()  # Sheet1
                self.import_chart_of_accounts()  # Sheet2
                self.import_customers()          # Sheet3
                self.import_vendors()            # Sheet4
                self.import_employees()          # Sheet5
                self.import_opening_balances()   # Sheet6
                self.import_sales_invoices()     # Sheet7
                self.import_purchase_bills()     # Sheet8
                self.import_expense_claims()     # Sheet9
                self.import_journal_entries()    # Sheets 12-15, 18
                
                if self.dry_run:
                    self.stdout.write(self.style.WARNING('DRY RUN - Rolling back changes'))
                    raise Exception('Dry run - rolling back')
                    
        except Exception as e:
            if 'Dry run' not in str(e):
                self.stderr.write(self.style.ERROR(f'Error: {e}'))
                raise
        
        self.stdout.write(self.style.SUCCESS('Import completed successfully!'))

    def get_cell_value(self, cell):
        """Get cell value, handling None and datetime."""
        if cell.value is None:
            return None
        if isinstance(cell.value, datetime):
            return cell.value.date() if hasattr(cell.value, 'date') else cell.value
        return cell.value

    def get_sheet_data(self, sheet_name, header_row=2):
        """Get data rows from a sheet."""
        ws = self.wb[sheet_name]
        headers = [cell.value for cell in ws[header_row]]
        
        data = []
        for row in ws.iter_rows(min_row=header_row + 1):
            row_data = {}
            has_data = False
            for idx, cell in enumerate(row):
                if idx < len(headers) and headers[idx]:
                    value = self.get_cell_value(cell)
                    if value is not None and value != '':
                        has_data = True
                    row_data[headers[idx]] = value
            if has_data:
                data.append(row_data)
        return data

    def import_company_settings(self):
        """Import company settings from Sheet1."""
        self.stdout.write('Importing Company Settings...')
        from apps.settings_app.models import CompanySettings
        
        ws = self.wb['Sheet1']
        settings_data = {}
        
        for row in ws.iter_rows(min_row=3):
            param = self.get_cell_value(row[0])
            value = self.get_cell_value(row[1])
            if param and value:
                settings_data[param] = value
        
        settings = CompanySettings.get_settings()
        
        if 'Company Name' in settings_data:
            settings.company_name = settings_data['Company Name']
        if 'Currency' in settings_data:
            settings.currency = settings_data['Currency']
        if 'Tax ID / TRN' in settings_data or 'TRN' in settings_data:
            settings.tax_id = settings_data.get('Tax ID / TRN') or settings_data.get('TRN', '')
        if 'Address' in settings_data:
            settings.address = settings_data['Address']
        if 'Phone' in settings_data:
            settings.phone = settings_data['Phone']
        if 'Email' in settings_data:
            settings.email = settings_data['Email']
        
        settings.save()
        self.stdout.write(self.style.SUCCESS(f'  Company: {settings.company_name}'))

    def import_chart_of_accounts(self):
        """Import Chart of Accounts from Sheet2."""
        self.stdout.write('Importing Chart of Accounts...')
        from apps.finance.models import Account, AccountType
        
        data = self.get_sheet_data('Sheet2')
        
        type_map = {
            'assets': AccountType.ASSET,
            'asset': AccountType.ASSET,
            'bank': AccountType.ASSET,
            'liabilities': AccountType.LIABILITY,
            'liability': AccountType.LIABILITY,
            'equity': AccountType.EQUITY,
            'income': AccountType.INCOME,
            'revenue': AccountType.INCOME,
            'expenses': AccountType.EXPENSE,
            'expense': AccountType.EXPENSE,
        }
        
        created_count = 0
        updated_count = 0
        
        for row in data:
            code = row.get('Account Code')
            name = row.get('Account Name')
            group = row.get('Group', '')
            acc_type = row.get('Type', '')
            
            if not code or not name:
                continue
            
            # Convert code to string
            code = str(int(code)) if isinstance(code, float) else str(code)
            
            # Determine account type
            account_type = type_map.get(group.lower() if group else '', 
                                       type_map.get(acc_type.lower() if acc_type else '', AccountType.EXPENSE))
            
            account, created = Account.objects.update_or_create(
                code=code,
                defaults={
                    'name': name,
                    'account_type': account_type,
                    'description': f'{group} - {acc_type}' if group and acc_type else '',
                }
            )
            
            self.account_map[code] = account
            
            if created:
                created_count += 1
            else:
                updated_count += 1
        
        self.stdout.write(self.style.SUCCESS(f'  Created: {created_count}, Updated: {updated_count}'))

    def import_customers(self):
        """Import Customers from Sheet3."""
        self.stdout.write('Importing Customers...')
        from apps.crm.models import Customer
        
        data = self.get_sheet_data('Sheet3')
        
        created_count = 0
        updated_count = 0
        
        for row in data:
            code = row.get('Customer Code')
            name = row.get('Customer Name')
            
            if not code or not name:
                continue
            
            trn = row.get('TRN Number', '')
            if isinstance(trn, float):
                trn = str(int(trn))
            
            credit_limit = row.get('Credit Limit (AED)', 0)
            if isinstance(credit_limit, str):
                credit_limit = Decimal(credit_limit.replace(',', '')) if credit_limit else Decimal('0')
            else:
                credit_limit = Decimal(str(credit_limit or 0))
            
            defaults = {
                'name': name,
                'trn': str(trn) if trn else '',
                'payment_terms': row.get('Payment Terms', 'Net 30') or 'Net 30',
                'credit_limit': credit_limit,
                'address': row.get('Address', '') or '',
                'city': row.get('City', '') or '',
                'customer_type': 'customer',
                'status': 'active',
            }
            
            # Try to find existing customer by code pattern
            customer = Customer.objects.filter(customer_number__endswith=code[-4:]).first()
            
            if customer:
                for key, value in defaults.items():
                    setattr(customer, key, value)
                customer.save()
                updated_count += 1
            else:
                customer = Customer.objects.create(**defaults)
                created_count += 1
            
            self.customer_map[code] = customer
        
        self.stdout.write(self.style.SUCCESS(f'  Created: {created_count}, Updated: {updated_count}'))

    def import_vendors(self):
        """Import Vendors from Sheet4."""
        self.stdout.write('Importing Vendors...')
        from apps.purchase.models import Vendor
        
        data = self.get_sheet_data('Sheet4')
        
        created_count = 0
        updated_count = 0
        
        for row in data:
            code = row.get('Vendor Code')
            name = row.get('Vendor Name')
            
            if not code or not name:
                continue
            
            trn = row.get('TRN Number', '')
            if isinstance(trn, float):
                trn = str(int(trn))
            
            credit_limit = row.get('Credit Limit (AED)', 0)
            if isinstance(credit_limit, str):
                credit_limit = Decimal(credit_limit.replace(',', '')) if credit_limit else Decimal('0')
            else:
                credit_limit = Decimal(str(credit_limit or 0))
            
            defaults = {
                'name': name,
                'trn': str(trn) if trn else '',
                'payment_terms': row.get('Payment Terms', 'Net 30') or 'Net 30',
                'credit_limit': credit_limit,
                'address': row.get('Address', '') or '',
                'city': row.get('City', '') or '',
                'status': 'active',
            }
            
            # Try to find existing vendor by code pattern
            vendor = Vendor.objects.filter(vendor_number__endswith=code[-4:]).first()
            
            if vendor:
                for key, value in defaults.items():
                    setattr(vendor, key, value)
                vendor.save()
                updated_count += 1
            else:
                vendor = Vendor.objects.create(**defaults)
                created_count += 1
            
            self.vendor_map[code] = vendor
        
        self.stdout.write(self.style.SUCCESS(f'  Created: {created_count}, Updated: {updated_count}'))

    def import_employees(self):
        """Import Employees from Sheet5."""
        self.stdout.write('Importing Employees...')
        from apps.hr.models import Employee, Department, Designation
        
        data = self.get_sheet_data('Sheet5')
        
        # Create departments cache
        dept_cache = {}
        
        created_count = 0
        updated_count = 0
        
        for row in data:
            code = row.get('Employee Code')
            name = row.get('Employee Name')
            
            if not code or not name:
                continue
            
            # Parse name
            name_parts = name.split(' ', 1)
            first_name = name_parts[0]
            last_name = name_parts[1] if len(name_parts) > 1 else ''
            
            # Get or create department
            dept_name = row.get('Department', 'General')
            if dept_name not in dept_cache:
                dept, _ = Department.objects.get_or_create(
                    name=dept_name,
                    defaults={'code': dept_name[:10].upper().replace(' ', '_')}
                )
                dept_cache[dept_name] = dept
            department = dept_cache[dept_name]
            
            # Get or create designation
            designation_name = row.get('Designation', 'Staff')
            designation, _ = Designation.objects.get_or_create(
                name=designation_name,
                department=department
            )
            
            # Parse salary
            salary = row.get('Salary (AED)', 0)
            if isinstance(salary, str):
                salary = Decimal(salary.replace(',', '')) if salary else Decimal('0')
            else:
                salary = Decimal(str(salary or 0))
            
            # Parse join date
            join_date = row.get('Join Date')
            if isinstance(join_date, str):
                try:
                    join_date = datetime.strptime(join_date, '%Y-%m-%d').date()
                except:
                    join_date = None
            elif isinstance(join_date, datetime):
                join_date = join_date.date()
            
            defaults = {
                'first_name': first_name,
                'last_name': last_name,
                'department': department,
                'designation': designation,
                'basic_salary': salary,
                'date_of_joining': join_date,
                'status': 'active' if row.get('Status', 'Active').lower() == 'active' else 'inactive',
                'email': f"{first_name.lower()}.{last_name.lower()}@company.com" if last_name else f"{first_name.lower()}@company.com",
            }
            
            # Try to find existing employee by code pattern
            employee = Employee.objects.filter(employee_code__endswith=code[-4:]).first()
            
            if employee:
                for key, value in defaults.items():
                    setattr(employee, key, value)
                employee.save()
                updated_count += 1
            else:
                employee = Employee.objects.create(**defaults)
                created_count += 1
            
            self.employee_map[code] = employee
        
        self.stdout.write(self.style.SUCCESS(f'  Created: {created_count}, Updated: {updated_count}'))

    def import_opening_balances(self):
        """Import Opening Balances from Sheet6."""
        self.stdout.write('Importing Opening Balances...')
        from apps.finance.models import Account, JournalEntry, JournalEntryLine
        
        data = self.get_sheet_data('Sheet6')
        
        # Skip header row if it's duplicated
        data = [row for row in data if row.get('Account Code') != 'Account Code']
        
        if not data:
            self.stdout.write('  No opening balance data found')
            return
        
        # Create a single opening balance journal entry
        journal = JournalEntry.objects.create(
            date=date(2024, 1, 1),
            reference='OB-2024-001',
            description='Opening Balances for 2024',
            entry_type='opening_balance',
            created_by=self.admin_user,
        )
        
        line_count = 0
        for row in data:
            code = row.get('Account Code')
            name = row.get('Account Name')
            debit = row.get('Debit (AED)') or row.get('Debit') or 0
            credit = row.get('Credit (AED)') or row.get('Credit') or 0
            
            if not code or (not debit and not credit):
                continue
            
            # Convert code to string
            code = str(int(code)) if isinstance(code, float) else str(code)
            
            # Find account
            account = self.account_map.get(code)
            if not account:
                account = Account.objects.filter(code=code).first()
            if not account:
                account = Account.objects.filter(name__icontains=name).first() if name else None
            
            if not account:
                self.stdout.write(self.style.WARNING(f'  Account not found: {code} - {name}'))
                continue
            
            # Convert amounts
            debit = Decimal(str(debit).replace(',', '')) if debit and debit != '-' else Decimal('0')
            credit = Decimal(str(credit).replace(',', '')) if credit and credit != '-' else Decimal('0')
            
            if debit > 0 or credit > 0:
                JournalEntryLine.objects.create(
                    journal_entry=journal,
                    account=account,
                    description=f'Opening Balance - {account.name}',
                    debit=debit,
                    credit=credit,
                )
                line_count += 1
        
        journal.calculate_totals()
        
        # Post if balanced
        if journal.total_debit == journal.total_credit and journal.total_debit > 0:
            journal.post(self.admin_user)
            self.stdout.write(self.style.SUCCESS(f'  Created {line_count} opening balance lines, Posted: AED {journal.total_debit:,.2f}'))
        else:
            self.stdout.write(self.style.WARNING(f'  Created {line_count} lines, Unbalanced: Dr={journal.total_debit} Cr={journal.total_credit}'))

    def import_sales_invoices(self):
        """Import Sales Invoices from Sheet7."""
        self.stdout.write('Importing Sales Invoices...')
        from apps.sales.models import Invoice, InvoiceItem
        from apps.crm.models import Customer
        
        data = self.get_sheet_data('Sheet7')
        
        created_count = 0
        
        for row in data:
            inv_no = row.get('Invoice No')
            customer_code = row.get('Customer Code')
            customer_name = row.get('Customer Name')
            
            if not inv_no or not customer_code:
                continue
            
            # Find customer
            customer = self.customer_map.get(customer_code)
            if not customer:
                customer = Customer.objects.filter(name__icontains=customer_name).first() if customer_name else None
            if not customer:
                customer = Customer.objects.first()
            
            if not customer:
                self.stdout.write(self.style.WARNING(f'  No customer found for invoice {inv_no}'))
                continue
            
            # Parse date
            inv_date = row.get('Date')
            if isinstance(inv_date, str):
                try:
                    inv_date = datetime.strptime(inv_date, '%Y-%m-%d').date()
                except:
                    inv_date = date.today()
            elif isinstance(inv_date, datetime):
                inv_date = inv_date.date()
            else:
                inv_date = date.today()
            
            # Parse amounts
            amount = row.get('Amount (AED)', 0)
            if isinstance(amount, str):
                amount = Decimal(amount.replace(',', '')) if amount else Decimal('0')
            else:
                amount = Decimal(str(amount or 0))
            
            vat_rate = row.get('VAT Rate', 5)
            if isinstance(vat_rate, str):
                if vat_rate.lower() in ['exempt', 'zero', 'n/a', '']:
                    vat_rate = Decimal('0')
                else:
                    try:
                        vat_rate = Decimal(vat_rate.replace('%', '').strip())
                    except:
                        vat_rate = Decimal('5')
            else:
                vat_rate = Decimal(str(vat_rate or 0))
            # Convert to percentage if needed (0.05 -> 5)
            if vat_rate < 1 and vat_rate > 0:
                vat_rate = vat_rate * 100
            
            vat_amount = row.get('VAT Amount', 0)
            if isinstance(vat_amount, str):
                vat_amount = Decimal(vat_amount.replace(',', '')) if vat_amount else Decimal('0')
            else:
                vat_amount = Decimal(str(vat_amount or 0))
            
            total = row.get('Total (AED)', 0)
            if isinstance(total, str):
                total = Decimal(total.replace(',', '')) if total else Decimal('0')
            else:
                total = Decimal(str(total or 0))
            
            # Determine status
            status_text = row.get('Status', 'Draft').lower() if row.get('Status') else 'draft'
            if 'paid' in status_text:
                status = 'paid'
            elif 'post' in status_text:
                status = 'posted'
            else:
                status = 'draft'
            
            # Create invoice
            invoice = Invoice.objects.create(
                customer=customer,
                invoice_date=inv_date,
                due_date=inv_date,  # Same as invoice date for simplicity
                status=status,
                subtotal=amount,
                vat_amount=vat_amount,
                total_amount=total if total else (amount + vat_amount),
                paid_amount=total if status == 'paid' else Decimal('0'),
                notes=f"Imported from Excel - Original: {inv_no}",
            )
            
            # Create invoice item
            description = row.get('Product/Service', 'Product/Service')
            InvoiceItem.objects.create(
                invoice=invoice,
                description=description,
                quantity=Decimal('1'),
                unit_price=amount,
                vat_rate=vat_rate,
                total=amount,
                vat_amount=vat_amount,
            )
            
            created_count += 1
        
        self.stdout.write(self.style.SUCCESS(f'  Created: {created_count} invoices'))

    def import_purchase_bills(self):
        """Import Purchase Bills from Sheet8."""
        self.stdout.write('Importing Purchase Bills...')
        from apps.purchase.models import Vendor, VendorBill, VendorBillItem
        
        data = self.get_sheet_data('Sheet8')
        
        created_count = 0
        
        for row in data:
            bill_no = row.get('Invoice No')
            vendor_code = row.get('Vendor Code')
            vendor_name = row.get('Vendor Name')
            
            if not bill_no or not vendor_code:
                continue
            
            # Find vendor
            vendor = self.vendor_map.get(vendor_code)
            if not vendor:
                vendor = Vendor.objects.filter(name__icontains=vendor_name).first() if vendor_name else None
            if not vendor:
                vendor = Vendor.objects.first()
            
            if not vendor:
                self.stdout.write(self.style.WARNING(f'  No vendor found for bill {bill_no}'))
                continue
            
            # Parse date
            bill_date = row.get('Date')
            if isinstance(bill_date, str):
                try:
                    bill_date = datetime.strptime(bill_date, '%Y-%m-%d').date()
                except:
                    bill_date = date.today()
            elif isinstance(bill_date, datetime):
                bill_date = bill_date.date()
            else:
                bill_date = date.today()
            
            # Parse amounts
            amount = row.get('Amount (AED)', 0)
            if isinstance(amount, str):
                amount = Decimal(amount.replace(',', '')) if amount else Decimal('0')
            else:
                amount = Decimal(str(amount or 0))
            
            vat_rate = row.get('VAT Rate', 5)
            if isinstance(vat_rate, str):
                if vat_rate.lower() in ['exempt', 'zero', 'n/a', '']:
                    vat_rate = Decimal('0')
                else:
                    try:
                        vat_rate = Decimal(vat_rate.replace('%', '').strip())
                    except:
                        vat_rate = Decimal('5')
            else:
                vat_rate = Decimal(str(vat_rate or 0))
            # Convert to percentage if needed (0.05 -> 5)
            if vat_rate < 1 and vat_rate > 0:
                vat_rate = vat_rate * 100
            
            vat_amount = row.get('VAT Amount', 0)
            if isinstance(vat_amount, str):
                vat_amount = Decimal(vat_amount.replace(',', '')) if vat_amount else Decimal('0')
            else:
                vat_amount = Decimal(str(vat_amount or 0))
            
            total = row.get('Total (AED)', 0)
            if isinstance(total, str):
                total = Decimal(total.replace(',', '')) if total else Decimal('0')
            else:
                total = Decimal(str(total or 0))
            
            # Determine status
            status_text = row.get('Status', 'Draft').lower() if row.get('Status') else 'draft'
            if 'paid' in status_text:
                status = 'paid'
            elif 'post' in status_text:
                status = 'posted'
            else:
                status = 'draft'
            
            # Create bill
            bill = VendorBill.objects.create(
                vendor=vendor,
                vendor_invoice_number=bill_no,
                bill_date=bill_date,
                due_date=bill_date,
                status=status,
                subtotal=amount,
                vat_amount=vat_amount,
                total_amount=total if total else (amount + vat_amount),
                paid_amount=total if status == 'paid' else Decimal('0'),
                notes=f"Imported from Excel - Original: {bill_no}",
            )
            
            # Create bill item
            description = row.get('Product/Service', 'Product/Service')
            VendorBillItem.objects.create(
                bill=bill,
                description=description,
                quantity=Decimal('1'),
                unit_price=amount,
                vat_rate=vat_rate,
                total=amount,
                vat_amount=vat_amount,
            )
            
            created_count += 1
        
        self.stdout.write(self.style.SUCCESS(f'  Created: {created_count} bills'))

    def import_expense_claims(self):
        """Import Expense Claims from Sheet9."""
        self.stdout.write('Importing Expense Claims...')
        from apps.purchase.models import ExpenseClaim, ExpenseClaimItem
        from apps.hr.models import Employee
        
        data = self.get_sheet_data('Sheet9')
        
        created_count = 0
        
        for row in data:
            claim_id = row.get('Claim ID')
            emp_code = row.get('Employee Code')
            emp_name = row.get('Employee Name')
            
            if not claim_id:
                continue
            
            # Find employee/user
            employee = self.employee_map.get(emp_code)
            if not employee:
                employee = Employee.objects.filter(first_name__icontains=emp_name.split()[0]).first() if emp_name else None
            
            # Use admin user if no employee found
            claim_user = employee.user if employee and employee.user else self.admin_user
            if not claim_user:
                claim_user = User.objects.first()
            
            if not claim_user:
                self.stdout.write(self.style.WARNING(f'  No user found for expense claim {claim_id}'))
                continue
            
            # Parse date
            claim_date = row.get('Date Submitted')
            if isinstance(claim_date, str):
                try:
                    claim_date = datetime.strptime(claim_date, '%Y-%m-%d').date()
                except:
                    claim_date = date.today()
            elif isinstance(claim_date, datetime):
                claim_date = claim_date.date()
            else:
                claim_date = date.today()
            
            # Parse amounts
            amount = row.get('Amount (AED)', 0)
            if isinstance(amount, str):
                amount = Decimal(amount.replace(',', '')) if amount else Decimal('0')
            else:
                amount = Decimal(str(amount or 0))
            
            vat_amount = row.get('VAT Amount', 0)
            if isinstance(vat_amount, str):
                vat_amount = Decimal(vat_amount.replace(',', '')) if vat_amount else Decimal('0')
            else:
                vat_amount = Decimal(str(vat_amount or 0))
            
            total = row.get('Total (AED)', 0)
            if isinstance(total, str):
                total = Decimal(total.replace(',', '')) if total else Decimal('0')
            else:
                total = Decimal(str(total or 0))
            
            # Determine status
            status_text = row.get('Status', 'Draft').lower() if row.get('Status') else 'draft'
            if 'paid' in status_text:
                status = 'paid'
            elif 'approv' in status_text:
                status = 'approved'
            elif 'submit' in status_text:
                status = 'submitted'
            else:
                status = 'draft'
            
            # Create expense claim
            expense_type = row.get('Expense Type', 'General')
            description = row.get('Description', '')
            
            # Map expense type to category
            category_map = {
                'travel': 'travel',
                'meals': 'meals',
                'accommodation': 'accommodation',
                'transport': 'transport',
                'office': 'office',
                'communication': 'communication',
                'it expense': 'office',
                'fuel': 'transport',
                'business': 'other',
            }
            category = category_map.get(expense_type.lower(), 'other') if expense_type else 'other'
            
            claim = ExpenseClaim.objects.create(
                employee=claim_user,
                claim_date=claim_date,
                description=f"{expense_type}: {description}" if description else expense_type,
                status=status,
                total_amount=total if total else (amount + vat_amount),
                total_vat=vat_amount,
            )
            
            # Create claim item
            ExpenseClaimItem.objects.create(
                expense_claim=claim,
                date=claim_date,
                category=category,
                description=description or expense_type,
                amount=amount,
                vat_amount=vat_amount,
                has_receipt=True,
            )
            
            created_count += 1
        
        self.stdout.write(self.style.SUCCESS(f'  Created: {created_count} expense claims'))

    def import_journal_entries(self):
        """Import Journal Entries from Sheets 12-15, 18."""
        self.stdout.write('Importing Journal Entries...')
        from apps.finance.models import Account, JournalEntry, JournalEntryLine
        
        journal_sheets = [
            ('Sheet12', 'Bank Charges'),
            ('Sheet13', 'Depreciation'),
            ('Sheet14', 'Bad Debt'),
            ('Sheet15', 'Reversal'),
            ('Sheet18', 'USD Transaction'),
        ]
        
        total_created = 0
        
        for sheet_name, description in journal_sheets:
            try:
                data = self.get_sheet_data(sheet_name)
                
                if not data:
                    continue
                
                # Group by date for multi-line entries
                entries_by_date = {}
                for row in data:
                    entry_date = row.get('Date')
                    if isinstance(entry_date, datetime):
                        entry_date = entry_date.date()
                    elif isinstance(entry_date, str):
                        try:
                            entry_date = datetime.strptime(entry_date, '%Y-%m-%d').date()
                        except:
                            entry_date = date.today()
                    else:
                        entry_date = date.today()
                    
                    if entry_date not in entries_by_date:
                        entries_by_date[entry_date] = []
                    entries_by_date[entry_date].append(row)
                
                for entry_date, rows in entries_by_date.items():
                    # Create journal entry
                    journal = JournalEntry.objects.create(
                        date=entry_date,
                        reference=f'IMP-{sheet_name}-{entry_date.strftime("%Y%m%d")}',
                        description=f'{description} - Imported from Excel',
                        entry_type='standard',
                        created_by=self.admin_user,
                    )
                    
                    for row in rows:
                        account_name = row.get('Account')
                        
                        # Parse amounts - handle AED and USD columns
                        debit = row.get('Debit (AED)') or row.get('Debit') or 0
                        credit = row.get('Credit (AED)') or row.get('Credit') or 0
                        
                        if not account_name or (debit == '-' and credit == '-'):
                            continue
                        
                        # Find account
                        account = Account.objects.filter(name__icontains=account_name).first()
                        if not account:
                            # Try to find by partial name match
                            name_parts = account_name.split(' - ')
                            for part in name_parts:
                                account = Account.objects.filter(name__icontains=part.strip()).first()
                                if account:
                                    break
                        
                        if not account:
                            self.stdout.write(self.style.WARNING(f'  Account not found: {account_name}'))
                            continue
                        
                        # Convert amounts
                        debit = Decimal(str(debit).replace(',', '')) if debit and debit != '-' else Decimal('0')
                        credit = Decimal(str(credit).replace(',', '')) if credit and credit != '-' else Decimal('0')
                        
                        if debit > 0 or credit > 0:
                            JournalEntryLine.objects.create(
                                journal_entry=journal,
                                account=account,
                                description=f'{description} - {account.name}',
                                debit=debit,
                                credit=credit,
                            )
                    
                    journal.calculate_totals()
                    
                    # Post if balanced
                    if journal.total_debit == journal.total_credit and journal.total_debit > 0:
                        journal.post(self.admin_user)
                        total_created += 1
                    else:
                        self.stdout.write(self.style.WARNING(
                            f'  {description} unbalanced: Dr={journal.total_debit} Cr={journal.total_credit}'
                        ))
                
            except Exception as e:
                self.stdout.write(self.style.WARNING(f'  Error processing {sheet_name}: {e}'))
        
        self.stdout.write(self.style.SUCCESS(f'  Created: {total_created} journal entries'))

